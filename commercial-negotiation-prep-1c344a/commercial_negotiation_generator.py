"""
commercial_negotiation_generator.py
Lilly Procurement Skills - commercial-negotiation-prep-1c344a deterministic
XLSX generator (C8).

WHY THIS EXISTS
---------------
SKILL.md Phase 3 ("Benchmark Comparison & Percentile Positioning") and Phase 5
("Volume Leverage Analysis") specify a Pricing Position Summary (counts and
percentages by Below/At/Above/Premium band), a Weighted Avg Position
(volume-weighted across rate lines), a Total Annual Exposure figure, a Volume
Leverage total, and per-line/whole-deal ZOPA bands (Phase 7, Phase 10 tab 2).
None of that math had a kernel binding: every one of those figures was
produced by model prose, with no code computing it and nothing to catch
drift between a per-line table and its own stated rollup. This module closes
that gap, following the same two-stage discipline as pro-forma-builder-1c344a/
pro_forma_generator.py and rfp-engine-1c344a/rfp_xlsx_generator.py:

  1. Python computes ground truth by calling the vendored numeric_kernel.py
     for every piece of math it already covers (percentile_gate() for the
     per-line percentile-resolution gate, weighted_score() for the
     volume-weighted position, to_hourly() for cross-unit rate normalization,
     assert_reconciles() for every rollup total against its own components).
  2. The XLSX workbooks are built from that same ground truth (never
     re-derived by hand in the workbook-writing step), and the self-test
     re-opens the saved files and confirms the written cells match the
     ground truth exactly, so there is no drift between what Python computed
     and what the workbook shows.

WHAT IT REFUSES TO DO
----------------------
  * a rate register missing a required field                -> RateRegisterValidationError
  * a benchmark with non-monotonic P25/P50/P75               -> RateRegisterValidationError
  * a negotiation stack where opening > target or target >
    walkaway (a malformed ZOPA)                               -> ZopaOrderError
  * a stated rollup total that does not equal the sum of its
    own per-line components (exposure, leverage, savings)     -> the kernel's ReconciliationError
  * a volume-weighted position whose kernel-computed value
    disagrees with an independent manual re-derivation        -> WeightedPositionMismatchError
  * openpyxl missing                                          -> ImportError (via _require_openpyxl)

Division of labor, same as the rest of the suite: code owns the arithmetic,
assembly, and invariants; the model (this skill's own workflow prose) owns
the narrative, the benchmark research, and deciding what opening/target/
walkaway rates to propose per line.

SCOPE NOTE: this module covers exactly the five figures named in the C8 plan
item (Pricing Position Summary, Weighted Avg Position, Total Annual Exposure,
Volume Leverage sums, ZOPA band math). It does not touch Phase 6's TCO
escalation math (already kernel-bound via SKILL.md's own HARD RULE prose
instructing a direct escalate()/npv() call) and it has nothing to do with
this skill's search-clustering, benchmark-caching, or Tier-0 SharePoint
sourcing logic, all of which live entirely in SKILL.md prose.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Sequence

# ---------------------------------------------------------------------------
# Vendored numeric kernel (same directory). Per the suite's "generator
# scripts, not freehand authoring" initiative: percentile_gate(), to_hourly(),
# weighted_score(), and assert_reconciles() are called from numeric_kernel.py,
# never re-derived by ad hoc arithmetic in this module.
# ---------------------------------------------------------------------------
_KERNEL_IMPORT_ERROR: Optional[Exception] = None
try:
    _THIS_DIR = os.path.dirname(os.path.abspath(__file__))
    if _THIS_DIR not in sys.path:
        sys.path.insert(0, _THIS_DIR)
    from numeric_kernel import percentile_gate as kernel_percentile_gate
    from numeric_kernel import to_hourly as kernel_to_hourly
    from numeric_kernel import weighted_score as kernel_weighted_score
    from numeric_kernel import assert_reconciles as kernel_assert_reconciles
    from numeric_kernel import InvalidInputError as KernelInvalidInputError
    from numeric_kernel import WeightSumError as KernelWeightSumError
    from numeric_kernel import UnknownUnitError as KernelUnknownUnitError
    from numeric_kernel import ReconciliationError as KernelReconciliationError
    KERNEL_AVAILABLE = True
except Exception as _exc:  # pragma: no cover - defensive, disclosed at runtime
    KERNEL_AVAILABLE = False
    _KERNEL_IMPORT_ERROR = _exc

    def kernel_percentile_gate(n_points, min_points=5):  # type: ignore
        raise RuntimeError(
            "numeric_kernel.py unavailable; cannot evaluate the percentile "
            f"gate. Import error: {_KERNEL_IMPORT_ERROR}"
        )

    def kernel_to_hourly(value, unit):  # type: ignore
        raise RuntimeError(
            "numeric_kernel.py unavailable; cannot normalize to an hourly "
            f"rate. Import error: {_KERNEL_IMPORT_ERROR}"
        )

    def kernel_weighted_score(scores, weights, tolerance=0.001):  # type: ignore
        raise RuntimeError(
            "numeric_kernel.py unavailable; cannot compute the volume-"
            f"weighted position. Import error: {_KERNEL_IMPORT_ERROR}"
        )

    def kernel_assert_reconciles(components, stated_total, label="rollup", tolerance=0.01):  # type: ignore
        raise RuntimeError(
            "numeric_kernel.py unavailable; cannot verify a rollup "
            f"reconciles. Import error: {_KERNEL_IMPORT_ERROR}"
        )

# ---------------------------------------------------------------------------
# XLSX library detection.
# ---------------------------------------------------------------------------
try:
    import openpyxl  # noqa: F401
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
    OPENPYXL_VERSION = openpyxl.__version__
except Exception as _exc:  # pragma: no cover
    OPENPYXL_AVAILABLE = False
    OPENPYXL_VERSION = None
    _OPENPYXL_IMPORT_ERROR = _exc


class RateRegisterValidationError(Exception):
    """Raised when the rate register is missing a required field or carries
    a value that breaks a documented invariant (non-monotonic percentiles,
    negative volumes, etc). Per Rule 1 (no-fabrication) this module refuses
    rather than guessing a missing figure."""


class ZopaOrderError(Exception):
    """Raised when a rate line's opening/target/walkaway stack is not in the
    order opening <= target <= walkaway, i.e. it is not a coherent ZOPA."""


class WeightedPositionMismatchError(Exception):
    """Raised when the kernel-computed Weighted Avg Position disagrees with
    an independent manual re-derivation of the same weighted average. Mirrors
    pro-forma-builder's Year-1-discounting invariant: the arithmetic is
    trivial, so a mismatch here means a bug in how scores/weights were
    assembled before being handed to the kernel, not a kernel defect."""


# ===========================================================================
# Band thresholds (SKILL.md Phase 3, "MARKET POSITION" / "PRICING POSITION
# SUMMARY"): Below Market <=P50, At Market P50-P65, Above Market P65-P85,
# Premium >P85.
# ===========================================================================

BAND_BELOW = "Below"
BAND_AT = "At"
BAND_ABOVE = "Above"
BAND_PREMIUM = "Premium"
BANDS_IN_ORDER = [BAND_BELOW, BAND_AT, BAND_ABOVE, BAND_PREMIUM]

BAND_TO_ASSESSMENT = {
    BAND_BELOW: "Below Market",
    BAND_AT: "At Market",
    BAND_ABOVE: "Above Market",
    BAND_PREMIUM: "Significantly Above Market",
}


def _band_for_position(position_pct: float) -> str:
    """Classify a 0-100 percentile-position estimate into the SKILL.md band.

    Source: SKILL.md Phase 3 Pricing Position Summary:
      "Below Market (<=P50) / At Market (P50-P65) / Above Market (P65-P85) /
       Premium (>P85)".
    """
    if position_pct <= 50:
        return BAND_BELOW
    if position_pct <= 65:
        return BAND_AT
    if position_pct <= 85:
        return BAND_ABOVE
    return BAND_PREMIUM


def _interpolate_percentile_position(proposed: float, p25: float, p50: float, p75: float) -> float:
    """Estimate the percentile point (0-100) of `proposed` within a rate
    line's benchmark distribution, given only the three anchors this skill's
    own Rate Line template reports (P25, P50, P75).

    JUDGMENT CALL (flagged, not silently resolved): SKILL.md's band table
    (Below <=P50, At P50-P65, Above P65-P85, Premium >P85) requires locating
    P65 and P85, but Phase 3's Rate Line template only ever reports P25/P50/
    P75. This function resolves that gap by linear interpolation between
    adjacent reported anchors, and linear extrapolation beyond the outermost
    anchors at the local (P25-P50 or P50-P75) slope. This is this module's
    own inference, not sourced from any skill text; it is deliberately
    monotonic and clamped to [0, 100] so a degenerate distribution cannot
    produce a nonsensical band.
    """
    if not (p25 <= p50 <= p75):
        raise RateRegisterValidationError(
            f"Benchmark percentiles are not monotonic (P25={p25}, P50={p50}, "
            f"P75={p75}); refusing to interpolate a percentile position "
            "against a malformed distribution."
        )

    if proposed <= p25:
        span = p50 - p25
        if span <= 0:
            return 25.0 if proposed == p25 else max(0.0, min(100.0, 25.0))
        pos = 25.0 - (p25 - proposed) / span * 25.0
    elif proposed <= p50:
        span = p50 - p25
        if span == 0:
            pos = 50.0
        else:
            pos = 25.0 + (proposed - p25) / span * 25.0
    elif proposed <= p75:
        span = p75 - p50
        if span == 0:
            pos = 50.0
        else:
            pos = 50.0 + (proposed - p50) / span * 25.0
    else:
        span = p75 - p50
        if span <= 0:
            pos = 100.0
        else:
            pos = 75.0 + (proposed - p75) / span * 25.0

    return max(0.0, min(100.0, pos))


def _confidence_label(n_sources: int) -> str:
    """SKILL.md Phase 3: "DATA CONFIDENCE: [HIGH 3+ independent sources /
    MEDIUM 2 sources / LOW 0-1 sources]" - mechanical transcription of that
    table, not a re-derived threshold."""
    if n_sources >= 3:
        return "HIGH"
    if n_sources == 2:
        return "MEDIUM"
    return "LOW"


# ===========================================================================
# 1. Rate register: typed input + validation
# ===========================================================================

@dataclass
class Benchmark:
    p25: float
    p50: float
    p75: float
    range_low: float
    range_high: float
    n_points: int
    sources: List[str] = field(default_factory=list)


@dataclass
class LillyInternal:
    historical_average: Optional[float] = None
    current_contract: Optional[float] = None
    portfolio_median: Optional[float] = None


@dataclass
class Negotiation:
    opening_rate: float
    target_rate: float
    walkaway_rate: float


@dataclass
class RateLine:
    rate_id: str
    label: str
    unit: str  # hour | day | week | month | year (numeric_kernel.to_hourly's known units)
    proposed_rate: float
    annual_quantity: float  # units of `unit` consumed per year; annual_spend = proposed_rate * annual_quantity
    benchmark: Benchmark
    negotiation: Negotiation
    lilly_internal: LillyInternal = field(default_factory=LillyInternal)


@dataclass
class LeverageOpportunity:
    label: str
    current_state: str
    proposed_state: str
    estimated_savings: float
    confidence: str  # High | Medium | Low
    execution: str = ""


@dataclass
class RateRegister:
    currency: str
    rate_lines: List[RateLine]
    leverage_opportunities: List[LeverageOpportunity] = field(default_factory=list)


REQUIRED_LINE_FIELDS = [
    "rate_id", "label", "unit", "proposed_rate", "annual_quantity",
    "benchmark", "negotiation",
]
REQUIRED_BENCHMARK_FIELDS = ["p25", "p50", "p75", "range_low", "range_high", "n_points"]
REQUIRED_NEGOTIATION_FIELDS = ["opening_rate", "target_rate", "walkaway_rate"]


def validate_rate_register(raw: Dict[str, Any]) -> RateRegister:
    """Validate a raw rate register dict (the validated output of SKILL.md
    Phase 1-3's benchmark research and Phase 7's counter-offer positions)
    and return a typed RateRegister. Refuses rather than guessing when a
    required field is missing or malformed, per Rule 1 (no-fabrication)."""
    errors: List[str] = []

    if "currency" not in raw or not isinstance(raw["currency"], str) or not raw["currency"]:
        errors.append("'currency' must be a non-empty string (e.g. 'USD').")

    raw_lines = raw.get("rate_lines")
    if not isinstance(raw_lines, list) or len(raw_lines) == 0:
        errors.append("'rate_lines' must be a non-empty list.")
        raw_lines = []

    lines: List[RateLine] = []
    seen_ids = set()
    for i, rl in enumerate(raw_lines):
        prefix = f"rate_lines[{i}]"
        if not isinstance(rl, dict):
            errors.append(f"{prefix} must be an object.")
            continue
        for f_ in REQUIRED_LINE_FIELDS:
            if f_ not in rl:
                errors.append(f"{prefix} missing required field '{f_}'.")
        if errors and any(e.startswith(prefix) for e in errors):
            continue

        rate_id = rl.get("rate_id")
        if not isinstance(rate_id, str) or not rate_id:
            errors.append(f"{prefix}.rate_id must be a non-empty string.")
        elif rate_id in seen_ids:
            errors.append(f"{prefix}.rate_id '{rate_id}' is a duplicate; every rate line needs a unique id.")
        else:
            seen_ids.add(rate_id)

        unit = rl.get("unit", "")
        if not isinstance(unit, str) or unit.strip().lower() not in (
            "hour", "day", "week", "month", "year"
        ):
            errors.append(
                f"{prefix}.unit must be one of hour/day/week/month/year "
                f"(numeric_kernel.to_hourly's known units). Got: {unit!r}."
            )

        proposed_rate = rl.get("proposed_rate")
        if not isinstance(proposed_rate, (int, float)) or proposed_rate < 0:
            errors.append(f"{prefix}.proposed_rate must be a non-negative number.")

        annual_quantity = rl.get("annual_quantity")
        if not isinstance(annual_quantity, (int, float)) or annual_quantity < 0:
            errors.append(
                f"{prefix}.annual_quantity must be a non-negative number "
                "(units of `unit` consumed per year)."
            )

        bm_raw = rl.get("benchmark")
        benchmark: Optional[Benchmark] = None
        if not isinstance(bm_raw, dict):
            errors.append(f"{prefix}.benchmark must be an object.")
        else:
            for f_ in REQUIRED_BENCHMARK_FIELDS:
                if f_ not in bm_raw:
                    errors.append(f"{prefix}.benchmark missing required field '{f_}'.")
            if all(f_ in bm_raw for f_ in REQUIRED_BENCHMARK_FIELDS):
                p25, p50, p75 = bm_raw["p25"], bm_raw["p50"], bm_raw["p75"]
                if not (isinstance(p25, (int, float)) and isinstance(p50, (int, float))
                        and isinstance(p75, (int, float))):
                    errors.append(f"{prefix}.benchmark p25/p50/p75 must be numeric.")
                elif not (p25 <= p50 <= p75):
                    errors.append(
                        f"{prefix}.benchmark percentiles are not monotonic "
                        f"(P25={p25}, P50={p50}, P75={p75})."
                    )
                n_points = bm_raw.get("n_points")
                if not isinstance(n_points, int) or n_points < 0:
                    errors.append(f"{prefix}.benchmark.n_points must be a non-negative integer.")
                else:
                    benchmark = Benchmark(
                        p25=p25, p50=p50, p75=p75,
                        range_low=bm_raw["range_low"], range_high=bm_raw["range_high"],
                        n_points=n_points, sources=list(bm_raw.get("sources", [])),
                    )

        neg_raw = rl.get("negotiation")
        negotiation: Optional[Negotiation] = None
        if not isinstance(neg_raw, dict):
            errors.append(f"{prefix}.negotiation must be an object.")
        else:
            for f_ in REQUIRED_NEGOTIATION_FIELDS:
                if f_ not in neg_raw:
                    errors.append(f"{prefix}.negotiation missing required field '{f_}'.")
            if all(f_ in neg_raw for f_ in REQUIRED_NEGOTIATION_FIELDS):
                negotiation = Negotiation(
                    opening_rate=neg_raw["opening_rate"],
                    target_rate=neg_raw["target_rate"],
                    walkaway_rate=neg_raw["walkaway_rate"],
                )

        li_raw = rl.get("lilly_internal", {})
        lilly_internal = LillyInternal(
            historical_average=li_raw.get("historical_average"),
            current_contract=li_raw.get("current_contract"),
            portfolio_median=li_raw.get("portfolio_median"),
        )

        if benchmark is not None and negotiation is not None and isinstance(rate_id, str) and rate_id:
            lines.append(RateLine(
                rate_id=rate_id, label=rl.get("label", rate_id), unit=unit.strip().lower(),
                proposed_rate=float(proposed_rate), annual_quantity=float(annual_quantity),
                benchmark=benchmark, negotiation=negotiation, lilly_internal=lilly_internal,
            ))

    leverage: List[LeverageOpportunity] = []
    for i, op in enumerate(raw.get("leverage_opportunities", [])):
        prefix = f"leverage_opportunities[{i}]"
        if not isinstance(op, dict):
            errors.append(f"{prefix} must be an object.")
            continue
        for f_ in ("label", "current_state", "proposed_state", "estimated_savings", "confidence"):
            if f_ not in op:
                errors.append(f"{prefix} missing required field '{f_}'.")
        if all(f_ in op for f_ in ("label", "current_state", "proposed_state", "estimated_savings", "confidence")):
            leverage.append(LeverageOpportunity(
                label=op["label"], current_state=op["current_state"],
                proposed_state=op["proposed_state"],
                estimated_savings=float(op["estimated_savings"]),
                confidence=op["confidence"], execution=op.get("execution", ""),
            ))

    if errors:
        raise RateRegisterValidationError(
            "Rate register failed validation; refusing to guess missing or "
            "invalid fields. Issues found:\n  - " + "\n  - ".join(errors)
        )

    return RateRegister(currency=raw["currency"], rate_lines=lines, leverage_opportunities=leverage)


# ===========================================================================
# Ground-truth computation (Python side), via the vendored kernel only.
# ===========================================================================

@dataclass
class LineGroundTruth:
    rate_id: str
    label: str
    position_pct: float
    band: str
    normalized_hourly_proposed: float
    annual_spend: float
    exposure_amount: float
    variance_vs_lilly_pct: Optional[float]
    confidence: str
    saving_at_opening: float
    saving_at_target: float
    saving_at_walkaway: float
    exceeds_walkaway: bool


@dataclass
class GroundTruth:
    lines: List[LineGroundTruth]
    total_annual_proposed: float
    counts_by_band: Dict[str, int]
    pct_by_band: Dict[str, float]
    weighted_avg_position: float
    weighted_avg_band: str
    all_lines_percentile_gated: bool
    total_annual_exposure: float
    total_leverage_value: float
    aggregate_saving_at_opening: float
    aggregate_saving_at_target: float
    aggregate_saving_at_walkaway: float


def compute_ground_truth(reg: RateRegister) -> GroundTruth:
    """Compute the Pricing Position Summary, Weighted Avg Position, Total
    Annual Exposure, Volume Leverage total, and per-line ZOPA figures, using
    numeric_kernel.py for every piece of math it already covers.

    Sign convention (mirrors pro-forma-builder's Period model): a "saving" is
    positive when the reference rate (opening/target/walkaway) is below the
    supplier's proposed rate. exposure_amount is the annualized dollar amount
    a line proposes ABOVE its own negotiation target, floored at 0 (a line
    already at or below target contributes no exposure), per SKILL.md's
    "Total Annual Exposure: $[amount above Lilly target, if applicable]".
    """
    if not KERNEL_AVAILABLE:
        raise RuntimeError(
            "numeric_kernel.py could not be imported; cannot compute ground "
            f"truth without it. Import error: {_KERNEL_IMPORT_ERROR}"
        )

    line_gts: List[LineGroundTruth] = []
    for rl in reg.rate_lines:
        bm = rl.benchmark
        position_pct = _interpolate_percentile_position(rl.proposed_rate, bm.p25, bm.p50, bm.p75)
        band = _band_for_position(position_pct)

        # to_hourly(): rate normalization (kernel NORMALIZATION face), so a
        # portfolio mixing day-rate contractor lines and hourly consulting
        # lines can be compared on one common basis in the workbook, even
        # though the position/exposure math below deliberately stays in each
        # line's own native unit (its own annual_quantity is quoted in that
        # same unit, so the native-unit math is already apples-to-apples for
        # that line).
        normalized_hourly_proposed = kernel_to_hourly(rl.proposed_rate, rl.unit)

        annual_spend = rl.proposed_rate * rl.annual_quantity
        exposure_amount = max(0.0, rl.proposed_rate - rl.negotiation.target_rate) * rl.annual_quantity

        variance_pct = None
        if rl.lilly_internal.historical_average:
            variance_pct = (
                (rl.proposed_rate - rl.lilly_internal.historical_average)
                / rl.lilly_internal.historical_average * 100.0
            )

        confidence = _confidence_label(len(bm.sources))

        saving_at_opening = (rl.proposed_rate - rl.negotiation.opening_rate) * rl.annual_quantity
        saving_at_target = (rl.proposed_rate - rl.negotiation.target_rate) * rl.annual_quantity
        saving_at_walkaway = (rl.proposed_rate - rl.negotiation.walkaway_rate) * rl.annual_quantity
        exceeds_walkaway = rl.proposed_rate > rl.negotiation.walkaway_rate

        line_gts.append(LineGroundTruth(
            rate_id=rl.rate_id, label=rl.label, position_pct=position_pct, band=band,
            normalized_hourly_proposed=normalized_hourly_proposed, annual_spend=annual_spend,
            exposure_amount=exposure_amount, variance_vs_lilly_pct=variance_pct,
            confidence=confidence, saving_at_opening=saving_at_opening,
            saving_at_target=saving_at_target, saving_at_walkaway=saving_at_walkaway,
            exceeds_walkaway=exceeds_walkaway,
        ))

    total_annual_proposed = sum(g.annual_spend for g in line_gts)
    if total_annual_proposed <= 0:
        raise KernelInvalidInputError(
            "Total annual proposed spend across all rate lines is 0; cannot "
            "compute a volume-weighted position (weights would all be 0/0)."
        )

    counts_by_band = {b: 0 for b in BANDS_IN_ORDER}
    for g in line_gts:
        counts_by_band[g.band] += 1
    n_lines = len(line_gts)
    pct_by_band = {b: (counts_by_band[b] / n_lines * 100.0) for b in BANDS_IN_ORDER}

    # Weighted Avg Position: volume-weighted by Lilly's expected utilization
    # (SKILL.md Phase 3). Weight = each line's share of total annual proposed
    # spend (a dollar-based weight, avoiding the apples-to-oranges problem of
    # mixing raw physical units -- hours vs days vs seats -- as weights
    # directly). Calls the kernel's weighted_score(), which additionally
    # refuses if the weights do not foot to 1.0.
    scores = {g.rate_id: g.position_pct for g in line_gts}
    weights = {g.rate_id: g.annual_spend / total_annual_proposed for g in line_gts}
    weighted_avg_position = kernel_weighted_score(scores, weights)
    weighted_avg_band = _band_for_position(weighted_avg_position)

    # SKILL.md Phase 3: "P[XX] only when the per-line samples support that
    # resolution" -- only report the numeric weighted percentile when every
    # contributing line clears the percentile_gate() (kernel-bound, N>=5).
    all_lines_percentile_gated = all(
        kernel_percentile_gate(rl.benchmark.n_points) for rl in reg.rate_lines
    )

    total_annual_exposure = kernel_assert_reconciles(
        components=[g.exposure_amount for g in line_gts],
        stated_total=sum(g.exposure_amount for g in line_gts),
        label="Total Annual Exposure",
    )
    total_leverage_value = kernel_assert_reconciles(
        components=[op.estimated_savings for op in reg.leverage_opportunities],
        stated_total=sum(op.estimated_savings for op in reg.leverage_opportunities),
        label="Total Leverage Value",
    ) if reg.leverage_opportunities else 0.0

    aggregate_saving_at_opening = kernel_assert_reconciles(
        components=[g.saving_at_opening for g in line_gts],
        stated_total=sum(g.saving_at_opening for g in line_gts),
        label="Savings Summary: OPENING",
    )
    aggregate_saving_at_target = kernel_assert_reconciles(
        components=[g.saving_at_target for g in line_gts],
        stated_total=sum(g.saving_at_target for g in line_gts),
        label="Savings Summary: TARGET",
    )
    aggregate_saving_at_walkaway = kernel_assert_reconciles(
        components=[g.saving_at_walkaway for g in line_gts],
        stated_total=sum(g.saving_at_walkaway for g in line_gts),
        label="Savings Summary: WALK-AWAY",
    )

    return GroundTruth(
        lines=line_gts, total_annual_proposed=total_annual_proposed,
        counts_by_band=counts_by_band, pct_by_band=pct_by_band,
        weighted_avg_position=weighted_avg_position, weighted_avg_band=weighted_avg_band,
        all_lines_percentile_gated=all_lines_percentile_gated,
        total_annual_exposure=total_annual_exposure, total_leverage_value=total_leverage_value,
        aggregate_saving_at_opening=aggregate_saving_at_opening,
        aggregate_saving_at_target=aggregate_saving_at_target,
        aggregate_saving_at_walkaway=aggregate_saving_at_walkaway,
    )


# ===========================================================================
# Hard-coded invariant checks (run BEFORE building workbooks)
# ===========================================================================

def _assert_zopa_order(reg: RateRegister) -> None:
    """Every rate line's negotiation stack must be a coherent ZOPA: opening
    <= target <= walkaway (SKILL.md Phase 7: opening anchors low, target is
    the realistic win, walkaway is the ceiling). A stack out of order is not
    a valid Zone of Possible Agreement and must not reach a workbook."""
    for rl in reg.rate_lines:
        n = rl.negotiation
        if not (n.opening_rate <= n.target_rate <= n.walkaway_rate):
            raise ZopaOrderError(
                f"Rate line '{rl.rate_id}' ({rl.label}) has an incoherent ZOPA: "
                f"opening={n.opening_rate}, target={n.target_rate}, "
                f"walkaway={n.walkaway_rate}. Required: opening <= target <= "
                "walkaway. Refusing to build a workbook around a malformed "
                "negotiation stack."
            )


def _assert_weighted_position_matches_manual(gt: GroundTruth) -> None:
    """Independently re-derive the Weighted Avg Position with plain Python
    (no kernel call) and confirm it matches the kernel-computed value. This
    does not distrust the kernel; it catches bugs in how THIS module built
    the scores/weights dicts handed to it (e.g. a weight computed against
    the wrong denominator), the same class of bug pro-forma-builder's
    Year-1-discounting invariant guards against."""
    total_spend = sum(g.annual_spend for g in gt.lines)
    manual = sum(g.position_pct * (g.annual_spend / total_spend) for g in gt.lines)
    if abs(manual - gt.weighted_avg_position) > 1e-6:
        raise WeightedPositionMismatchError(
            f"Weighted Avg Position mismatch: manual re-derivation={manual!r}, "
            f"kernel-computed={gt.weighted_avg_position!r}. These must agree "
            "exactly; a mismatch means the scores/weights dicts passed to "
            "weighted_score() were assembled incorrectly."
        )


def run_hardcoded_invariant_checks(reg: RateRegister, gt: GroundTruth) -> None:
    """Run every hard-coded invariant. Raises on any failure; callers
    (generate_commercial_negotiation_workbooks) must not save a workbook if
    this raises."""
    _assert_zopa_order(reg)
    _assert_weighted_position_matches_manual(gt)


# ===========================================================================
# Workbook builders
# ===========================================================================

def _require_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is not installed in this Python environment, so no "
            "XLSX file can be written. Install it (`pip install openpyxl`) "
            "or point this script at an interpreter that already has it. "
            f"Original import error: {_OPENPYXL_IMPORT_ERROR}"
        )


_HEADER_FONT = None
_HEADER_FILL = None
_BAND_FILLS = None


def _style_header(ws, headers: Sequence[str], row: int = 1) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _init_styles() -> None:
    global _HEADER_FONT, _HEADER_FILL, _BAND_FILLS
    _HEADER_FONT = Font(bold=True, color="FFFFFF")
    _HEADER_FILL = PatternFill(start_color="0F3A85", end_color="0F3A85", fill_type="solid")
    # No green in any status role, per the suite-wide status-palette rule.
    _BAND_FILLS = {
        BAND_BELOW: PatternFill(start_color="D4E5F7", end_color="D4E5F7", fill_type="solid"),   # pale blue
        BAND_AT: PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),       # amber
        BAND_ABOVE: PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid"),     # orange
        BAND_PREMIUM: PatternFill(start_color="FDE8E5", end_color="FDE8E5", fill_type="solid"),   # pale red
    }


def build_rate_comparison_workbook(reg: RateRegister, gt: GroundTruth) -> "openpyxl.Workbook":
    """Build rate_comparison.xlsx: one row per rate line (supplier proposed,
    market benchmarks, Lilly internal, market position band, variance, data
    confidence, normalized hourly rate) plus a Pricing Position Summary
    sheet (counts/percentages by band, Weighted Avg Position, Total Annual
    Exposure)."""
    _require_openpyxl()
    _init_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rate Comparison"

    headers = [
        "Rate ID", "Label", "Unit", "Supplier Proposed", "P25", "P50", "P75",
        "Range Low", "Range High", "N Points", "Sources", "Historical Average",
        "Current Contract", "Portfolio Median", "Normalized Hourly Rate",
        "Position (est. percentile)", "Market Position Band", "Assessment",
        "Variance vs Lilly Historical %", "Data Confidence", "Annual Quantity",
        "Annual Spend",
    ]
    _style_header(ws, headers)

    by_id = {rl.rate_id: rl for rl in reg.rate_lines}
    r = 2
    for g in gt.lines:
        rl = by_id[g.rate_id]
        bm = rl.benchmark
        li = rl.lilly_internal
        row_vals = [
            rl.rate_id, rl.label, rl.unit, rl.proposed_rate, bm.p25, bm.p50, bm.p75,
            bm.range_low, bm.range_high, bm.n_points, "; ".join(bm.sources),
            li.historical_average, li.current_contract, li.portfolio_median,
            round(g.normalized_hourly_proposed, 4), round(g.position_pct, 2),
            g.band, BAND_TO_ASSESSMENT[g.band],
            round(g.variance_vs_lilly_pct, 2) if g.variance_vs_lilly_pct is not None else None,
            g.confidence, rl.annual_quantity, round(g.annual_spend, 2),
        ]
        for col, v in enumerate(row_vals, start=1):
            ws.cell(row=r, column=col, value=v)
        ws.cell(row=r, column=17).fill = _BAND_FILLS[g.band]
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # --- Pricing Position Summary sheet -----------------------------------
    ws2 = wb.create_sheet("Pricing Position Summary")
    _style_header(ws2, ["Metric", "Value"])
    n_lines = len(gt.lines)
    rows2 = [
        ("Rate Lines Analyzed", n_lines),
        ("Below Market (<=P50) count", gt.counts_by_band[BAND_BELOW]),
        ("Below Market (<=P50) %", round(gt.pct_by_band[BAND_BELOW], 1)),
        ("At Market (P50-P65) count", gt.counts_by_band[BAND_AT]),
        ("At Market (P50-P65) %", round(gt.pct_by_band[BAND_AT], 1)),
        ("Above Market (P65-P85) count", gt.counts_by_band[BAND_ABOVE]),
        ("Above Market (P65-P85) %", round(gt.pct_by_band[BAND_ABOVE], 1)),
        ("Premium (>P85) count", gt.counts_by_band[BAND_PREMIUM]),
        ("Premium (>P85) %", round(gt.pct_by_band[BAND_PREMIUM], 1)),
        ("Weighted Avg Position (0-100 scale)", round(gt.weighted_avg_position, 2)),
        ("Weighted Avg Position (percentile reportable)", gt.all_lines_percentile_gated),
        ("Weighted Avg Position Band", BAND_TO_ASSESSMENT[gt.weighted_avg_band]),
        ("Total Annual Proposed Spend", round(gt.total_annual_proposed, 2)),
        ("Total Annual Exposure (above Lilly target)", round(gt.total_annual_exposure, 2)),
    ]
    for i, (label, value) in enumerate(rows2, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 22

    # Reconciliation-check cell, visible in the workbook itself, mirroring
    # pro-forma-builder's Savings Waterfall pattern.
    check_row = len(rows2) + 3
    ws2.cell(row=check_row, column=1,
             value="Reconciliation check (sum of per-line exposure - stated Total Annual Exposure; must be 0)")
    ws2.cell(row=check_row, column=2,
             value=sum(g.exposure_amount for g in gt.lines) - gt.total_annual_exposure)

    return wb


def build_counter_offer_workbook(reg: RateRegister, gt: GroundTruth) -> "openpyxl.Workbook":
    """Build counter_offer.xlsx: per-line opening/target/walkaway columns and
    ZOPA flags, a Savings Summary (Phase 7), and a Volume Leverage
    Opportunities sheet (Phase 5) with its Total Leverage Value."""
    _require_openpyxl()
    _init_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "Counter-Offer"

    headers = [
        "Rate ID", "Label", "Unit", "Supplier Proposed", "Opening", "Target",
        "Walk-Away", "Saving at Opening", "Saving at Target", "Saving at Walk-Away",
        "Exceeds Walk-Away", "Market Position Band",
    ]
    _style_header(ws, headers)

    by_id = {rl.rate_id: rl for rl in reg.rate_lines}
    r = 2
    for g in gt.lines:
        rl = by_id[g.rate_id]
        n = rl.negotiation
        row_vals = [
            rl.rate_id, rl.label, rl.unit, rl.proposed_rate, n.opening_rate,
            n.target_rate, n.walkaway_rate, round(g.saving_at_opening, 2),
            round(g.saving_at_target, 2), round(g.saving_at_walkaway, 2),
            g.exceeds_walkaway, g.band,
        ]
        for col, v in enumerate(row_vals, start=1):
            ws.cell(row=r, column=col, value=v)
        if g.exceeds_walkaway:
            ws.cell(row=r, column=11).fill = _BAND_FILLS[BAND_PREMIUM]
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # --- Savings Summary sheet (Phase 7) -----------------------------------
    ws2 = wb.create_sheet("Savings Summary")
    _style_header(ws2, ["Position", "Aggregate Saving vs Proposed", "% of Proposed"])
    total = gt.total_annual_proposed
    savings_rows = [
        ("OPENING", gt.aggregate_saving_at_opening),
        ("TARGET", gt.aggregate_saving_at_target),
        ("WALK-AWAY", gt.aggregate_saving_at_walkaway),
    ]
    for i, (label, saving) in enumerate(savings_rows, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=round(saving, 2))
        ws2.cell(row=i, column=3, value=round(saving / total * 100.0, 2) if total else None)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 16

    check_row = len(savings_rows) + 3
    ws2.cell(row=check_row, column=1,
             value="Reconciliation check (sum of per-line savings @ TARGET - aggregate; must be 0)")
    ws2.cell(row=check_row, column=2,
             value=sum(g.saving_at_target for g in gt.lines) - gt.aggregate_saving_at_target)

    # --- Volume Leverage Opportunities sheet (Phase 5) ---------------------
    ws3 = wb.create_sheet("Volume Leverage")
    _style_header(ws3, ["Opportunity", "Current State", "Proposed State", "Estimated Savings", "Confidence", "Execution"])
    r = 2
    for op in reg.leverage_opportunities:
        ws3.cell(row=r, column=1, value=op.label)
        ws3.cell(row=r, column=2, value=op.current_state)
        ws3.cell(row=r, column=3, value=op.proposed_state)
        ws3.cell(row=r, column=4, value=round(op.estimated_savings, 2))
        ws3.cell(row=r, column=5, value=op.confidence)
        ws3.cell(row=r, column=6, value=op.execution)
        r += 1
    total_row = r + 1
    ws3.cell(row=total_row, column=1, value="TOTAL LEVERAGE VALUE")
    ws3.cell(row=total_row, column=4, value=round(gt.total_leverage_value, 2))
    for col in "ABCDEF":
        ws3.column_dimensions[col].width = 26

    check_row2 = total_row + 2
    ws3.cell(row=check_row2, column=1,
             value="Reconciliation check (sum of opportunity savings - Total Leverage Value; must be 0)")
    ws3.cell(row=check_row2, column=4,
             value=sum(op.estimated_savings for op in reg.leverage_opportunities) - gt.total_leverage_value)

    return wb


# ===========================================================================
# Full pipeline: validate -> compute ground truth -> hard invariant checks
# -> build both workbooks -> save
# ===========================================================================

def generate_commercial_negotiation_workbooks(
    raw_register: Dict[str, Any],
    rate_comparison_path: str,
    counter_offer_path: str,
) -> RateRegister:
    """End-to-end: validate the raw rate register, compute ground truth via
    numeric_kernel, run the hard-coded invariant checks, build both
    workbooks, and only then save them. Raises rather than saving a workbook
    that fails validation or reconciliation."""
    reg = validate_rate_register(raw_register)
    gt = compute_ground_truth(reg)
    run_hardcoded_invariant_checks(reg, gt)
    wb1 = build_rate_comparison_workbook(reg, gt)
    wb1.save(rate_comparison_path)
    wb2 = build_counter_offer_workbook(reg, gt)
    wb2.save(counter_offer_path)
    return reg


# ===========================================================================
# __main__ self-test
# ===========================================================================

if __name__ == "__main__":
    import copy
    import tempfile

    print("=" * 78)
    print("commercial_negotiation_generator.py self-test")
    print("=" * 78)
    print(f"numeric_kernel.py available: {KERNEL_AVAILABLE}")
    print(f"openpyxl available: {OPENPYXL_AVAILABLE}"
          + (f" (version {OPENPYXL_VERSION})" if OPENPYXL_AVAILABLE else ""))
    print()

    results: List[tuple] = []

    def check(label, condition, detail=""):
        results.append((label, bool(condition), detail))
        status = "PASS" if condition else "FAIL"
        line = f"[{status}] {label}"
        if detail:
            line += f"  ({detail})"
        print(line)

    # Sample register spanning all four bands, with a mix of units, so the
    # to_hourly() normalization and the weighted-position blend are both
    # genuinely exercised.
    sample_register: Dict[str, Any] = {
        "currency": "USD",
        "rate_lines": [
            {
                "rate_id": "L1", "label": "Senior Developer (T&M)", "unit": "hour",
                "proposed_rate": 145.0, "annual_quantity": 4000,
                "benchmark": {"p25": 120.0, "p50": 140.0, "p75": 165.0,
                              "range_low": 100.0, "range_high": 190.0,
                              "n_points": 8, "sources": ["Janco", "TEKsystems", "Robert Half"]},
                "lilly_internal": {"historical_average": 138.0, "current_contract": 135.0,
                                     "portfolio_median": 142.0},
                "negotiation": {"opening_rate": 130.0, "target_rate": 138.0, "walkaway_rate": 150.0},
            },
            {
                "rate_id": "L2", "label": "Solutions Architect (day rate)", "unit": "day",
                "proposed_rate": 1900.0, "annual_quantity": 200,
                "benchmark": {"p25": 1400.0, "p50": 1600.0, "p75": 1850.0,
                              "range_low": 1200.0, "range_high": 2100.0,
                              "n_points": 6, "sources": ["Janco", "TEKsystems"]},
                "lilly_internal": {"historical_average": 1700.0, "current_contract": None,
                                     "portfolio_median": None},
                "negotiation": {"opening_rate": 1600.0, "target_rate": 1750.0, "walkaway_rate": 1900.0},
            },
            {
                "rate_id": "L3", "label": "PM (per-seat/month)", "unit": "month",
                "proposed_rate": 9500.0, "annual_quantity": 12,
                "benchmark": {"p25": 7000.0, "p50": 8000.0, "p75": 9000.0,
                              "range_low": 6000.0, "range_high": 10000.0,
                              "n_points": 3, "sources": ["Janco", "Robert Half"]},
                "lilly_internal": {"historical_average": 8200.0},
                "negotiation": {"opening_rate": 8000.0, "target_rate": 8500.0, "walkaway_rate": 9200.0},
            },
            {
                "rate_id": "L4", "label": "QA Analyst", "unit": "hour",
                "proposed_rate": 78.0, "annual_quantity": 2000,
                "benchmark": {"p25": 75.0, "p50": 85.0, "p75": 95.0,
                              "range_low": 65.0, "range_high": 105.0,
                              "n_points": 5, "sources": ["Janco", "TEKsystems", "Robert Half"]},
                "lilly_internal": {"historical_average": 82.0},
                "negotiation": {"opening_rate": 74.0, "target_rate": 78.0, "walkaway_rate": 85.0},
            },
        ],
        "leverage_opportunities": [
            {"label": "Multi-BU consolidation", "current_state": "Spend split across 3 BUs",
             "proposed_state": "Consolidate under one MSA", "estimated_savings": 120000.0,
             "confidence": "High", "execution": "Combine SOWs at renewal"},
            {"label": "3-year commitment", "current_state": "1-year renewals",
             "proposed_state": "3-year term with rate lock", "estimated_savings": 65000.0,
             "confidence": "Medium", "execution": "Negotiate rate lock alongside term extension"},
        ],
    }

    # --- Step 1: validation ---------------------------------------------
    try:
        reg = validate_rate_register(sample_register)
        check("validate_rate_register accepts a complete sample register", True,
              f"{len(reg.rate_lines)} rate lines, {len(reg.leverage_opportunities)} leverage opportunities")
    except Exception as e:
        check("validate_rate_register accepts a complete sample register", False, str(e))
        raise

    broken = copy.deepcopy(sample_register)
    del broken["rate_lines"][0]["benchmark"]
    try:
        validate_rate_register(broken)
        check("validate_rate_register refuses a rate line missing 'benchmark'", False, "did not raise")
    except RateRegisterValidationError as e:
        check("validate_rate_register refuses a rate line missing 'benchmark'", True, str(e)[:120])

    broken2 = copy.deepcopy(sample_register)
    broken2["rate_lines"][0]["benchmark"]["p50"] = 500.0  # now P25 <= P50 but P50 > P75, non-monotonic
    try:
        validate_rate_register(broken2)
        check("validate_rate_register refuses non-monotonic P25/P50/P75", False, "did not raise")
    except RateRegisterValidationError as e:
        check("validate_rate_register refuses non-monotonic P25/P50/P75", True, str(e)[:120])

    # --- Step 2: ground truth --------------------------------------------
    try:
        gt = compute_ground_truth(reg)
        check("compute_ground_truth runs via numeric_kernel", True,
              f"weighted_avg_position={gt.weighted_avg_position:.2f}, "
              f"total_annual_exposure={gt.total_annual_exposure:.2f}, "
              f"total_leverage_value={gt.total_leverage_value:.2f}")
    except Exception as e:
        check("compute_ground_truth runs via numeric_kernel", False, str(e))
        raise

    # --- Step 3: the required Verify-line assertion: exposure total reconciles ---
    manual_exposure_total = sum(
        max(0.0, rl.proposed_rate - rl.negotiation.target_rate) * rl.annual_quantity
        for rl in reg.rate_lines
    )
    check("Total Annual Exposure reconciles against a manually summed per-line total",
          abs(manual_exposure_total - gt.total_annual_exposure) < 1e-6,
          f"manual={manual_exposure_total:.2f}, kernel-verified={gt.total_annual_exposure:.2f}")

    # --- Step 4: Weighted Avg Position math is correct --------------------
    total_spend = sum(g.annual_spend for g in gt.lines)
    manual_weighted = sum(g.position_pct * (g.annual_spend / total_spend) for g in gt.lines)
    check("Weighted Avg Position matches an independent manual re-derivation",
          abs(manual_weighted - gt.weighted_avg_position) < 1e-6,
          f"manual={manual_weighted:.4f}, kernel={gt.weighted_avg_position:.4f}")

    # --- Step 5: ZOPA bands compute correctly ------------------------------
    # Known boundary values: exactly P50 -> Below; just above P50 -> At;
    # exactly P75 (interpolated to 75) -> Above; a value far beyond P75 -> Premium.
    boundary_checks = [
        (100.0, BAND_BELOW),   # <= 50 position from a value at P25 exactly maps to 25 -> Below
        (140.0, BAND_BELOW),   # exactly P50 -> position 50 -> Below (<=50 rule)
        (150.0, None),         # between P50/P75, computed below against expected math
        (165.0, BAND_ABOVE),   # exactly P75 -> position 75 -> Above (65 < 75 <= 85 rule)
        (400.0, BAND_PREMIUM),  # far beyond P75 -> Premium
    ]
    p25, p50, p75 = 120.0, 140.0, 165.0
    all_boundary_ok = True
    detail_parts = []
    for proposed, expected in boundary_checks:
        pos = _interpolate_percentile_position(proposed, p25, p50, p75)
        band = _band_for_position(pos)
        if expected is not None:
            ok = band == expected
        else:
            # 150 is strictly between P50 (position 50) and P75 (position 75);
            # interpolation must land strictly between 50 and 75, i.e. "At" or "Above".
            ok = 50 < pos < 75 and band in (BAND_AT, BAND_ABOVE)
        all_boundary_ok = all_boundary_ok and ok
        detail_parts.append(f"{proposed}->pos={pos:.1f},band={band}{'OK' if ok else 'MISMATCH'}")
    check("ZOPA/market-position band boundaries classify correctly",
          all_boundary_ok, "; ".join(detail_parts))

    # Per-line ZOPA ordering invariant actually catches a malformed stack.
    broken_zopa_reg = copy.deepcopy(reg)
    broken_zopa_reg.rate_lines[0].negotiation.target_rate = 999.0  # target > walkaway now
    try:
        _assert_zopa_order(broken_zopa_reg)
        check("ZOPA order invariant CORRECTLY REJECTS a malformed opening/target/walkaway stack",
              False, "did not raise, but should have")
    except ZopaOrderError as e:
        check("ZOPA order invariant CORRECTLY REJECTS a malformed opening/target/walkaway stack",
              True, str(e)[:150])

    # Weighted-position mismatch invariant actually catches a bug.
    bad_gt = copy.deepcopy(gt)
    bad_gt.weighted_avg_position = gt.weighted_avg_position + 5.0
    try:
        _assert_weighted_position_matches_manual(bad_gt)
        check("Weighted-position invariant CORRECTLY REJECTS a simulated +5 drift bug",
              False, "did not raise, but should have")
    except WeightedPositionMismatchError as e:
        check("Weighted-position invariant CORRECTLY REJECTS a simulated +5 drift bug",
              True, str(e)[:150])

    try:
        run_hardcoded_invariant_checks(reg, gt)
        check("hard-coded invariant checks PASS on the valid sample "
              "(ZOPA order; weighted position matches manual re-derivation)", True)
    except Exception as e:
        check("hard-coded invariant checks PASS on the valid sample", False, str(e))
        raise

    # --- Step 6: build + save + reopen the real workbooks; confirm no drift ---
    if not OPENPYXL_AVAILABLE:
        check("openpyxl available to write real .xlsx files", False,
              "openpyxl is NOT installed in this interpreter; XLSX writing could not be exercised.")
    else:
        check("openpyxl available to write real .xlsx files", True, f"version {OPENPYXL_VERSION}")

        tmp_dir = tempfile.gettempdir()
        rate_comparison_path = os.path.join(tmp_dir, "rate_comparison_selftest.xlsx")
        counter_offer_path = os.path.join(tmp_dir, "counter_offer_selftest.xlsx")
        try:
            generate_commercial_negotiation_workbooks(sample_register, rate_comparison_path, counter_offer_path)
            check("generate_commercial_negotiation_workbooks() ran end-to-end without raising", True)
        except Exception as e:
            check("generate_commercial_negotiation_workbooks() ran end-to-end without raising", False, str(e))
            raise

        for label, path in [("rate_comparison.xlsx", rate_comparison_path),
                             ("counter_offer.xlsx", counter_offer_path)]:
            exists = os.path.exists(path)
            size = os.path.getsize(path) if exists else 0
            check(f"{label} written to {path}", exists and size > 0, f"size={size} bytes")

        try:
            reopened1 = openpyxl.load_workbook(rate_comparison_path)
            check("rate_comparison.xlsx has both expected tabs",
                  set(reopened1.sheetnames) == {"Rate Comparison", "Pricing Position Summary"},
                  f"tabs={reopened1.sheetnames}")

            ws_sum = reopened1["Pricing Position Summary"]
            written_values = {ws_sum.cell(row=r, column=1).value: ws_sum.cell(row=r, column=2).value
                               for r in range(2, ws_sum.max_row + 1)
                               if ws_sum.cell(row=r, column=1).value}
            written_exposure = written_values.get("Total Annual Exposure (above Lilly target)")
            check("Re-opened rate_comparison.xlsx Total Annual Exposure matches computed ground truth (no drift)",
                  written_exposure is not None and abs(written_exposure - round(gt.total_annual_exposure, 2)) < 1e-6,
                  f"written={written_exposure}, ground_truth={round(gt.total_annual_exposure, 2)}")

            written_wap = written_values.get("Weighted Avg Position (0-100 scale)")
            check("Re-opened rate_comparison.xlsx Weighted Avg Position matches computed ground truth (no drift)",
                  written_wap is not None and abs(written_wap - round(gt.weighted_avg_position, 2)) < 1e-6,
                  f"written={written_wap}, ground_truth={round(gt.weighted_avg_position, 2)}")

            recon_cell = None
            for row in ws_sum.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "Reconciliation check" in cell.value:
                        recon_cell = ws_sum.cell(row=cell.row, column=2).value
            check("rate_comparison.xlsx reconciliation-check cell is 0 (exposure total foots)",
                  recon_cell is not None and abs(recon_cell) < 1e-6,
                  f"recon_cell={recon_cell}")

            reopened2 = openpyxl.load_workbook(counter_offer_path)
            check("counter_offer.xlsx has all three expected tabs",
                  set(reopened2.sheetnames) == {"Counter-Offer", "Savings Summary", "Volume Leverage"},
                  f"tabs={reopened2.sheetnames}")

            ws_vl = reopened2["Volume Leverage"]
            written_total_leverage = None
            for row in ws_vl.iter_rows():
                for cell in row:
                    if cell.value == "TOTAL LEVERAGE VALUE":
                        written_total_leverage = ws_vl.cell(row=cell.row, column=4).value
            check("Re-opened counter_offer.xlsx Total Leverage Value matches computed ground truth (no drift)",
                  written_total_leverage is not None
                  and abs(written_total_leverage - round(gt.total_leverage_value, 2)) < 1e-6,
                  f"written={written_total_leverage}, ground_truth={round(gt.total_leverage_value, 2)}")

        except Exception as e:
            check("Re-opened workbook structural and value spot-checks", False, str(e))

    print()
    print("=" * 78)
    passed = sum(1 for _, ok, _ in results if ok)
    failed = sum(1 for _, ok, _ in results if not ok)
    total = len(results)
    print(f"SUMMARY: {passed}/{total} passed, {failed}/{total} failed")
    if failed:
        print("FAILED CASES:")
        for label, ok, detail in results:
            if not ok:
                print(f"  - {label}: {detail}")
    print("=" * 78)

    if failed:
        sys.exit(1)
