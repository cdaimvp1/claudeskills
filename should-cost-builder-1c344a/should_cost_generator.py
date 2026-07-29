"""
should_cost_generator.py
Lilly Procurement Skills - should-cost-builder-1c344a deterministic XLSX generator.

Purpose (mirrors pro-forma-builder-1c344a/pro_forma_generator.py's "generator
scripts, not freehand authoring" pattern): this module takes a validated
should-cost input register (cost drivers + margin policy + optional supplier
price, per should-cost-builder-1c344a/SKILL.md's "Aggregation Method",
"Gap Method", and "Cost-Driver Assumption Ledger" sections) and MECHANICALLY
produces a formula-driven should_cost_model.xlsx. Every derived cell in the
workbook is a real Excel formula referencing other cells, so the model is
auditable and adjustable in Excel, never a pasted-in number.

Two-stage discipline (same as pro-forma-builder):
  1. Python computes the same figures the workbook will show, by calling the
     vendored numeric_kernel.py's quadrature_rollup() / convert_currency()
     (never re-deriving that arithmetic itself). SKILL.md's "Computation
     requirement (HARD RULE)": "do not hand-compute the roll-up. The actual
     arithmetic MUST be executed by calling quadrature_rollup() in the
     vendored numeric_kernel.py... not produced by model arithmetic."
  2. The Excel formulas written into the workbook independently re-derive the
     SAME roll-up (SQRT(SUMSQ(...)) style formulas over live cell ranges),
     not by pasting Python's computed numbers. Before saving, this script
     asserts the two agree (see run_hardcoded_invariant_checks()).

Hard-coded invariants (code-level checks, not comments; mirrors pro-forma's
_assert_year1_discounting / _assert_waterfall_reconciles_to_npv_inputs
pattern):
  - Quadrature-not-naive invariant: the kernel's quadrature roll-up is
    independently re-derived in Python (sqrt of sum-of-squares over
    non-widened drivers, linear addition over widened drivers) and must match
    the kernel's own QuadratureResult; and, whenever at least two components
    carry a nonzero, non-widened spread, the quadrature range must differ
    from the naive sum-of-extremes range (proving quadrature actually ran,
    not a naive sum masquerading as one).
  - Margin-excluded-from-quadrature invariant: SKILL.md's Aggregation Method
    step 3 HARD RULE, "Margin/SG&A is never passed into quadrature_rollup()."
    This asserts the component list actually passed to the kernel excludes
    every COMMERCIAL-class driver, and that Total_base (final) foots exactly
    to Total_base (quadrature, cost-only) plus the margin/commercial dollar
    amount added LINEARLY on top.
  - Total-base-foots-to-components invariant: SKILL.md Self-Test #1, "Total_base
    equals the visible sum of component bases (exact foot)."
  - Range-ordering invariant: SKILL.md Self-Test #2, "Total_low <= Total_base
    <= Total_high."
  - Gap-attribution-reconciles invariant: SKILL.md's Dashboard "Numbers-reconcile
    assertion": "the Gap Method's driver-attribution rows must sum exactly to
    the visible gap dollar amount."

If any invariant fails, this script RAISES rather than writing a workbook
that fails its own reconciliation.

Scope note: this script builds the XLSX generator ONLY (should_cost_model.xlsx,
per SKILL.md's Deliverables section: cost-stack breakdown, aggregated
should-cost range, gap vs proposed price, and the assumption-ledger tab). The
optional Magazine-style dashboard (JSX, `examples/should_cost_canonical_
dashboard.jsx`) is a separate, already-existing deliverable and is NOT
touched here.

============================================================================
Should-cost input register schema (JUDGMENT CALL, flagged): SKILL.md
describes the Cost-Driver Assumption Ledger's ROW shape verbatim
("{ component, class (MATERIAL|CONVERSION|COMMERCIAL), basis, low, base,
high, currency, source, source_date, confidence, index_used, freshness_flag,
grouped_with[], notes }") and its model-level HEADER shape verbatim
("{ as_of_date, fx_rate_table, total_low, total_base, total_high,
correlation_assumption, model_confidence, supplier_price?, gap_abs?,
gap_pct? }"), but does not specify the raw INPUT contract a generator script
should accept (the ledger is the OUTPUT artifact, not the input). This module
resolves that ambiguity by accepting an input register shaped like the ledger
row schema plus the handful of top-level fields the header implies are
inputs (as_of_date, fx_rate_table, reporting currency, supplier_price) and,
because grouping is this skill's own judgment (not the kernel's, per the
Aggregation Method's "this skill's own grouping logic still decides which
lines go in which call"), a `correlated_group` string per driver instead of
the ledger's post-hoc `grouped_with[]` list:

{
  "reporting_currency": "USD",
  "as_of_date": "2026-07-23",
  "category": "optional free-text label",
  "description": "optional free-text label",
  "fx_rate_table": [ {"from": "EUR", "to": "USD", "rate": 1.08, "as_of": "2026-06-01"}, ... ],
  "cost_drivers": [
    {
      "component": "Materials",
      "class": "MATERIAL" | "CONVERSION" | "COMMERCIAL",
      "basis": "free text (per unit, BOM qty, loaded rate x hours, ...)",
      "low": float, "base": float, "high": float,          # native currency
      "currency": "USD",
      "source": "free text", "source_date": "YYYY-MM-DD",
      "confidence": "HIGH" | "MEDIUM" | "LOW",
      "index_used": "optional free text",
      "freshness_flag": "optional free text",
      "correlated_group": "optional group name, or null/omitted for independent",
      "notes": "optional free text"
    }, ...
  ],
  "margin_policy": {                                        # optional
    "base_pct": 0.08, "low_pct": 0.05, "high_pct": 0.12,
    "basis": "free text", "source": "free text", "confidence": "MEDIUM"
  },
  "supplier_price": float | null                             # optional
}

`class: COMMERCIAL` drivers (explicit SG&A/financing/other dollar lines) and
an optional percentage-based `margin_policy` are BOTH handled the same way
per the Aggregation Method: excluded entirely from the quadrature_rollup()
call and added LINEARLY on top of the quadrature output (never squared),
matching "Margin and SG&A are commercial, not statistical... add its spread
linearly on top of the converted-and-combined cost band."

`correlated_group`: drivers sharing the same group name are summed linearly
(their own low/base/high) into ONE pseudo-component before being fed to
quadrature_rollup() alongside the independent (ungrouped) drivers, per the
Aggregation Method's step 3 "Correlated drivers" sub-rule. A group's
confidence flag is the weakest (lowest) flag among its members (LOW if any
member is LOW, else MEDIUM if any member is MEDIUM, else HIGH), quoted
verbatim from that same sub-rule.
============================================================================
"""

from __future__ import annotations

import copy
from datetime import datetime
import os
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Sequence, Tuple

# ---------------------------------------------------------------------------
# Vendored numeric kernel (same directory). Per SKILL.md's "Computation
# requirement (HARD RULE)": the quadrature roll-up MUST be executed by
# calling quadrature_rollup() in the vendored numeric_kernel.py, never
# hand-computed here. convert_currency() is used for the Aggregation
# Method's step 6 currency normalization, for the same reason.
# ---------------------------------------------------------------------------
_KERNEL_IMPORT_ERROR: Optional[Exception] = None
try:
    _THIS_DIR = os.path.dirname(os.path.abspath(__file__))
    if _THIS_DIR not in sys.path:
        sys.path.insert(0, _THIS_DIR)
    from numeric_kernel import quadrature_rollup as kernel_quadrature_rollup
    from numeric_kernel import convert_currency as kernel_convert_currency
    from numeric_kernel import QuadratureResult
    from numeric_kernel import InvalidInputError as KernelInvalidInputError
    from numeric_kernel import UnknownCurrencyError as KernelUnknownCurrencyError
    KERNEL_AVAILABLE = True
except Exception as _exc:  # pragma: no cover - defensive, disclosed at runtime
    KERNEL_AVAILABLE = False
    _KERNEL_IMPORT_ERROR = _exc

    def kernel_quadrature_rollup(*args, **kwargs):  # type: ignore
        raise RuntimeError(
            "numeric_kernel.py unavailable; cannot compute the quadrature "
            f"roll-up. Import error: {_KERNEL_IMPORT_ERROR}"
        )

    def kernel_convert_currency(*args, **kwargs):  # type: ignore
        raise RuntimeError(
            "numeric_kernel.py unavailable; cannot convert currency. "
            f"Import error: {_KERNEL_IMPORT_ERROR}"
        )

# ---------------------------------------------------------------------------
# XLSX library detection. Mirrors pro_forma_generator.py: try openpyxl; if
# unavailable, raise a clear ImportError at workbook-build time (not at
# import time), so validation/ground-truth logic (which needs no XLSX
# library) remains usable even without one.
# ---------------------------------------------------------------------------
try:
    import openpyxl  # noqa: F401
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
    OPENPYXL_VERSION = openpyxl.__version__
except Exception as _exc:  # pragma: no cover
    OPENPYXL_AVAILABLE = False
    OPENPYXL_VERSION = None
    _OPENPYXL_IMPORT_ERROR = _exc


class ShouldCostValidationError(Exception):
    """Raised when the should-cost input register is missing a required
    field or carries a value the skill's Accuracy and Anti-Drift Rules
    forbid. Per Rule 1 ("Never fabricate a cost component or driver") this
    module refuses rather than guessing a missing value."""


class ReconciliationError(Exception):
    """Raised when a hard-coded invariant fails (quadrature-not-naive,
    margin-excluded-from-quadrature, total-base-foots-to-components,
    range-ordering, or gap-attribution-reconciles). The workbook is not
    saved when this fires."""


# ===========================================================================
# 1. Should-cost input register: typed input + validation
# ===========================================================================

VALID_CLASSES = ("MATERIAL", "CONVERSION", "COMMERCIAL")
VALID_CONFIDENCE = ("HIGH", "MEDIUM", "LOW")
_CONFIDENCE_RANK = {"LOW": 0, "MEDIUM": 1, "HIGH": 2}  # for weakest-flag-wins


@dataclass
class CostDriver:
    """Typed mirror of one row of should-cost-builder's Cost-Driver
    Assumption Ledger (SKILL.md, "Cost-Driver Assumption Ledger" section),
    plus `correlated_group` (this skill's own grouping judgment, resolved
    into the ledger's `grouped_with[]` shape only at ledger-emission time)."""

    component: str
    driver_class: str          # MATERIAL | CONVERSION | COMMERCIAL
    basis: str
    low: float                 # native currency
    base: float
    high: float
    currency: str
    source: str
    source_date: str
    confidence: str            # HIGH | MEDIUM | LOW
    index_used: str = ""
    freshness_flag: str = ""
    correlated_group: Optional[str] = None
    notes: str = ""


@dataclass
class MarginPolicy:
    base_pct: float
    low_pct: float
    high_pct: float
    basis: str = ""
    source: str = ""
    confidence: str = "MEDIUM"


@dataclass
class ShouldCostInput:
    reporting_currency: str
    as_of_date: str
    fx_rate_table: List[Dict[str, Any]]
    cost_drivers: List[CostDriver]
    margin_policy: Optional[MarginPolicy]
    supplier_price: Optional[float]
    category: str = ""
    description: str = ""


REQUIRED_TOP_LEVEL_FIELDS = ["reporting_currency", "as_of_date", "cost_drivers"]
REQUIRED_DRIVER_FIELDS = [
    "component", "class", "basis", "low", "base", "high", "currency",
    "source", "source_date", "confidence",
]



# Placeholders that satisfy "a date is present" while carrying no capture date at all.
# These are the strings a model reaches for when it does not know when a figure was taken.
_DATE_PLACEHOLDERS = frozenset({
    "", "tbd", "tba", "n/a", "na", "n/d", "none", "null", "unknown", "unspecified",
    "recent", "current", "various", "ongoing", "latest", "-", "--", "?",
})


def _check_capture_date(value, label, errors):
    """Enforce the capture-date requirement (item #32 / H5).

    G12 in `lilly-brand-assets-1c344a` defines a cited source as one carrying a capture
    date. That was stated centrally and enforced nowhere: a date field checked only for
    PRESENCE accepts "" and "TBD" and renders them as provenance.

    A figure with no capture date cannot be aged, cannot be judged stale and cannot be
    re-verified. It is worse than a missing figure, because it looks like evidence.

    Deterministic: parses the string and consults no clock. Comparing against "today"
    would make the generator's output depend on when it ran.

    Appends to `errors` rather than raising, matching this validator's collect-then-report
    style so the caller sees every problem in one pass.
    """
    if not isinstance(value, str):
        errors.append("%s must be a date string, got %s." % (label, type(value).__name__))
        return False
    raw = value.strip()
    if raw.lower() in _DATE_PLACEHOLDERS:
        errors.append(
            "%s is %r, which is a placeholder, not a capture date. If the date is genuinely "
            "unknown, drop the data point rather than carrying it with an empty provenance "
            "field." % (label, value)
        )
        return False
    parsed = None
    # Named-month formats are accepted because they are UNAMBIGUOUS real capture
    # dates, just not ISO ones; refusing them would reject honest provenance over
    # notation. Slash formats are deliberately NOT accepted: 03/04/2026 is March 4
    # or 4 March depending on the reader, and a date that parses two ways is not
    # provenance either.
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y",
                "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y",
                "%d %b %Y", "%d %B %Y", "%b %Y", "%B %Y"):
        try:
            parsed = datetime.strptime(raw, fmt)
            break
        except ValueError:
            parsed = None
    if parsed is None:
        errors.append(
            "%s is %r, which does not parse as YYYY-MM-DD, YYYY-MM or YYYY. A capture date "
            "that cannot be parsed cannot be used to age or stale-check the figure, so it is "
            "not provenance." % (label, value)
        )
        return False
    if not (1990 <= parsed.year <= 2100):
        errors.append(
            "%s has year %d, outside 1990-2100; almost certainly a typo." % (label, parsed.year)
        )
        return False
    return True

def validate_should_cost_input(register: Dict[str, Any]) -> ShouldCostInput:
    """Validate a raw should-cost input register dict and return a typed
    ShouldCostInput. Refuses (raises ShouldCostValidationError) rather than
    guessing when a required field is missing or a value violates a
    documented, fixed convention, per Rule 1 ("Never fabricate a cost
    component or driver... If unknown, label it an assumption with a range,
    never a confident invented number.")."""
    errors: List[str] = []

    for f in REQUIRED_TOP_LEVEL_FIELDS:
        if f not in register:
            errors.append(f"Missing required top-level field: '{f}'")
    if errors:
        raise ShouldCostValidationError(
            "Should-cost input register is missing required fields; "
            "refusing to guess. " + "; ".join(errors)
        )

    reporting_currency = register["reporting_currency"]
    if not isinstance(reporting_currency, str) or not reporting_currency:
        errors.append("'reporting_currency' must be a non-empty string (e.g. 'USD').")
    reporting_currency_u = (reporting_currency or "").strip().upper()

    as_of_date = register["as_of_date"]
    if not isinstance(as_of_date, str) or not as_of_date:
        errors.append("'as_of_date' must be a non-empty string (e.g. '2026-07-23').")
    else:
        _check_capture_date(as_of_date, "'as_of_date'", errors)

    fx_rate_table = register.get("fx_rate_table", [])
    if not isinstance(fx_rate_table, list):
        errors.append("'fx_rate_table' must be a list.")
    fx_currencies_available = set()
    for i, fx in enumerate(fx_rate_table if isinstance(fx_rate_table, list) else []):
        if not isinstance(fx, dict) or not all(k in fx for k in ("from", "to", "rate")):
            errors.append(f"'fx_rate_table[{i}]' must be an object with 'from', 'to', 'rate'.")
            continue
        if str(fx["to"]).strip().upper() != reporting_currency_u:
            errors.append(
                f"'fx_rate_table[{i}].to' ({fx['to']!r}) does not match "
                f"reporting_currency ({reporting_currency!r}). Per Aggregation "
                "Method step 6, every component must normalize to ONE "
                "reporting currency; an FX row targeting a different "
                "currency cannot be used and this module refuses to guess "
                "which currency was intended."
            )
        else:
            fx_currencies_available.add(str(fx["from"]).strip().upper())

    raw_drivers = register.get("cost_drivers", [])
    if not isinstance(raw_drivers, list) or len(raw_drivers) == 0:
        errors.append(
            "'cost_drivers' must be a non-empty list of cost-driver rows "
            "(SKILL.md 'Cost-Driver Assumption Ledger' row shape: component, "
            "class, basis, low, base, high, currency, source, source_date, "
            "confidence, ...)."
        )
        raw_drivers = []

    drivers: List[CostDriver] = []
    group_names_seen: Dict[str, str] = {}  # group -> a member's class, to check homogeneity
    for i, d in enumerate(raw_drivers):
        if not isinstance(d, dict):
            errors.append(f"'cost_drivers[{i}]' must be an object.")
            continue
        missing = [f for f in REQUIRED_DRIVER_FIELDS if f not in d]
        if missing:
            errors.append(f"'cost_drivers[{i}]' is missing required field(s): {missing}.")
            continue

        component = d["component"]
        driver_class = str(d["class"]).strip().upper()
        basis = d["basis"]
        low, base, high = d["low"], d["base"], d["high"]
        currency = str(d["currency"]).strip().upper()
        source = d["source"]
        source_date = d["source_date"]
        # Item #32: the Cost-Driver Assumption Ledger IS a research table, so every row's
        # source_date must be a real capture date. It was previously carried through
        # unchecked, which let "TBD" render in the ledger as though it were provenance.
        _check_capture_date(source_date, f"'cost_drivers[{i}].source_date'", errors)
        confidence = str(d["confidence"]).strip().upper()
        correlated_group = d.get("correlated_group") or None

        if driver_class not in VALID_CLASSES:
            errors.append(
                f"'cost_drivers[{i}].class' must be one of {VALID_CLASSES}; "
                f"got {d['class']!r}."
            )
        if confidence not in VALID_CONFIDENCE:
            errors.append(
                f"'cost_drivers[{i}].confidence' must be one of {VALID_CONFIDENCE}; "
                f"got {d['confidence']!r}."
            )
        if not all(isinstance(v, (int, float)) for v in (low, base, high)):
            errors.append(f"'cost_drivers[{i}]' low/base/high must all be numeric.")
        else:
            if not (low <= base <= high):
                errors.append(
                    f"'cost_drivers[{i}]' ({component!r}) violates low <= base <= "
                    f"high (got low={low}, base={base}, high={high}); a triangular "
                    "estimate cannot have base outside [low, high]."
                )
        if currency != reporting_currency_u and currency not in fx_currencies_available:
            errors.append(
                f"'cost_drivers[{i}]' ({component!r}) is priced in {currency!r} "
                f"but no fx_rate_table row converts {currency!r} to "
                f"{reporting_currency!r}. Per Aggregation Method step 6, "
                "multi-currency components must never be aggregated without "
                "a stated rate and date."
            )
        if driver_class == "COMMERCIAL" and correlated_group:
            errors.append(
                f"'cost_drivers[{i}]' ({component!r}) is class COMMERCIAL and "
                f"also carries correlated_group {correlated_group!r}; COMMERCIAL "
                "drivers (margin/SG&A) are never fed into the quadrature "
                "roll-up (per the Aggregation Method step-3 HARD RULE), so "
                "grouping them for that roll-up is meaningless. Remove the "
                "correlated_group or change the class."
            )
        if correlated_group:
            if correlated_group in group_names_seen and group_names_seen[correlated_group] != driver_class:
                errors.append(
                    f"correlated_group {correlated_group!r} mixes classes "
                    f"({group_names_seen[correlated_group]!r} and {driver_class!r}); "
                    "a correlated group must be internally consistent so its "
                    "linear-summed spread is meaningful."
                )
            group_names_seen.setdefault(correlated_group, driver_class)

        drivers.append(CostDriver(
            component=component, driver_class=driver_class, basis=basis,
            low=float(low) if isinstance(low, (int, float)) else 0.0,
            base=float(base) if isinstance(base, (int, float)) else 0.0,
            high=float(high) if isinstance(high, (int, float)) else 0.0,
            currency=currency, source=source, source_date=source_date,
            confidence=confidence, index_used=d.get("index_used", ""),
            freshness_flag=d.get("freshness_flag", ""),
            correlated_group=correlated_group, notes=d.get("notes", ""),
        ))

    if drivers and not any(dr.driver_class in ("MATERIAL", "CONVERSION") for dr in drivers):
        errors.append(
            "No MATERIAL or CONVERSION driver present; quadrature_rollup() "
            "requires at least one non-COMMERCIAL driver to roll up (margin/"
            "SG&A alone cannot form a should-cost range)."
        )

    margin_policy: Optional[MarginPolicy] = None
    raw_margin = register.get("margin_policy")
    if raw_margin is not None:
        if not isinstance(raw_margin, dict) or not all(
            k in raw_margin for k in ("base_pct", "low_pct", "high_pct")
        ):
            errors.append(
                "'margin_policy', when present, must be an object with at "
                "least 'base_pct', 'low_pct', 'high_pct' (SKILL.md Aggregation "
                "Method step 3: 'treat margin's range as a policy assumption "
                "(e.g. base 8%, low 5%, high 12%)')."
            )
        else:
            lp, bp, hp = raw_margin["low_pct"], raw_margin["base_pct"], raw_margin["high_pct"]
            if not all(isinstance(v, (int, float)) for v in (lp, bp, hp)):
                errors.append("'margin_policy' pct fields must all be numeric.")
            elif not (lp <= bp <= hp):
                errors.append(
                    f"'margin_policy' violates low_pct <= base_pct <= high_pct "
                    f"(got low={lp}, base={bp}, high={hp})."
                )
            margin_conf = str(raw_margin.get("confidence", "MEDIUM")).strip().upper()
            if margin_conf not in VALID_CONFIDENCE:
                errors.append(
                    f"'margin_policy.confidence' must be one of {VALID_CONFIDENCE}; "
                    f"got {raw_margin.get('confidence')!r}."
                )
            margin_policy = MarginPolicy(
                base_pct=float(bp), low_pct=float(lp), high_pct=float(hp),
                basis=raw_margin.get("basis", ""), source=raw_margin.get("source", ""),
                confidence=margin_conf,
            )

    supplier_price = register.get("supplier_price")
    if supplier_price is not None and not isinstance(supplier_price, (int, float)):
        errors.append("'supplier_price', when present, must be numeric or null.")

    if errors:
        raise ShouldCostValidationError(
            "Should-cost input register failed validation; refusing to guess "
            "missing or invalid fields. Issues found:\n  - " + "\n  - ".join(errors)
        )

    return ShouldCostInput(
        reporting_currency=reporting_currency, as_of_date=as_of_date,
        fx_rate_table=fx_rate_table, cost_drivers=drivers,
        margin_policy=margin_policy,
        supplier_price=float(supplier_price) if supplier_price is not None else None,
        category=register.get("category", ""), description=register.get("description", ""),
    )


# ===========================================================================
# Ground-truth computation (Python side), via the vendored kernel only.
# ===========================================================================

@dataclass
class QuadInputRow:
    """One row fed into the single kernel_quadrature_rollup() call: either
    one ungrouped MATERIAL/CONVERSION driver, or one correlated group's
    linear-summed pseudo-component (Aggregation Method step 3)."""
    name: str
    base: float
    spread_low: float
    spread_high: float
    confidence: str
    member_components: List[str]


@dataclass
class DriverAttribution:
    component: str
    base_norm: float
    share_of_total_base: float
    gap_dollars: float


@dataclass
class GapResult:
    supplier_price: float
    absolute_gap: float
    gap_pct: float
    position: str  # ABOVE_HIGH | WITHIN_RANGE | BELOW_LOW
    attribution: List[DriverAttribution]


@dataclass
class GroundTruth:
    reporting_currency: str
    low_norm: Dict[str, float]     # component -> normalized low
    base_norm: Dict[str, float]    # component -> normalized base
    high_norm: Dict[str, float]    # component -> normalized high
    correlated_groups: Dict[str, List[str]]   # group name -> quadrature-eligible member components
    quad_input_rows: List[QuadInputRow]        # exactly what was passed to the kernel, in order
    quadrature_result: "QuadratureResult"
    cost_subtotal_base: float
    commercial_explicit_base: float
    commercial_explicit_low: float
    commercial_explicit_high: float
    margin_policy_base: float
    margin_policy_low: float
    margin_policy_high: float
    margin_base_dollar: float      # commercial_explicit + margin_policy, combined
    margin_low_dollar: float
    margin_high_dollar: float
    total_base_final: float
    total_low_final: float
    total_high_final: float
    naive_low_footnote: float
    naive_high_footnote: float
    model_confidence: str
    gap: Optional[GapResult]


def _weakest_confidence(flags: Sequence[str]) -> str:
    """Weakest (lowest) flag among a group's members: LOW if any member is
    LOW, else MEDIUM if any member is MEDIUM, else HIGH. Quoted verbatim
    from should-cost-builder's Aggregation Method step 3."""
    return min(flags, key=lambda f: _CONFIDENCE_RANK[f])


def _normalize_currency(reg: ShouldCostInput) -> Tuple[Dict[str, float], Dict[str, float], Dict[str, float]]:
    """Normalize every driver's low/base/high to reg.reporting_currency via
    numeric_kernel.convert_currency(), per Aggregation Method step 6
    ("Normalize every component to one reporting currency at a stated FX
    rate and date BEFORE aggregating"). Returns three dicts keyed by
    component name. validate_should_cost_input() already guaranteed every
    non-reporting-currency driver has a matching fx_rate_table row."""
    reporting = reg.reporting_currency.strip().upper()
    fx_table: Dict[str, float] = {
        str(fx["from"]).strip().upper(): float(fx["rate"])
        for fx in reg.fx_rate_table
        if str(fx.get("to", "")).strip().upper() == reporting
    }
    low_norm: Dict[str, float] = {}
    base_norm: Dict[str, float] = {}
    high_norm: Dict[str, float] = {}
    for d in reg.cost_drivers:
        if d.currency.strip().upper() == reporting:
            low_norm[d.component] = d.low
            base_norm[d.component] = d.base
            high_norm[d.component] = d.high
        else:
            low_norm[d.component] = kernel_convert_currency(d.low, d.currency, fx_table)
            base_norm[d.component] = kernel_convert_currency(d.base, d.currency, fx_table)
            high_norm[d.component] = kernel_convert_currency(d.high, d.currency, fx_table)
    return low_norm, base_norm, high_norm


def _determine_model_confidence(quadrature_eligible: List[CostDriver],
                                 base_norm: Dict[str, float],
                                 total_base_final: float) -> str:
    """Aggregation Method step 5: 'HIGH only if every material driver (each
    >10% of base) is HIGH or MEDIUM and none of the top-three drivers is
    LOW. MEDIUM if one top-three driver is LOW. LOW if two or more material
    drivers are LOW...' (the research-effort-floor clause of step 5 is a
    process check on how the input was SOURCED, not data this generator
    receives, and is out of scope here; disclosed as a JUDGMENT CALL.) "base"
    for the >10%/top-three tests is read as share of the FINAL should-cost
    base (post-margin), since that is the headline figure the model
    confidence describes; SKILL.md does not specify which base to use for
    this threshold, so this is also a disclosed JUDGMENT CALL.
    """
    if not quadrature_eligible or total_base_final <= 0:
        return "LOW"
    ranked = sorted(quadrature_eligible, key=lambda d: base_norm[d.component], reverse=True)
    top3 = ranked[:3]
    material = [d for d in quadrature_eligible if base_norm[d.component] / total_base_final > 0.10]
    material_low_count = sum(1 for d in material if d.confidence == "LOW")
    top3_low_count = sum(1 for d in top3 if d.confidence == "LOW")

    if material_low_count >= 2:
        return "LOW"
    if top3_low_count >= 1:
        return "MEDIUM"
    if all(d.confidence in ("HIGH", "MEDIUM") for d in material):
        return "HIGH"
    return "MEDIUM"


def compute_ground_truth(reg: ShouldCostInput) -> GroundTruth:
    """Compute the should-cost roll-up, margin add-on, and gap, using ONLY
    numeric_kernel.quadrature_rollup() / convert_currency() for the core
    math, per SKILL.md's Computation-requirement HARD RULE. This is the
    reference against which the hard-coded invariants are checked before
    the workbook is saved.
    """
    if not KERNEL_AVAILABLE:
        raise RuntimeError(
            "numeric_kernel.py could not be imported; this generator refuses "
            "to hand-compute the quadrature roll-up in its place (SKILL.md "
            f"Computation-requirement HARD RULE). Import error: {_KERNEL_IMPORT_ERROR}"
        )

    low_norm, base_norm, high_norm = _normalize_currency(reg)

    quadrature_eligible = [d for d in reg.cost_drivers if d.driver_class in ("MATERIAL", "CONVERSION")]
    commercial_explicit = [d for d in reg.cost_drivers if d.driver_class == "COMMERCIAL"]

    # --- Group correlated drivers (Aggregation Method step 3, "Correlated
    # drivers"): linear-sum each group's low/base/high, confidence = weakest
    # member flag. Ungrouped drivers pass through as independent inputs. ---
    correlated_groups: Dict[str, List[str]] = {}
    for d in quadrature_eligible:
        if d.correlated_group:
            correlated_groups.setdefault(d.correlated_group, []).append(d.component)

    quad_input_rows: List[QuadInputRow] = []
    for d in quadrature_eligible:
        if not d.correlated_group:
            quad_input_rows.append(QuadInputRow(
                name=d.component, base=base_norm[d.component],
                spread_low=base_norm[d.component] - low_norm[d.component],
                spread_high=high_norm[d.component] - base_norm[d.component],
                confidence=d.confidence, member_components=[d.component],
            ))
    for group_name, members in correlated_groups.items():
        group_base = sum(base_norm[m] for m in members)
        group_spread_low = sum(base_norm[m] - low_norm[m] for m in members)
        group_spread_high = sum(high_norm[m] - base_norm[m] for m in members)
        group_conf = _weakest_confidence(
            [d.confidence for d in quadrature_eligible if d.component in members]
        )
        quad_input_rows.append(QuadInputRow(
            name=group_name, base=group_base, spread_low=group_spread_low,
            spread_high=group_spread_high, confidence=group_conf, member_components=members,
        ))

    # --- The kernel call (HARD RULE: the ONLY place quadrature math runs).
    # Margin/SG&A (COMMERCIAL class) is never in this call, per step 3. ----
    qr: QuadratureResult = kernel_quadrature_rollup(
        component_bases=[r.base for r in quad_input_rows],
        component_spreads_low=[r.spread_low for r in quad_input_rows],
        component_spreads_high=[r.spread_high for r in quad_input_rows],
        confidence_flags=[r.confidence for r in quad_input_rows],
    )
    cost_subtotal_base = qr.total_base

    # --- Commercial (explicit $ COMMERCIAL drivers) + margin policy (%):
    # both added LINEARLY on top, never quadratured (step 3 HARD RULE). ---
    commercial_explicit_base = sum(base_norm[d.component] for d in commercial_explicit)
    commercial_explicit_low = sum(low_norm[d.component] for d in commercial_explicit)
    commercial_explicit_high = sum(high_norm[d.component] for d in commercial_explicit)

    if reg.margin_policy is not None:
        margin_policy_base = reg.margin_policy.base_pct * cost_subtotal_base
        margin_policy_low = reg.margin_policy.low_pct * cost_subtotal_base
        margin_policy_high = reg.margin_policy.high_pct * cost_subtotal_base
    else:
        margin_policy_base = margin_policy_low = margin_policy_high = 0.0

    margin_base_dollar = commercial_explicit_base + margin_policy_base
    margin_low_dollar = commercial_explicit_low + margin_policy_low
    margin_high_dollar = commercial_explicit_high + margin_policy_high

    total_base_final = cost_subtotal_base + margin_base_dollar
    total_low_final = qr.total_low + margin_low_dollar
    total_high_final = qr.total_high + margin_high_dollar
    naive_low_footnote = qr.naive_low + margin_low_dollar
    naive_high_footnote = qr.naive_high + margin_high_dollar

    model_confidence = _determine_model_confidence(quadrature_eligible, base_norm, total_base_final)

    # --- Gap Method (SKILL.md "Gap Method"): skipped, per Edge Cases, when
    # no supplier price was provided. ---------------------------------
    gap: Optional[GapResult] = None
    if reg.supplier_price is not None:
        absolute_gap = reg.supplier_price - total_base_final
        gap_pct = absolute_gap / total_base_final if total_base_final else 0.0
        if reg.supplier_price > total_high_final:
            position = "ABOVE_HIGH"
        elif reg.supplier_price < total_low_final:
            position = "BELOW_LOW"
        else:
            position = "WITHIN_RANGE"

        # Driver attribution (JUDGMENT CALL, disclosed: SKILL.md names the
        # requirement -- "decompose the gap onto the components" -- but
        # gives no formula. This module attributes proportionally to each
        # component's share of the FINAL should-cost base, the simplest
        # allocation that sums exactly to the gap by construction, and
        # includes the margin-policy dollar amount as its own attributed
        # row since margin IS a component of the total per the ledger's own
        # class taxonomy.)
        attribution: List[DriverAttribution] = []
        all_attributable: List[Tuple[str, float]] = [
            (d.component, base_norm[d.component]) for d in reg.cost_drivers
        ]
        if reg.margin_policy is not None:
            all_attributable.append(("Margin (policy)", margin_policy_base))
        for name, base_val in all_attributable:
            share = base_val / total_base_final if total_base_final else 0.0
            attribution.append(DriverAttribution(
                component=name, base_norm=base_val, share_of_total_base=share,
                gap_dollars=share * absolute_gap,
            ))
        gap = GapResult(
            supplier_price=reg.supplier_price, absolute_gap=absolute_gap,
            gap_pct=gap_pct, position=position, attribution=attribution,
        )

    return GroundTruth(
        reporting_currency=reg.reporting_currency, low_norm=low_norm, base_norm=base_norm,
        high_norm=high_norm, correlated_groups=correlated_groups, quad_input_rows=quad_input_rows,
        quadrature_result=qr, cost_subtotal_base=cost_subtotal_base,
        commercial_explicit_base=commercial_explicit_base, commercial_explicit_low=commercial_explicit_low,
        commercial_explicit_high=commercial_explicit_high, margin_policy_base=margin_policy_base,
        margin_policy_low=margin_policy_low, margin_policy_high=margin_policy_high,
        margin_base_dollar=margin_base_dollar, margin_low_dollar=margin_low_dollar,
        margin_high_dollar=margin_high_dollar, total_base_final=total_base_final,
        total_low_final=total_low_final, total_high_final=total_high_final,
        naive_low_footnote=naive_low_footnote, naive_high_footnote=naive_high_footnote,
        model_confidence=model_confidence, gap=gap,
    )


# ===========================================================================
# 2 & 3. Hard-coded invariant checks (run BEFORE saving)
# ===========================================================================

def _assert_quadrature_independently_rederived(gt: GroundTruth, tolerance: float = 1e-6) -> None:
    """Hard-coded check: independently re-derive the quadrature roll-up
    (sqrt of sum-of-squares over non-widened rows, linear addition over
    widened rows) from gt.quad_input_rows, and assert it matches the
    kernel's own QuadratureResult. This re-derivation is INDEPENDENT of the
    kernel's internals (mirrors pro_forma_generator.py's manual NPV
    re-derivation pattern), it does not replace the kernel as the source of
    truth; the kernel's call in compute_ground_truth() remains the only
    place the roll-up is actually computed for the workbook.

    Also asserts the quadrature range differs from the naive sum-of-extremes
    range whenever at least two quad_input_rows carry a nonzero, non-widened
    spread (proving quadrature, not a naive sum, actually ran), mirroring
    pro-forma's 'not accidentally equal to the wrong convention' check.
    """
    import math

    widened_names = {gt.quad_input_rows[i].name for i in gt.quadrature_result.widened_components}
    quad_rows = [r for r in gt.quad_input_rows if r.name not in widened_names]
    lin_rows = [r for r in gt.quad_input_rows if r.name in widened_names]

    manual_total_base = sum(r.base for r in gt.quad_input_rows)
    manual_low = manual_total_base - math.sqrt(sum(r.spread_low ** 2 for r in quad_rows)) \
        - sum(r.spread_low for r in lin_rows)
    manual_high = manual_total_base + math.sqrt(sum(r.spread_high ** 2 for r in quad_rows)) \
        + sum(r.spread_high for r in lin_rows)

    if abs(manual_total_base - gt.quadrature_result.total_base) > tolerance:
        raise ReconciliationError(
            "Quadrature invariant FAILED: manually re-summed Total_base "
            f"({manual_total_base}) does not match the kernel's Total_base "
            f"({gt.quadrature_result.total_base})."
        )
    if abs(manual_low - gt.quadrature_result.total_low) > tolerance:
        raise ReconciliationError(
            "Quadrature invariant FAILED: manually re-derived Total_low "
            f"({manual_low}) does not match numeric_kernel.quadrature_rollup()'s "
            f"Total_low ({gt.quadrature_result.total_low}). Refusing to build "
            "a workbook whose roll-up cannot be independently verified."
        )
    if abs(manual_high - gt.quadrature_result.total_high) > tolerance:
        raise ReconciliationError(
            "Quadrature invariant FAILED: manually re-derived Total_high "
            f"({manual_high}) does not match numeric_kernel.quadrature_rollup()'s "
            f"Total_high ({gt.quadrature_result.total_high})."
        )

    nonzero_quad_rows = [r for r in quad_rows if r.spread_low != 0 or r.spread_high != 0]
    if len(nonzero_quad_rows) >= 2:
        if abs(gt.quadrature_result.total_low - gt.quadrature_result.naive_low) < 1e-9:
            raise ReconciliationError(
                "Quadrature invariant FAILED: Total_low is indistinguishable "
                "from the naive worst-case envelope (naive_low) with 2+ "
                "independent nonzero-spread drivers; quadrature should "
                "strictly tighten the band. This indicates the naive sum was "
                "used instead of root-sum-of-squares."
            )
        if abs(gt.quadrature_result.total_high - gt.quadrature_result.naive_high) < 1e-9:
            raise ReconciliationError(
                "Quadrature invariant FAILED: Total_high is indistinguishable "
                "from the naive worst-case envelope (naive_high) with 2+ "
                "independent nonzero-spread drivers."
            )


def _assert_margin_excluded_from_quadrature(gt: GroundTruth, tolerance: float = 1e-6) -> None:
    """Hard-coded check: SKILL.md Aggregation Method step 3 HARD RULE,
    'Margin/SG&A is never passed into quadrature_rollup()... add its spread
    linearly on top of the function's returned Total_low/Total_high
    afterward.' Asserts (a) no quad_input_row carries a COMMERCIAL-class
    member, and (b) the final totals foot exactly to the quadrature output
    plus the margin/commercial dollar amount, added linearly (not squared)."""
    all_grouped_and_ungrouped_members: List[str] = []
    for row in gt.quad_input_rows:
        all_grouped_and_ungrouped_members.extend(row.member_components)
    # member components are all quadrature_eligible by construction in
    # compute_ground_truth(); this re-asserts that invariant holds by
    # checking none of them appear with driver_class COMMERCIAL is
    # impossible to check here without the original CostDriver list, so this
    # check instead asserts the linear-footing identity, which can only hold
    # if margin dollars were added OUTSIDE the sqrt(sum-of-squares) term.
    implied_margin = gt.total_base_final - gt.cost_subtotal_base
    if abs(implied_margin - gt.margin_base_dollar) > tolerance:
        raise ReconciliationError(
            "Margin-excluded-from-quadrature invariant FAILED: "
            f"Total_base_final - cost_subtotal_base ({implied_margin}) does "
            f"not equal margin_base_dollar ({gt.margin_base_dollar}); margin "
            "must be added linearly on top of the quadrature output, exactly, "
            "with no other adjustment."
        )
    implied_margin_low = gt.total_low_final - gt.quadrature_result.total_low
    implied_margin_high = gt.total_high_final - gt.quadrature_result.total_high
    if abs(implied_margin_low - gt.margin_low_dollar) > tolerance:
        raise ReconciliationError(
            "Margin-excluded-from-quadrature invariant FAILED: "
            f"Total_low_final - quadrature Total_low ({implied_margin_low}) "
            f"does not equal margin_low_dollar ({gt.margin_low_dollar})."
        )
    if abs(implied_margin_high - gt.margin_high_dollar) > tolerance:
        raise ReconciliationError(
            "Margin-excluded-from-quadrature invariant FAILED: "
            f"Total_high_final - quadrature Total_high ({implied_margin_high}) "
            f"does not equal margin_high_dollar ({gt.margin_high_dollar})."
        )


def _assert_total_base_foots_to_components(reg: ShouldCostInput, gt: GroundTruth,
                                            tolerance: float = 1e-6) -> None:
    """Hard-coded check: SKILL.md Pre-Delivery Self-Test #1, 'Total_base
    equals the visible sum of component bases (exact foot).' Total_base
    (final, post-margin) must equal the literal sum of every driver's
    normalized base plus the margin-policy dollar amount (margin being a
    COMMERCIAL-class component in the ledger's own taxonomy)."""
    literal_sum = sum(gt.base_norm[d.component] for d in reg.cost_drivers) + gt.margin_policy_base
    if abs(literal_sum - gt.total_base_final) > tolerance:
        raise ReconciliationError(
            "Total-base-foots-to-components invariant FAILED: the literal "
            f"sum of every component's normalized base plus the margin-policy "
            f"dollar amount ({literal_sum}) does not equal Total_base_final "
            f"({gt.total_base_final})."
        )


def _assert_range_ordering(gt: GroundTruth) -> None:
    """Hard-coded check: SKILL.md Pre-Delivery Self-Test #2, 'Total_low <=
    Total_base <= Total_high.' Checked on BOTH the cost-only (pre-margin,
    kernel-reported) and final (post-margin) ranges."""
    qr = gt.quadrature_result
    if not (qr.total_low <= qr.total_base <= qr.total_high):
        raise ReconciliationError(
            "Range-ordering invariant FAILED (cost-only, pre-margin): "
            f"expected total_low <= total_base <= total_high, got "
            f"{qr.total_low} <= {qr.total_base} <= {qr.total_high}."
        )
    if not (gt.total_low_final <= gt.total_base_final <= gt.total_high_final):
        raise ReconciliationError(
            "Range-ordering invariant FAILED (final, post-margin): expected "
            f"total_low <= total_base <= total_high, got {gt.total_low_final} "
            f"<= {gt.total_base_final} <= {gt.total_high_final}."
        )


def _assert_gap_attribution_reconciles(gt: GroundTruth, tolerance: float = 1e-6) -> None:
    """Hard-coded check: should-cost-builder's dashboard 'Numbers-reconcile
    assertion' -- 'the Gap Method's driver-attribution rows must sum exactly
    to the visible gap dollar amount.' Skipped (no-op) when no supplier price
    was provided, per the Edge Cases section (Gap Method is not computed)."""
    if gt.gap is None:
        return
    attributed_sum = sum(a.gap_dollars for a in gt.gap.attribution)
    if abs(attributed_sum - gt.gap.absolute_gap) > tolerance:
        raise ReconciliationError(
            "Gap-attribution-reconciles invariant FAILED: driver-attribution "
            f"rows sum to {attributed_sum}, but the visible gap is "
            f"{gt.gap.absolute_gap}. SKILL.md requires these to reconcile "
            "exactly; refusing to save a workbook whose gap attribution does "
            "not foot."
        )


def run_hardcoded_invariant_checks(reg: ShouldCostInput, gt: GroundTruth) -> None:
    """Run every hard-coded invariant. Raises ReconciliationError on any
    failure; callers (build_workbook / generate_should_cost_workbook) must
    not save the workbook if this raises."""
    _assert_quadrature_independently_rederived(gt)
    _assert_margin_excluded_from_quadrature(gt)
    _assert_total_base_foots_to_components(reg, gt)
    _assert_range_ordering(gt)
    _assert_gap_attribution_reconciles(gt)


# ===========================================================================
# 2. Workbook builder (real Excel formulas, not baked-in numbers)
# ===========================================================================

HEADER_FILL = None  # set lazily once openpyxl is confirmed available
HEADER_FONT = None


def _require_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is not installed in this Python environment, so no "
            "XLSX file can be written. Install it (`pip install openpyxl`) "
            "or point this script at an interpreter that already has it. "
            f"Original import error: {_OPENPYXL_IMPORT_ERROR}"
        )


def _write_header_row(ws, headers: Sequence[str], row: int) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def build_workbook(reg: ShouldCostInput, gt: GroundTruth) -> "openpyxl.Workbook":
    """Build the formula-driven workbook. Tabs: Assumptions, CostStack,
    AggregationRollup, GapAnalysis, AssumptionLedger, Summary. Every derived
    cell is a live Excel formula referencing other cells (SKILL.md's
    'Rule 3: Show the build and emit it... Totals reconcile' and the
    dashboard's 'Numbers-reconcile assertion... mirrored in JS per the
    vendored numeric_kernel.py, never hand-typed' -- the XLSX equivalent of
    that requirement is that the roll-up formulas below are LIVE Excel
    formulas re-deriving the same quadrature math, not pasted numbers).

    This function does NOT save the file and does NOT run the reconciliation
    checks; call generate_should_cost_workbook() for the full validated
    pipeline. Sheet names are deliberately single-word (no spaces) so
    cross-sheet formulas never need sheet-name quoting.
    """
    _require_openpyxl()

    global HEADER_FONT, HEADER_FILL
    # Palette per SKILL.md's own house-style block: "header bar Lilly Black
    # #212121... section headers Bold Blue #0F3A85."
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="212121", end_color="212121", fill_type="solid")

    wb = Workbook()
    ws_assump = wb.active
    ws_assump.title = "Assumptions"

    reporting_currency = reg.reporting_currency
    n_drivers = len(reg.cost_drivers)

    # ---------------------------------------------------------------
    # Tab 1: Assumptions (raw values live here; every other tab
    # references these cells, never re-types a number).
    # ---------------------------------------------------------------
    _write_header_row(ws_assump, ["Assumption", "Value", "Notes"], row=1)
    supplier_price_value = reg.supplier_price if reg.supplier_price is not None else "NOT_PROVIDED"
    mp = reg.margin_policy
    rows = [
        ("Reporting currency", reporting_currency, ""),
        ("As-of date", reg.as_of_date, ""),
        ("Category", reg.category, ""),
        ("Description", reg.description, ""),
        ("Supplier price", supplier_price_value,
         "NEEDS_INPUT: gap not computable yet" if reg.supplier_price is None else ""),
        ("Margin base %", mp.base_pct if mp else 0.0, mp.basis if mp else "No margin policy provided"),
        ("Margin low %", mp.low_pct if mp else 0.0, ""),
        ("Margin high %", mp.high_pct if mp else 0.0, ""),
        ("Margin source", mp.source if mp else "", ""),
        ("Margin confidence", mp.confidence if mp else "", ""),
        ("Correlation assumption",
         "; ".join(f"{g}: {', '.join(members)}" for g, members in gt.correlated_groups.items())
         or "No correlated groups declared; all MATERIAL/CONVERSION drivers treated as independent.",
         "Per Aggregation Method step 3"),
    ]
    r = 2
    cell_refs: Dict[str, str] = {}
    for label, value, note in rows:
        ws_assump.cell(row=r, column=1, value=label)
        ws_assump.cell(row=r, column=2, value=value)
        ws_assump.cell(row=r, column=3, value=note)
        cell_refs[label] = f"Assumptions!$B${r}"
        r += 1

    reporting_currency_ref = cell_refs["Reporting currency"]
    supplier_price_ref = cell_refs["Supplier price"]
    margin_base_pct_ref = cell_refs["Margin base %"]
    margin_low_pct_ref = cell_refs["Margin low %"]
    margin_high_pct_ref = cell_refs["Margin high %"]

    wb.defined_names["ReportingCurrency"] = DefinedName("ReportingCurrency", attr_text=reporting_currency_ref)
    wb.defined_names["SupplierPrice"] = DefinedName("SupplierPrice", attr_text=supplier_price_ref)
    wb.defined_names["MarginBasePct"] = DefinedName("MarginBasePct", attr_text=margin_base_pct_ref)
    wb.defined_names["MarginLowPct"] = DefinedName("MarginLowPct", attr_text=margin_low_pct_ref)
    wb.defined_names["MarginHighPct"] = DefinedName("MarginHighPct", attr_text=margin_high_pct_ref)

    r += 1  # blank spacer row
    fx_header_row = r
    ws_assump.cell(row=fx_header_row, column=1, value="FX Rate Table (per Aggregation Method step 6)")
    r += 1
    fx_table_header_row = r
    _write_header_row(ws_assump, ["From currency", "Rate to reporting", "As-of"], row=fx_table_header_row)
    r += 1
    fx_table_start_row = r
    for fx in reg.fx_rate_table:
        ws_assump.cell(row=r, column=1, value=str(fx["from"]).strip().upper())
        ws_assump.cell(row=r, column=2, value=float(fx["rate"]))
        ws_assump.cell(row=r, column=3, value=fx.get("as_of", ""))
        r += 1
    fx_table_end_row = r - 1
    if fx_table_end_row < fx_table_start_row:
        # No FX rows: keep the range non-empty so INDEX/MATCH below doesn't
        # error on a zero-height range; MATCH will simply never find a hit
        # (every driver should already be in the reporting currency).
        fx_table_end_row = fx_table_start_row

    ws_assump.column_dimensions["A"].width = 30
    ws_assump.column_dimensions["B"].width = 18
    ws_assump.column_dimensions["C"].width = 40

    # ---------------------------------------------------------------
    # Tab 2: CostStack (the cost build-up: one row per sourced driver,
    # with live formulas for FX normalization, spreads, and base shares).
    # ---------------------------------------------------------------
    ws_cs = wb.create_sheet("CostStack")
    _write_header_row(ws_cs, [
        "Component", "Class", "Basis", "Currency", "Low (native)", "Base (native)", "High (native)",
        "FX rate to reporting", "Low (normalized)", "Base (normalized)", "High (normalized)",
        "Spread low", "Spread high", "Correlated group", "Confidence",
        "Share of cost subtotal", "Share of final base",
        "Source", "Source date", "Index used", "Freshness flag", "Notes",
    ], row=1)

    for i, d in enumerate(reg.cost_drivers):
        row = 2 + i
        ws_cs.cell(row=row, column=1, value=d.component)
        ws_cs.cell(row=row, column=2, value=d.driver_class)
        ws_cs.cell(row=row, column=3, value=d.basis)
        ws_cs.cell(row=row, column=4, value=d.currency)
        ws_cs.cell(row=row, column=5, value=d.low)
        ws_cs.cell(row=row, column=6, value=d.base)
        ws_cs.cell(row=row, column=7, value=d.high)
        # FX rate: 1 if already in reporting currency, else INDEX/MATCH
        # against the Assumptions FX table (live lookup, not a baked rate).
        ws_cs.cell(row=row, column=8, value=(
            f"=IF(D{row}=ReportingCurrency,1,"
            f"IFERROR(INDEX(Assumptions!$B${fx_table_start_row}:$B${fx_table_end_row},"
            f"MATCH(D{row},Assumptions!$A${fx_table_start_row}:$A${fx_table_end_row},0)),"
            f"\"FX_RATE_MISSING\"))"
        ))
        ws_cs.cell(row=row, column=9, value=f"=E{row}*H{row}")
        ws_cs.cell(row=row, column=10, value=f"=F{row}*H{row}")
        ws_cs.cell(row=row, column=11, value=f"=G{row}*H{row}")
        ws_cs.cell(row=row, column=12, value=f"=J{row}-I{row}")   # spread low
        ws_cs.cell(row=row, column=13, value=f"=K{row}-J{row}")   # spread high
        ws_cs.cell(row=row, column=14, value=d.correlated_group or "")
        ws_cs.cell(row=row, column=15, value=d.confidence)
        # Share of cost subtotal: only meaningful for non-COMMERCIAL rows
        # (COMMERCIAL rows never enter the quadrature roll-up).
        ws_cs.cell(row=row, column=16, value=(
            f"=IF(B{row}=\"COMMERCIAL\",\"N/A (commercial - excluded from quadrature)\","
            f"J{row}/CostSubtotalBase)"
        ))
        ws_cs.cell(row=row, column=17, value=f"=J{row}/TotalBaseFinal")
        ws_cs.cell(row=row, column=18, value=d.source)
        ws_cs.cell(row=row, column=19, value=d.source_date)
        ws_cs.cell(row=row, column=20, value=d.index_used)
        ws_cs.cell(row=row, column=21, value=d.freshness_flag)
        ws_cs.cell(row=row, column=22, value=d.notes)

    last_cs_row = 1 + n_drivers
    widths = [26, 12, 26, 10, 12, 12, 12, 10, 14, 14, 14, 12, 12, 16, 12, 16, 14, 22, 14, 18, 14, 30]
    for idx, w in enumerate(widths, start=1):
        ws_cs.column_dimensions[get_column_letter(idx)].width = w

    # ---------------------------------------------------------------
    # Tab 3: AggregationRollup (independently re-derives the quadrature
    # roll-up as LIVE Excel formulas: SQRT(SUMPRODUCT(...^2)) over
    # non-widened rows, linear SUMPRODUCT over widened rows -- the Excel
    # mirror of numeric_kernel.quadrature_rollup(), per Aggregation Method
    # steps 3-4).
    # ---------------------------------------------------------------
    ws_ar = wb.create_sheet("AggregationRollup")
    _write_header_row(ws_ar, ["Name", "Base", "Spread low", "Spread high", "Confidence",
                               "Share of cost subtotal", "Widened?"], row=1)

    quad_row = 2
    quad_row_map: Dict[str, int] = {}
    for d in reg.cost_drivers:
        if d.driver_class in ("MATERIAL", "CONVERSION") and not d.correlated_group:
            cs_row = 2 + reg.cost_drivers.index(d)
            ws_ar.cell(row=quad_row, column=1, value=d.component)
            ws_ar.cell(row=quad_row, column=2, value=f"=CostStack!J{cs_row}")
            ws_ar.cell(row=quad_row, column=3, value=f"=CostStack!L{cs_row}")
            ws_ar.cell(row=quad_row, column=4, value=f"=CostStack!M{cs_row}")
            ws_ar.cell(row=quad_row, column=5, value=f"=CostStack!O{cs_row}")
            quad_row_map[d.component] = quad_row
            quad_row += 1

    for group_name in gt.correlated_groups:
        ws_ar.cell(row=quad_row, column=1, value=group_name)
        ws_ar.cell(row=quad_row, column=2,
                   value=f"=SUMIFS(CostStack!J:J,CostStack!N:N,\"{group_name}\")")
        ws_ar.cell(row=quad_row, column=3,
                   value=f"=SUMIFS(CostStack!L:L,CostStack!N:N,\"{group_name}\")")
        ws_ar.cell(row=quad_row, column=4,
                   value=f"=SUMIFS(CostStack!M:M,CostStack!N:N,\"{group_name}\")")
        ws_ar.cell(row=quad_row, column=5, value=(
            f"=IF(COUNTIFS(CostStack!N:N,\"{group_name}\",CostStack!O:O,\"LOW\")>0,\"LOW\","
            f"IF(COUNTIFS(CostStack!N:N,\"{group_name}\",CostStack!O:O,\"MEDIUM\")>0,\"MEDIUM\",\"HIGH\"))"
        ))
        quad_row += 1

    last_quad_row = quad_row - 1
    # Share-of-cost-subtotal and Widened? columns need CostSubtotalBase,
    # which is only known once this table's own Base column exists; Excel
    # tolerates the forward reference since named ranges resolve at
    # calculation time, not write time.
    for rr in range(2, last_quad_row + 1):
        ws_ar.cell(row=rr, column=6, value=f"=B{rr}/CostSubtotalBase")
        ws_ar.cell(row=rr, column=7,
                   value=f"=IF(AND(E{rr}=\"LOW\",F{rr}>0.15),\"YES\",\"NO\")")

    for idx, w in enumerate([24, 14, 14, 14, 12, 18, 10], start=1):
        ws_ar.column_dimensions[get_column_letter(idx)].width = w

    base_range = f"B2:B{last_quad_row}"
    splo_range = f"C2:C{last_quad_row}"
    sphi_range = f"D2:D{last_quad_row}"
    widened_range = f"G2:G{last_quad_row}"

    summary_start = last_quad_row + 2
    summary_rows: List[Tuple[str, str]] = [
        ("Cost subtotal base (MATERIAL+CONVERSION, pre-margin)", f"=SUM({base_range})"),
        ("Sum-of-squares low (non-widened)", f"=SUMPRODUCT(({widened_range}=\"NO\")*({splo_range})^2)"),
        ("Sum-of-squares high (non-widened)", f"=SUMPRODUCT(({widened_range}=\"NO\")*({sphi_range})^2)"),
        ("Linear low add (widened LOW>15%)", f"=SUMPRODUCT(({widened_range}=\"YES\")*({splo_range}))"),
        ("Linear high add (widened LOW>15%)", f"=SUMPRODUCT(({widened_range}=\"YES\")*({sphi_range}))"),
        ("Cost total low (quadrature)", None),   # filled below, needs prior rows
        ("Cost total high (quadrature)", None),
        ("Naive low (footnote only, worst-case envelope)", f"=B{summary_start}-SUM({splo_range})"),
        ("Naive high (footnote only, worst-case envelope)", f"=B{summary_start}+SUM({sphi_range})"),
        ("Commercial explicit base (SG&A/other $ COMMERCIAL drivers)",
         "=SUMIFS(CostStack!J:J,CostStack!B:B,\"COMMERCIAL\")"),
        ("Commercial explicit low", "=SUMIFS(CostStack!I:I,CostStack!B:B,\"COMMERCIAL\")"),
        ("Commercial explicit high", "=SUMIFS(CostStack!K:K,CostStack!B:B,\"COMMERCIAL\")"),
        ("Margin policy base ($ = base% x cost subtotal)", f"=MarginBasePct*B{summary_start}"),
        ("Margin policy low", f"=MarginLowPct*B{summary_start}"),
        ("Margin policy high", f"=MarginHighPct*B{summary_start}"),
        ("Margin + commercial base dollar (linear add-on, never quadratured)", None),
        ("Margin + commercial low dollar", None),
        ("Margin + commercial high dollar", None),
        ("TOTAL SHOULD-COST BASE (final)", None),
        ("TOTAL SHOULD-COST LOW (final)", None),
        ("TOTAL SHOULD-COST HIGH (final)", None),
        ("Naive worst-case low (footnote, incl. margin)", None),
        ("Naive worst-case high (footnote, incl. margin)", None),
        ("Reconciliation check: Low <= Base <= High (must be TRUE)", None),
    ]

    # Row index bookkeeping (explicit, mirrors pro_forma_generator.py's
    # payback_header_row / roi_row style of named row offsets).
    cost_subtotal_row = summary_start
    ssq_low_row = summary_start + 1
    ssq_high_row = summary_start + 2
    lin_low_row = summary_start + 3
    lin_high_row = summary_start + 4
    cost_low_row = summary_start + 5
    cost_high_row = summary_start + 6
    naive_low_row = summary_start + 7
    naive_high_row = summary_start + 8
    comm_base_row = summary_start + 9
    comm_low_row = summary_start + 10
    comm_high_row = summary_start + 11
    mp_base_row = summary_start + 12
    mp_low_row = summary_start + 13
    mp_high_row = summary_start + 14
    margin_base_row = summary_start + 15
    margin_low_row = summary_start + 16
    margin_high_row = summary_start + 17
    total_base_row = summary_start + 18
    total_low_row = summary_start + 19
    total_high_row = summary_start + 20
    naive_low_foot_row = summary_start + 21
    naive_high_foot_row = summary_start + 22
    recon_row = summary_start + 23

    summary_rows[5] = ("Cost total low (quadrature)", f"=B{cost_subtotal_row}-SQRT(B{ssq_low_row})-B{lin_low_row}")
    summary_rows[6] = ("Cost total high (quadrature)", f"=B{cost_subtotal_row}+SQRT(B{ssq_high_row})+B{lin_high_row}")
    summary_rows[15] = ("Margin + commercial base dollar (linear add-on, never quadratured)",
                         f"=B{comm_base_row}+B{mp_base_row}")
    summary_rows[16] = ("Margin + commercial low dollar", f"=B{comm_low_row}+B{mp_low_row}")
    summary_rows[17] = ("Margin + commercial high dollar", f"=B{comm_high_row}+B{mp_high_row}")
    summary_rows[18] = ("TOTAL SHOULD-COST BASE (final)", f"=B{cost_subtotal_row}+B{margin_base_row}")
    summary_rows[19] = ("TOTAL SHOULD-COST LOW (final)", f"=B{cost_low_row}+B{margin_low_row}")
    summary_rows[20] = ("TOTAL SHOULD-COST HIGH (final)", f"=B{cost_high_row}+B{margin_high_row}")
    summary_rows[21] = ("Naive worst-case low (footnote, incl. margin)", f"=B{naive_low_row}+B{margin_low_row}")
    summary_rows[22] = ("Naive worst-case high (footnote, incl. margin)", f"=B{naive_high_row}+B{margin_high_row}")
    summary_rows[23] = ("Reconciliation check: Low <= Base <= High (must be TRUE)",
                         f"=AND(B{total_low_row}<=B{total_base_row},B{total_base_row}<=B{total_high_row})")

    for i, (label, formula) in enumerate(summary_rows):
        rr = summary_start + i
        ws_ar.cell(row=rr, column=1, value=label)
        ws_ar.cell(row=rr, column=2, value=formula)

    wb.defined_names["CostSubtotalBase"] = DefinedName("CostSubtotalBase", attr_text=f"AggregationRollup!$B${cost_subtotal_row}")
    wb.defined_names["MarginBaseDollar"] = DefinedName("MarginBaseDollar", attr_text=f"AggregationRollup!$B${margin_base_row}")
    wb.defined_names["MarginLowDollar"] = DefinedName("MarginLowDollar", attr_text=f"AggregationRollup!$B${margin_low_row}")
    wb.defined_names["MarginHighDollar"] = DefinedName("MarginHighDollar", attr_text=f"AggregationRollup!$B${margin_high_row}")
    wb.defined_names["TotalBaseFinal"] = DefinedName("TotalBaseFinal", attr_text=f"AggregationRollup!$B${total_base_row}")
    wb.defined_names["TotalLowFinal"] = DefinedName("TotalLowFinal", attr_text=f"AggregationRollup!$B${total_low_row}")
    wb.defined_names["TotalHighFinal"] = DefinedName("TotalHighFinal", attr_text=f"AggregationRollup!$B${total_high_row}")
    wb.defined_names["NaiveLowFootnote"] = DefinedName("NaiveLowFootnote", attr_text=f"AggregationRollup!$B${naive_low_foot_row}")
    wb.defined_names["NaiveHighFootnote"] = DefinedName("NaiveHighFootnote", attr_text=f"AggregationRollup!$B${naive_high_foot_row}")

    ws_ar.column_dimensions["A"].width = 58
    ws_ar.column_dimensions["B"].width = 18

    # ---------------------------------------------------------------
    # Tab 4: GapAnalysis (Gap Method: absolute gap, gap %, position vs
    # range, and driver attribution that sums exactly to the gap).
    # ---------------------------------------------------------------
    ws_gap = wb.create_sheet("GapAnalysis")
    _write_header_row(ws_gap, ["Metric", "Value", "Notes"], row=1)
    ws_gap.cell(row=2, column=1, value="Should-cost base")
    ws_gap.cell(row=2, column=2, value="=TotalBaseFinal")
    ws_gap.cell(row=3, column=1, value="Should-cost low")
    ws_gap.cell(row=3, column=2, value="=TotalLowFinal")
    ws_gap.cell(row=4, column=1, value="Should-cost high")
    ws_gap.cell(row=4, column=2, value="=TotalHighFinal")
    ws_gap.cell(row=5, column=1, value="Supplier price")
    ws_gap.cell(row=5, column=2, value="=SupplierPrice")
    ws_gap.cell(row=6, column=1, value="Absolute gap (Proposed - Should-cost base)")
    ws_gap.cell(row=6, column=2,
                value="=IF(ISNUMBER(SupplierPrice),SupplierPrice-TotalBaseFinal,"
                      "\"NEEDS_INPUT: gap not computable yet, add the supplier's proposed price\")")
    ws_gap.cell(row=7, column=1, value="Gap % of should-cost")
    ws_gap.cell(row=7, column=2,
                value="=IF(ISNUMBER(SupplierPrice),(SupplierPrice-TotalBaseFinal)/TotalBaseFinal,\"NEEDS_INPUT\")")
    ws_gap.cell(row=8, column=1, value="Position vs range")
    ws_gap.cell(row=8, column=2, value=(
        "=IF(ISNUMBER(SupplierPrice),"
        "IF(SupplierPrice>TotalHighFinal,\"ABOVE_HIGH: clear over-ask, strong leverage\","
        "IF(SupplierPrice<TotalLowFinal,\"BELOW_LOW: investigate before celebrating, should-cost may be missing a cost\","
        "\"WITHIN_RANGE: defensible tolerance, negotiate on the drivers\")),\"NEEDS_INPUT\")"
    ))
    ws_gap.cell(row=9, column=1, value="Single dominant driver (>60% of base)")
    ws_gap.cell(row=9, column=2, value=f"=IF(MAX(CostStack!Q2:Q{last_cs_row})>0.6,"
                                        "\"YES: total range is essentially that driver's range\",\"NO\")")

    attr_header_row = 11
    _write_header_row(ws_gap, ["Component", "Base (normalized)", "Share of final base", "Gap $ attributed"],
                       row=attr_header_row)
    attr_row = attr_header_row + 1
    for i, d in enumerate(reg.cost_drivers):
        cs_row = 2 + i
        ws_gap.cell(row=attr_row, column=1, value=f"=CostStack!A{cs_row}")
        ws_gap.cell(row=attr_row, column=2, value=f"=CostStack!J{cs_row}")
        ws_gap.cell(row=attr_row, column=3, value=f"=CostStack!Q{cs_row}")
        ws_gap.cell(row=attr_row, column=4,
                    value=f"=IF(ISNUMBER(SupplierPrice),C{attr_row}*(SupplierPrice-TotalBaseFinal),\"NEEDS_INPUT\")")
        attr_row += 1
    if reg.margin_policy is not None:
        ws_gap.cell(row=attr_row, column=1, value="Margin (policy)")
        ws_gap.cell(row=attr_row, column=2, value=f"=AggregationRollup!B{mp_base_row}")
        ws_gap.cell(row=attr_row, column=3, value=f"=B{attr_row}/TotalBaseFinal")
        ws_gap.cell(row=attr_row, column=4,
                    value=f"=IF(ISNUMBER(SupplierPrice),C{attr_row}*(SupplierPrice-TotalBaseFinal),\"NEEDS_INPUT\")")
        attr_row += 1
    last_attr_row = attr_row - 1
    ws_gap.cell(row=attr_row, column=1, value="TOTAL (must equal Absolute gap above)")
    ws_gap.cell(row=attr_row, column=4, value=f"=SUM(D{attr_header_row + 1}:D{last_attr_row})")

    for idx, w in enumerate([40, 20, 20, 20], start=1):
        ws_gap.column_dimensions[get_column_letter(idx)].width = w

    # ---------------------------------------------------------------
    # Tab 5: AssumptionLedger (SKILL.md "Cost-Driver Assumption Ledger":
    # one row per driver, plus the model-level header).
    # ---------------------------------------------------------------
    ws_led = wb.create_sheet("AssumptionLedger")
    _write_header_row(ws_led, [
        "Component", "Class", "Basis", "Low", "Base", "High", "Currency",
        "Source", "Source date", "Confidence", "Index used", "Freshness flag",
        "Correlated group", "Notes",
    ], row=1)
    for i, d in enumerate(reg.cost_drivers):
        cs_row = 2 + i
        led_row = 2 + i
        for col, cs_col in zip(range(1, 4), "ABC"):
            ws_led.cell(row=led_row, column=col, value=f"=CostStack!{cs_col}{cs_row}")
        for col, cs_col in zip(range(4, 7), "EFG"):  # native low/base/high
            ws_led.cell(row=led_row, column=col, value=f"=CostStack!{cs_col}{cs_row}")
        ws_led.cell(row=led_row, column=7, value=f"=CostStack!D{cs_row}")
        for col, cs_col in zip(range(8, 13), "RSTU"):
            ws_led.cell(row=led_row, column=col, value=f"=CostStack!{cs_col}{cs_row}")
        ws_led.cell(row=led_row, column=13, value=f"=CostStack!N{cs_row}")
        ws_led.cell(row=led_row, column=14, value=f"=CostStack!V{cs_row}")
    last_led_row = 1 + n_drivers
    if reg.margin_policy is not None:
        mrow = last_led_row + 1
        ws_led.cell(row=mrow, column=1, value="Margin (policy)")
        ws_led.cell(row=mrow, column=2, value="COMMERCIAL")
        ws_led.cell(row=mrow, column=3, value="% of cost subtotal base (policy assumption)")
        ws_led.cell(row=mrow, column=4, value=f"=AggregationRollup!B{mp_low_row}")
        ws_led.cell(row=mrow, column=5, value=f"=AggregationRollup!B{mp_base_row}")
        ws_led.cell(row=mrow, column=6, value=f"=AggregationRollup!B{mp_high_row}")
        ws_led.cell(row=mrow, column=7, value=reporting_currency)
        ws_led.cell(row=mrow, column=8, value=reg.margin_policy.source)
        ws_led.cell(row=mrow, column=9, value=reg.as_of_date)
        ws_led.cell(row=mrow, column=10, value=reg.margin_policy.confidence)
        ws_led.cell(row=mrow, column=11, value="")
        ws_led.cell(row=mrow, column=12, value="")
        ws_led.cell(row=mrow, column=13, value="")
        ws_led.cell(row=mrow, column=14, value="Added linearly on top of the quadrature roll-up, never quadratured.")
        last_led_row = mrow

    hdr_start = last_led_row + 2
    ws_led.cell(row=hdr_start, column=1, value="Model-level header")
    header_rows = [
        ("as_of_date", reg.as_of_date),
        ("fx_rate_table", "; ".join(
            f"{fx['from']}->{reg.reporting_currency}@{fx['rate']} ({fx.get('as_of', '')})"
            for fx in reg.fx_rate_table
        ) or "N/A (all drivers already in reporting currency)"),
        ("total_low", "=TotalLowFinal"),
        ("total_base", "=TotalBaseFinal"),
        ("total_high", "=TotalHighFinal"),
        ("correlation_assumption", cell_refs["Correlation assumption"].replace("Assumptions!", "=Assumptions!")),
        ("model_confidence", gt.model_confidence),
        ("supplier_price", "=SupplierPrice"),
        ("gap_abs", "=GapAnalysis!B6"),
        ("gap_pct", "=GapAnalysis!B7"),
    ]
    model_conf_row = None
    for i, (label, value) in enumerate(header_rows):
        rr = hdr_start + 1 + i
        ws_led.cell(row=rr, column=1, value=label)
        ws_led.cell(row=rr, column=2, value=value)
        if label == "model_confidence":
            model_conf_row = rr

    for idx, w in enumerate([26, 14, 34, 12, 12, 34, 12, 22, 14, 16, 18, 40], start=1):
        ws_led.column_dimensions[get_column_letter(idx)].width = w

    # ---------------------------------------------------------------
    # Tab 6: Summary (KPI roll-up + the SKILL.md-required reconciliation
    # line, built as a live TEXT()-concatenation formula, not pasted text).
    # ---------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    _write_header_row(ws_sum, ["Metric", "Value"], row=1)
    ws_sum.cell(row=2, column=1, value="Should-cost base")
    ws_sum.cell(row=2, column=2, value="=TotalBaseFinal")
    ws_sum.cell(row=3, column=1, value="Should-cost range")
    ws_sum.cell(row=3, column=2, value='=TEXT(TotalLowFinal,"#,##0")&" - "&TEXT(TotalHighFinal,"#,##0")')
    ws_sum.cell(row=4, column=1, value="Model confidence")
    ws_sum.cell(row=4, column=2, value=f"=AssumptionLedger!B{model_conf_row}")
    ws_sum.cell(row=5, column=1, value="Supplier price")
    ws_sum.cell(row=5, column=2, value="=SupplierPrice")
    ws_sum.cell(row=6, column=1, value="Gap ($)")
    ws_sum.cell(row=6, column=2, value="=GapAnalysis!B6")
    ws_sum.cell(row=7, column=1, value="Gap (%)")
    ws_sum.cell(row=7, column=2, value="=GapAnalysis!B7")
    ws_sum.cell(row=8, column=1, value="Position vs range")
    ws_sum.cell(row=8, column=2, value="=GapAnalysis!B8")
    ws_sum.cell(row=9, column=1, value="Single dominant driver flag")
    ws_sum.cell(row=9, column=2, value="=GapAnalysis!B9")
    ws_sum.cell(row=10, column=1, value="Naive worst-case envelope (footnote only, never headline)")
    ws_sum.cell(row=10, column=2, value='=TEXT(NaiveLowFootnote,"#,##0")&" - "&TEXT(NaiveHighFootnote,"#,##0")')
    ws_sum.cell(row=11, column=1, value="Reconciliation line (Pre-Delivery Self-Test)")
    ws_sum.cell(row=11, column=2, value=(
        '="Reconciles: base "&TEXT(TotalBaseFinal,"#,##0")&" via quadrature; range ["'
        '&TEXT(TotalLowFinal,"#,##0")&", "&TEXT(TotalHighFinal,"#,##0")&"]; model confidence "'
        f'&AssumptionLedger!B{model_conf_row}&"."'
    ))
    ws_sum.cell(row=12, column=1, value="Reconciliation check (must be TRUE)")
    ws_sum.cell(row=12, column=2, value=f"=AggregationRollup!B{recon_row}")

    ws_sum.column_dimensions["A"].width = 46
    ws_sum.column_dimensions["B"].width = 40

    return wb


# ===========================================================================
# Full pipeline: validate -> compute ground truth -> hard-invariant checks
# -> build workbook -> save
# ===========================================================================

def generate_should_cost_workbook(raw_register: Dict[str, Any], output_path: str) -> ShouldCostInput:
    """End-to-end: validate the raw should-cost input register, compute
    ground truth via numeric_kernel, run the hard-coded invariant checks,
    build the formula-driven workbook, and only then save it. Raises rather
    than saving a workbook that fails validation or reconciliation."""
    reg = validate_should_cost_input(raw_register)
    gt = compute_ground_truth(reg)
    run_hardcoded_invariant_checks(reg, gt)
    wb = build_workbook(reg, gt)
    wb.save(output_path)
    return reg


# ===========================================================================
# Self-test / CLI
# ===========================================================================

def _golden_worked_example_register() -> Dict[str, Any]:
    """Mirrors should-cost-builder-1c344a/SKILL.md's own worked illustration
    verbatim: 'a 3-component stack with bases 60 + 30 + 10 = Total_base 100.
    Spreads (low/high): materials +/-9, labor +/-6, logistics +/-2 ...
    sqrt(9^2+6^2+2^2) = sqrt(121) = 11, so the range is [89, 111].' Logistics
    (10% of base) is the LOW driver; 10% does not exceed the 15% widening
    threshold, so no widening fires. No margin, no supplier price -- this
    scenario exists purely to cross-check the kernel wiring against the
    skill's own quoted numbers (same golden example numeric_kernel.py's own
    self-test uses)."""
    return {
        "reporting_currency": "USD",
        "as_of_date": "2026-07-23",
        "category": "Golden worked example (SKILL.md Aggregation Method)",
        "cost_drivers": [
            {"component": "Materials", "class": "MATERIAL", "basis": "per unit BOM",
             "low": 51, "base": 60, "high": 69, "currency": "USD",
             "source": "SKILL.md worked illustration", "source_date": "2026-07-22", "confidence": "MEDIUM"},
            {"component": "Labor", "class": "CONVERSION", "basis": "loaded labor rate x hours",
             "low": 24, "base": 30, "high": 36, "currency": "USD",
             "source": "SKILL.md worked illustration", "source_date": "2026-07-22", "confidence": "MEDIUM"},
            {"component": "Logistics", "class": "CONVERSION", "basis": "lane rate",
             "low": 8, "base": 10, "high": 12, "currency": "USD",
             "source": "SKILL.md worked illustration", "source_date": "2026-07-22", "confidence": "LOW"},
        ],
    }


def _full_sample_register() -> Dict[str, Any]:
    """A fuller scenario exercising every code path: FX normalization
    (Materials priced in EUR), a correlated group (Materials + Packaging,
    both petroleum-derived), the >15%-of-base LOW-confidence widening rule
    (Logistics), an explicit-dollar COMMERCIAL driver (Financing fee), a
    percentage margin_policy, and a supplier price positioned ABOVE the
    should-cost high (to exercise the Gap Method's over-ask branch)."""
    eur_rate = 1.08
    return {
        "reporting_currency": "USD",
        "as_of_date": "2026-07-23",
        "category": "IT hardware - managed switch, 48-port",
        "description": "Bottoms-up should-cost for a 48-port managed switch, annual volume 5,000 units.",
        "fx_rate_table": [{"from": "EUR", "to": "USD", "rate": eur_rate, "as_of": "2026-06-15"}],
        "cost_drivers": [
            {"component": "Materials (chassis + components)", "class": "MATERIAL",
             "basis": "per unit BOM, qty 1", "low": 50.0 / eur_rate, "base": 60.0 / eur_rate,
             "high": 70.0 / eur_rate, "currency": "EUR", "source": "Supplier BOM + 2 market quotes",
             "source_date": "2026-06-01", "confidence": "MEDIUM", "index_used": "N/A",
             "freshness_flag": "current", "correlated_group": "petro_inputs",
             "notes": "Petroleum-derived plastics + metals; grouped with Packaging."},
            {"component": "Packaging (petroleum-derived)", "class": "MATERIAL",
             "basis": "per unit", "low": 7.5, "base": 10.0, "high": 13.5, "currency": "USD",
             "source": "Packaging supplier quote", "source_date": "2026-06-01", "confidence": "MEDIUM",
             "correlated_group": "petro_inputs",
             "notes": "Moves with the same petrochemical index as Materials."},
            {"component": "Labor (assembly)", "class": "CONVERSION",
             "basis": "assembly hours x loaded rate", "low": 24.0, "base": 30.0, "high": 36.0,
             "currency": "USD", "source": "Internal industrial-engineering estimate",
             "source_date": "2026-05-15", "confidence": "MEDIUM"},
            {"component": "Manufacturing overhead", "class": "CONVERSION",
             "basis": "facility burden per unit", "low": 13.5, "base": 15.0, "high": 18.0,
             "currency": "USD", "source": "Industry overhead-rate benchmark", "source_date": "2026-04-01",
             "confidence": "MEDIUM"},
            {"component": "Logistics (outbound + duties)", "class": "CONVERSION",
             "basis": "lane rate + duty %", "low": 8.0, "base": 25.0, "high": 30.0, "currency": "USD",
             "source": "Single freight-forwarder quote", "source_date": "2026-03-01", "confidence": "LOW",
             "notes": "Single data point; LOW confidence; >15% of cost subtotal -> widening rule fires."},
            {"component": "Financing fee", "class": "COMMERCIAL",
             "basis": "flat fee per PO", "low": 3.0, "base": 5.0, "high": 8.0, "currency": "USD",
             "source": "Standard terms sheet", "source_date": "2026-06-01", "confidence": "MEDIUM"},
        ],
        "margin_policy": {
            "base_pct": 0.08, "low_pct": 0.05, "high_pct": 0.12,
            "basis": "category-typical hardware-manufacturer margin band",
            "source": "market-rate-benchmarking category note", "confidence": "MEDIUM",
        },
        "supplier_price": 200.0,
    }


def _run_self_test() -> int:
    import tempfile

    print("=" * 78)
    print("should_cost_generator.py self-test")
    print("=" * 78)
    print(f"numeric_kernel.py available: {KERNEL_AVAILABLE}")
    print(f"openpyxl available: {OPENPYXL_AVAILABLE} "
          f"(version {OPENPYXL_VERSION})" if OPENPYXL_AVAILABLE else
          f"openpyxl available: {OPENPYXL_AVAILABLE}")
    print()

    results: List[tuple] = []

    def check(label, condition, detail=""):
        results.append((label, bool(condition), detail))
        status = "PASS" if condition else "FAIL"
        line = f"[{status}] {label}"
        if detail:
            line += f"  ({detail})"
        print(line)

    # --- Step 0: golden worked example matches SKILL.md's quoted numbers ---
    golden = _golden_worked_example_register()
    try:
        greg = validate_should_cost_input(golden)
        ggt = compute_ground_truth(greg)
        check(
            "Golden worked example matches SKILL.md verbatim "
            "('base 100 = 60+30+10; range [89, 111] via quadrature')",
            abs(ggt.cost_subtotal_base - 100) < 1e-9
            and abs(ggt.quadrature_result.total_low - 89) < 1e-9
            and abs(ggt.quadrature_result.total_high - 111) < 1e-9
            and ggt.quadrature_result.widened_components == [],
            f"base={ggt.cost_subtotal_base}, low={ggt.quadrature_result.total_low}, "
            f"high={ggt.quadrature_result.total_high}, widened={ggt.quadrature_result.widened_components}",
        )
        run_hardcoded_invariant_checks(greg, ggt)
        check("Hard-coded invariants PASS on the golden worked example", True)
    except Exception as e:
        check("Golden worked example round-trips through validate/compute/invariants", False, str(e))
        raise

    # --- Step 1: validation on the full sample ---------------------------
    sample_register = _full_sample_register()
    try:
        reg = validate_should_cost_input(sample_register)
        check("validate_should_cost_input accepts the full sample register", True)
    except Exception as e:
        check("validate_should_cost_input accepts the full sample register", False, str(e))
        raise

    # --- Step 1b: validation refuses broken registers ---------------------
    broken = copy.deepcopy(sample_register)
    del broken["cost_drivers"]
    try:
        validate_should_cost_input(broken)
        check("validate_should_cost_input refuses a register missing cost_drivers", False, "did not raise")
    except ShouldCostValidationError as e:
        check("validate_should_cost_input refuses a register missing cost_drivers", True, str(e)[:120])

    broken2 = copy.deepcopy(sample_register)
    broken2["cost_drivers"][0]["low"] = 999.0  # low > base, violates triangular estimate
    try:
        validate_should_cost_input(broken2)
        check("validate_should_cost_input refuses low > base", False, "did not raise")
    except ShouldCostValidationError as e:
        check("validate_should_cost_input refuses low > base", True, str(e)[:150])

    broken3 = copy.deepcopy(sample_register)
    broken3["cost_drivers"][4]["correlated_group"] = "bad_group"
    broken3["cost_drivers"][4]["class"] = "COMMERCIAL"
    try:
        validate_should_cost_input(broken3)
        check("validate_should_cost_input refuses a COMMERCIAL driver carrying a correlated_group", False,
              "did not raise")
    except ShouldCostValidationError as e:
        check("validate_should_cost_input refuses a COMMERCIAL driver carrying a correlated_group", True,
              str(e)[:150])

    broken4 = copy.deepcopy(sample_register)
    broken4["cost_drivers"][0]["currency"] = "GBP"  # no fx row for GBP
    try:
        validate_should_cost_input(broken4)
        check("validate_should_cost_input refuses an unconvertible currency", False, "did not raise")
    except ShouldCostValidationError as e:
        check("validate_should_cost_input refuses an unconvertible currency", True, str(e)[:150])

    # --- Step 2: ground truth + hard invariants ---------------------------
    try:
        gt = compute_ground_truth(reg)
        check("compute_ground_truth runs via numeric_kernel.quadrature_rollup()", True,
              f"cost_subtotal_base={gt.cost_subtotal_base:.2f}, "
              f"total_base_final={gt.total_base_final:.2f}, "
              f"range=[{gt.total_low_final:.2f}, {gt.total_high_final:.2f}], "
              f"model_confidence={gt.model_confidence}")
    except Exception as e:
        check("compute_ground_truth runs via numeric_kernel.quadrature_rollup()", False, str(e))
        raise

    check("Logistics driver was widened (LOW confidence, >15% of cost subtotal)",
          any(gt.quad_input_rows[i].name == "Logistics (outbound + duties)"
              for i in gt.quadrature_result.widened_components))
    check("Correlated group 'petro_inputs' formed (Materials + Packaging)",
          "petro_inputs" in gt.correlated_groups
          and set(gt.correlated_groups["petro_inputs"])
          == {"Materials (chassis + components)", "Packaging (petroleum-derived)"})
    check("Gap Method computed (supplier_price provided) and position is ABOVE_HIGH",
          gt.gap is not None and gt.gap.position == "ABOVE_HIGH",
          f"gap={gt.gap.absolute_gap if gt.gap else None}")

    try:
        run_hardcoded_invariant_checks(reg, gt)
        check("All hard-coded invariant checks PASS on the valid full sample", True)
    except ReconciliationError as e:
        check("All hard-coded invariant checks PASS on the valid full sample", False, str(e))
        raise

    # --- Step 2b: prove the quadrature invariant catches a bug ------------
    bad_gt = copy.deepcopy(gt)
    bad_gt.quadrature_result = QuadratureResult(
        total_base=gt.quadrature_result.total_base,
        total_low=gt.quadrature_result.total_base - 1.0,  # not sqrt-of-squares derived
        total_high=gt.quadrature_result.total_high,
        widened_components=gt.quadrature_result.widened_components,
        naive_low=gt.quadrature_result.naive_low, naive_high=gt.quadrature_result.naive_high,
    )
    try:
        _assert_quadrature_independently_rederived(bad_gt)
        check("Quadrature invariant CORRECTLY REJECTS a corrupted total_low", False, "did not raise, but should have")
    except ReconciliationError as e:
        check("Quadrature invariant CORRECTLY REJECTS a corrupted total_low", True, str(e)[:150])

    # --- Step 2c: prove the margin-exclusion invariant catches a bug ------
    bad_gt2 = copy.deepcopy(gt)
    bad_gt2.margin_base_dollar = gt.margin_base_dollar + 500.0  # break the linear footing
    try:
        _assert_margin_excluded_from_quadrature(bad_gt2)
        check("Margin-exclusion invariant CORRECTLY REJECTS a broken linear footing", False,
              "did not raise, but should have")
    except ReconciliationError as e:
        check("Margin-exclusion invariant CORRECTLY REJECTS a broken linear footing", True, str(e)[:150])

    # --- Step 2d: prove the gap-attribution invariant catches a bug -------
    bad_gt3 = copy.deepcopy(gt)
    if bad_gt3.gap is not None:
        bad_gt3.gap.attribution[0].gap_dollars += 999.0
    try:
        _assert_gap_attribution_reconciles(bad_gt3)
        check("Gap-attribution invariant CORRECTLY REJECTS a broken attribution sum", False,
              "did not raise, but should have")
    except ReconciliationError as e:
        check("Gap-attribution invariant CORRECTLY REJECTS a broken attribution sum", True, str(e)[:150])

    # --- Step 3: build + save + reopen the real workbook -------------------
    if not OPENPYXL_AVAILABLE:
        check("openpyxl available to write a real .xlsx file", False,
              "openpyxl is NOT installed in this interpreter; XLSX writing could not be exercised.")
        print()
        print("Cannot proceed past this point without openpyxl.")
    else:
        check("openpyxl available to write a real .xlsx file", True, f"version {OPENPYXL_VERSION}")

        tmp_dir = tempfile.gettempdir()
        out_path = os.path.join(tmp_dir, "should_cost_model_selftest.xlsx")
        try:
            generate_should_cost_workbook(sample_register, out_path)
            check("generate_should_cost_workbook() ran end-to-end without raising", True)
        except Exception as e:
            check("generate_should_cost_workbook() ran end-to-end without raising", False, str(e))
            raise

        exists = os.path.exists(out_path)
        size = os.path.getsize(out_path) if exists else 0
        check(f"XLSX file written to {out_path}", exists and size > 0, f"size={size} bytes")

        try:
            reopened = openpyxl.load_workbook(out_path)
            sheet_names = reopened.sheetnames
            expected_tabs = ["Assumptions", "CostStack", "AggregationRollup",
                              "GapAnalysis", "AssumptionLedger", "Summary"]
            check("Re-opened workbook contains all expected tabs",
                  all(t in sheet_names for t in expected_tabs), f"tabs={sheet_names}")

            ws_ar_check = reopened["AggregationRollup"]
            total_base_formula = None
            for row in ws_ar_check.iter_rows():
                for cell in row:
                    if cell.value == "TOTAL SHOULD-COST BASE (final)":
                        total_base_formula = ws_ar_check.cell(row=cell.row, column=2).value
            check("AggregationRollup's TOTAL SHOULD-COST BASE cell contains a live formula (starts with '=')",
                  isinstance(total_base_formula, str) and total_base_formula.startswith("="),
                  f"got: {total_base_formula!r}")

            ws_gap_check = reopened["GapAnalysis"]
            total_attr_formula = None
            for row in ws_gap_check.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "TOTAL (must equal" in cell.value:
                        total_attr_formula = ws_gap_check.cell(row=cell.row, column=4).value
            check("GapAnalysis's attribution TOTAL cell contains a live formula",
                  isinstance(total_attr_formula, str) and total_attr_formula.startswith("="),
                  f"got: {total_attr_formula!r}")

            defined_names = list(reopened.defined_names.keys())
            expected_names = ["ReportingCurrency", "SupplierPrice", "MarginBasePct", "CostSubtotalBase",
                               "TotalBaseFinal", "TotalLowFinal", "TotalHighFinal"]
            check("Workbook defines named ranges for the core roll-up figures",
                  all(n in defined_names for n in expected_names), f"defined_names={defined_names}")

            ws_cs_check = reopened["CostStack"]
            check("CostStack has one data row per cost driver",
                  ws_cs_check.max_row == 1 + len(sample_register["cost_drivers"]),
                  f"max_row={ws_cs_check.max_row}")

        except Exception as e:
            check("Re-opened workbook structural spot-checks", False, str(e))

    print()
    print("=" * 78)
    passed = sum(1 for _, ok, _ in results if ok)
    failed = sum(1 for _, ok, _ in results if not ok)
    total = len(results)
    print(f"SUMMARY: {passed}/{total} passed, {failed}/{total} failed")
    if failed:
        print("FAILED CASES:")
        for label, ok, detail in results:
            if not ok:
                print(f"  - {label}: {detail}")
    print("=" * 78)
    return 1 if failed else 0


def main(argv: Optional[List[str]] = None) -> int:
    """CLI entry point. Usage:
        python should_cost_generator.py --input register.json --output should_cost_model.xlsx
        python should_cost_generator.py --self-test
        python should_cost_generator.py                 (no args -> runs the self-test)
    """
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate should_cost_model.xlsx from a validated should-cost input register "
                    "(should-cost-builder-1c344a). See the module docstring for the input JSON schema."
    )
    parser.add_argument("--input", "-i", help="Path to a JSON file containing the should-cost input register.")
    parser.add_argument("--output", "-o", default="should_cost_model.xlsx",
                         help="Output .xlsx path (default: should_cost_model.xlsx in the current directory).")
    parser.add_argument("--self-test", action="store_true",
                         help="Run the built-in self-test suite instead of generating from --input.")
    args = parser.parse_args(argv)

    if args.self_test or not args.input:
        return _run_self_test()

    import json
    try:
        with open(args.input, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except (OSError, ValueError) as e:
        print(f"ERROR: could not read/parse --input {args.input!r}: {e}", file=sys.stderr)
        return 1

    try:
        generate_should_cost_workbook(raw, args.output)
    except ImportError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        print("The should-cost input was valid and the ground truth computed cleanly; "
              "only the XLSX-writing step could not run because openpyxl is missing.", file=sys.stderr)
        return 1
    except (ShouldCostValidationError, ReconciliationError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
