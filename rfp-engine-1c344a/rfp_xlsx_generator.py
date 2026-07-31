#!/usr/bin/env python3
"""
rfp_xlsx_generator.py - real builders for `requirements_matrix.xlsx` and
`pricing_template.xlsx` (E4).

WHY THIS EXISTS
---------------
`SKILL.md:292` and the Outputs table promised a 5-tier response scale with data-validation
dropdowns, conditional formatting and a locked structure. No code produced them, so nothing
guaranteed the dropdowns existed at all: the artifact was described, not built. The DOCX
path already had a real builder (`assets/lilly_rfx_template.js`), which made this the
widest gap between a claim and its implementation in this skill.

The schema is NOT invented here. `references/artifact-schemas.md` sections 3 and 7 own it;
this module implements those sections exactly, including the named hexes.

Same division of labour as the rest of the suite:
  code  owns validation, arithmetic, assembly and invariants
  model owns narrative

WHAT IT REFUSES TO DO
---------------------
  * category weights that do not sum to 100          -> the kernel's WeightSumError
  * duplicate Req_ID                                 -> SchemaError
  * a value outside a controlled vocabulary          -> SchemaError
  * fewer rows or categories than the package needs  -> SchemaError
  * openpyxl missing                                 -> XlsxUnavailableError
Silently normalizing a weight set is specifically forbidden: this skill builds the grid
that evaluation-engine and rfp-response-analysis later score against, so a quiet
renormalization here distorts every downstream ranking with nothing left to show it
happened.
"""
from __future__ import annotations

import os
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    _OPENPYXL = True
except ImportError:                                          # pragma: no cover
    _OPENPYXL = False

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from numeric_kernel import assert_weight_sum                  # noqa: E402  (G11)


class XlsxError(Exception):
    """Base for every refusal."""


class XlsxUnavailableError(XlsxError):
    pass


class SchemaError(XlsxError):
    pass


# --- controlled vocabularies, from references/artifact-schemas.md section 3 -------------

RESPONSE_TIERS = [
    "Meets OOB",
    "Standard Config",
    "Major Config",
    "Vendor Customization",
    "Does Not Meet",
]

# Each tier a distinct, uniquely-named hex. NO GREEN, per the suite status-palette rule:
# a green "Meets OOB" would read as approval of a supplier claim nobody has verified.
TIER_FILLS = {
    "Meets OOB":            ("0F3A85", "FFFFFF"),   # Bold Blue, white text
    "Standard Config":      ("D4E5F7", "000000"),   # Pale Blue
    "Major Config":         ("FFF3CD", "000000"),   # Amber
    "Vendor Customization": ("FFE0CC", "000000"),   # Orange
    "Does Not Meet":        ("FDE8E5", "000000"),   # Pale Red
}

PRIORITIES = ["Must Have", "Should Have", "Nice to Have"]
RESPONSE_FORMATS = ["Meets Scale (5-tier)", "Narrative", "Yes-No", "File Upload"]
SOURCES = ["Business Stakeholder", "Regulatory", "Security", "Landscape Analysis", "Prior RFP"]

BASE_COLUMNS = [
    "Req_ID", "Category", "Subcategory", "Requirement", "Priority",
    "Response_Format", "Evaluation_Weight", "Source",
    "Supplier_Response", "Supplier_Comments", "Draft_Flag",
]
# Structure columns are locked; the two supplier columns are the only unlocked ones.
STRUCTURE_COLUMNS = BASE_COLUMNS[:8]
SUPPLIER_COLUMNS = ["Supplier_Response", "Supplier_Comments"]
OPTIONAL_COLUMNS = ["Dependencies", "Amendment_Ref"]

HEADER_FILL = "D9D9D9"

# Minimums from the schema: Full = 20 rows / 4 categories, Brief = 10 / 3.
PACKAGE_MINIMUMS = {"full": (20, 4), "brief": (10, 3)}

PRICING_TABS = [
    "Instructions", "Commercial_Summary", "License_Subscription",
    "Implementation_Services", "Volume_Scenarios", "Assumptions", "Exclusions",
]


def _require_openpyxl():
    if not _OPENPYXL:
        raise XlsxUnavailableError(
            "openpyxl is not importable, so the workbook cannot be built. Refusing to "
            "emit a CSV and call it requirements_matrix.xlsx: the data validation and "
            "conditional formatting ARE the artifact, and a CSV silently carries neither."
        )


def validate_requirements(rows, package="full"):
    """Every check that must pass before a workbook is worth writing. Returns the set of
    optional columns that are actually populated."""
    pkg = (package or "full").strip().lower()
    if pkg not in PACKAGE_MINIMUMS:
        raise SchemaError("unknown package %r; expected one of %s"
                          % (package, sorted(PACKAGE_MINIMUMS)))
    if not rows:
        raise SchemaError("no requirements supplied; there is no matrix to build")

    min_rows, min_cats = PACKAGE_MINIMUMS[pkg]
    if len(rows) < min_rows:
        raise SchemaError(
            "a %s package needs at least %d requirements, got %d. Refusing to pad the "
            "matrix with invented rows to reach the minimum." % (pkg, min_rows, len(rows))
        )

    seen = set()
    by_category = {}
    for i, r in enumerate(rows):
        rid = (r.get("Req_ID") or "").strip()
        if not rid:
            raise SchemaError("row %d has no Req_ID" % i)
        if rid in seen:
            raise SchemaError(
                "duplicate Req_ID %r. Addenda and the Q&A consolidation path both key off "
                "Req_ID, so a duplicate silently amends the wrong requirement." % rid
            )
        seen.add(rid)

        for field, vocab in (("Priority", PRIORITIES),
                             ("Response_Format", RESPONSE_FORMATS),
                             ("Source", SOURCES)):
            val = (r.get(field) or "").strip()
            if val not in vocab:
                raise SchemaError(
                    "row %s has %s=%r, which is outside the controlled vocabulary %s"
                    % (rid, field, val, vocab)
                )

        cat = (r.get("Category") or "").strip()
        if not cat:
            raise SchemaError("row %s has no Category" % rid)
        w = r.get("Evaluation_Weight")
        if w is None:
            raise SchemaError(
                "row %s has no Evaluation_Weight. Refusing to treat it as 0, which would "
                "drop the requirement out of scoring without saying so." % rid
            )
        by_category.setdefault(cat, {})[rid] = float(w)

    if len(by_category) < min_cats:
        raise SchemaError(
            "a %s package needs at least %d categories, got %d (%s)"
            % (pkg, min_cats, len(by_category), sorted(by_category))
        )

    # G11 HARD RULE: the kernel does this, not a hand-rolled sum. Once per category.
    for cat, weights in sorted(by_category.items()):
        assert_weight_sum(weights, expected=100.0)

    # Optional columns appear ONLY when populated, so the locked structure column count
    # does not change for runs that never used the synthesizer or issued an addendum.
    present = [c for c in OPTIONAL_COLUMNS
               if any((r.get(c) or "").strip() for r in rows)]
    return present


def build_requirements_matrix(rows, path, package="full"):
    """Write requirements_matrix.xlsx. Validates first; writes nothing if validation fails."""
    _require_openpyxl()
    optional = validate_requirements(rows, package)
    columns = BASE_COLUMNS + optional

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    struct_fill = PatternFill("solid", fgColor=HEADER_FILL)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.protection = Protection(locked=True)

    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(columns, start=1):
            val = row.get(name)
            if name == "Draft_Flag":
                val = bool(val)
            cell = ws.cell(row=r, column=c, value=val)
            if name in SUPPLIER_COLUMNS:
                cell.protection = Protection(locked=False)
                cell.fill = white_fill
                if name == "Supplier_Comments":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.protection = Protection(locked=True)
                cell.fill = struct_fill

    last_row = len(rows) + 1
    resp_idx = columns.index("Supplier_Response") + 1
    resp_col = get_column_letter(resp_idx)
    comm_col = get_column_letter(columns.index("Supplier_Comments") + 1)

    # The 5-tier dropdown. This is the thing SKILL.md promised and nothing produced.
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(RESPONSE_TIERS),
        allow_blank=True,
        showDropDown=False,          # openpyxl inverts this: False MEANS show the dropdown
        showErrorMessage=True,
        errorTitle="Not a valid response",
        error="Choose one of the five response tiers.",
    )
    ws.add_data_validation(dv)
    dv.add("%s2:%s%d" % (resp_col, resp_col, last_row))

    # Conditional formatting, one rule per tier, exact-match on the tier text.
    for tier, (bg, fg) in TIER_FILLS.items():
        ws.conditional_formatting.add(
            "%s2:%s%d" % (resp_col, resp_col, last_row),
            FormulaRule(
                formula=['EXACT($%s2,"%s")' % (resp_col, tier)],
                fill=PatternFill("solid", fgColor=bg),
                font=Font(color=fg),
                stopIfTrue=False,
            ),
        )

    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(columns)), last_row)
    ws.freeze_panes = "A2"
    ws.column_dimensions[comm_col].width = 60      # the schema's "minimum width 300px"
    for c, name in enumerate(columns, start=1):
        if name not in ("Supplier_Comments",):
            ws.column_dimensions[get_column_letter(c)].width = \
                40 if name == "Requirement" else 18

    # Structure locked, supplier columns unlocked. Protection only takes effect on the
    # sheet, so cell-level locked=True is inert without this line.
    ws.protection.sheet = True
    ws.protection.enable()

    wb.save(path)
    return path


def build_pricing_template(path, domain_tabs=None):
    """Write pricing_template.xlsx with the seven required tabs, plus any domain tabs."""
    _require_openpyxl()
    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)

    tabs = list(PRICING_TABS) + [t for t in (domain_tabs or []) if t not in PRICING_TABS]

    first = True
    for name in tabs:
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        first = False
        if name == "Instructions":
            ws["A1"] = "How to complete this template"
            ws["A1"].font = header_font
            for i, line in enumerate([
                "Complete every tab. A blank cell is read as no response, not as zero.",
                "State the currency and the period each figure covers.",
                "Assumptions that affect pricing belong on the Assumptions tab, not in a cell comment.",
                "Anything excluded from the quoted price belongs on the Exclusions tab.",
            ], start=3):
                ws.cell(row=i, column=1, value=line)
            ws.column_dimensions["A"].width = 100
        elif name == "Commercial_Summary":
            for c, h in enumerate(["Cost Element", "Year 1", "Year 2", "Year 3", "Total"], start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
            elements = ["License / Subscription", "Implementation Services", "Support", "Other"]
            for r, el in enumerate(elements, start=2):
                ws.cell(row=r, column=1, value=el)
                ws.cell(row=r, column=5, value="=SUM(B%d:D%d)" % (r, r))
            total_row = len(elements) + 2
            ws.cell(row=total_row, column=1, value="Total").font = header_font
            for c in range(2, 6):
                col = get_column_letter(c)
                ws.cell(row=total_row, column=c,
                        value="=SUM(%s2:%s%d)" % (col, col, total_row - 1))
            ws.column_dimensions["A"].width = 30
        elif name == "Volume_Scenarios":
            for c, h in enumerate(["Scenario", "Volume", "Unit Price", "Annual Cost"], start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for r, s in enumerate(["Base", "+25%", "+50%"], start=2):
                ws.cell(row=r, column=1, value=s)
                ws.cell(row=r, column=4, value="=B%d*C%d" % (r, r))
        else:
            ws["A1"] = name.replace("_", " ")
            ws["A1"].font = header_font
            ws.column_dimensions["A"].width = 40

    wb.save(path)
    return path


def describe(path):
    """Re-open a workbook and report what is actually in it. Used by the self-test so the
    assertion is made against the FILE, not against the code that wrote it."""
    _require_openpyxl()
    wb = load_workbook(path)
    ws = wb["Requirements"] if "Requirements" in wb.sheetnames else wb.active
    return {
        "sheets": wb.sheetnames,
        "validations": [
            {"sqref": str(dv.sqref), "formula1": dv.formula1, "type": dv.type}
            for dv in ws.data_validations.dataValidation
        ],
        "conditional_formats": [
            {"range": str(rng.sqref), "rules": len(rules)}
            for rng, rules in ws.conditional_formatting._cf_rules.items()
        ],
        "protected": bool(ws.protection.sheet),
        "autofilter": str(ws.auto_filter.ref) if ws.auto_filter.ref else None,
        "headers": [c.value for c in ws[1]],
        "rows": ws.max_row - 1,
    }


def main(argv):
    if not _OPENPYXL:
        print("openpyxl is not available; this generator cannot run here.", file=sys.stderr)
        return 2
    print(__doc__.strip())
    print("\nImport build_requirements_matrix() / build_pricing_template(), or run "
          "rfp_xlsx_selftest.py.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
