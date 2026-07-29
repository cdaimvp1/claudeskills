#!/usr/bin/env python3
"""
Self-test for rfp_xlsx_generator.py (E4).

E4's verify clause: "open the generated workbook and assert data validations and
conditional formats are present." So every positive assertion here runs against a workbook
that has been WRITTEN TO DISK AND REOPENED. Asserting against the in-memory object would
only prove the code did what the code did; it would not prove the features survive the
round trip into the file, which is exactly where they are lost in practice.

Run: python rfp_xlsx_selftest.py
"""
from __future__ import annotations

import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from numeric_kernel import WeightSumError                      # noqa: E402
from rfp_xlsx_generator import (                              # noqa: E402
    PRICING_TABS,
    RESPONSE_TIERS,
    SchemaError,
    TIER_FILLS,
    build_pricing_template,
    build_requirements_matrix,
    describe,
    validate_requirements,
)

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl unavailable; cannot run E4 self-test")
    sys.exit(2)

PASS, FAIL = [], []


def ok(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(("  ok   " if cond else "  FAIL ") + name + (("  <- " + detail) if detail and not cond else ""))


def raises(name, exc, fn):
    try:
        fn()
    except exc:
        ok(name, True)
        return
    except Exception as e:                                   # noqa: BLE001
        ok(name, False, "raised %s, expected %s" % (type(e).__name__, exc.__name__))
        return
    ok(name, False, "did not raise " + exc.__name__)


def sample_rows(n_per_cat=5):
    """20 requirements across 4 categories, weights summing to 100 within each."""
    cats = ["Security", "Integration", "Reporting", "Support"]
    rows = []
    i = 1
    for cat in cats:
        # 5 rows at 20% each = 100% per category.
        for k in range(n_per_cat):
            rows.append({
                "Req_ID": "REQ-%03d" % i,
                "Category": cat,
                "Subcategory": "",
                "Requirement": "%s requirement %d." % (cat, k + 1),
                "Priority": "Must Have" if k == 0 else "Should Have",
                "Response_Format": "Meets Scale (5-tier)",
                "Evaluation_Weight": 100.0 / n_per_cat,
                "Source": "Business Stakeholder",
                "Supplier_Response": "",
                "Supplier_Comments": "",
                "Draft_Flag": False,
            })
            i += 1
    return rows


def run():
    print("=" * 84)
    print("rfp_xlsx_generator self-test (assertions run against the REOPENED file)")
    print("=" * 84)

    tmp = tempfile.mkdtemp(prefix="e4_")
    matrix = os.path.join(tmp, "requirements_matrix.xlsx")
    pricing = os.path.join(tmp, "pricing_template.xlsx")

    build_requirements_matrix(sample_rows(), matrix)
    ok("T1  the workbook was actually written", os.path.isfile(matrix))

    d = describe(matrix)

    # --- E4's stated verify: validations + conditional formats present ----------------
    ok("T2  a data validation survives the round trip", len(d["validations"]) >= 1,
       repr(d["validations"]))
    dv = d["validations"][0] if d["validations"] else {}
    ok("T3  it is a list validation", dv.get("type") == "list", repr(dv.get("type")))
    for tier in RESPONSE_TIERS:
        ok("T4  dropdown offers %-21s" % tier, tier in (dv.get("formula1") or ""))
    ok("T5  the dropdown covers every data row, not just the first",
       dv.get("sqref", "").endswith("21"), repr(dv.get("sqref")))

    ok("T6  conditional formatting survives the round trip",
       len(d["conditional_formats"]) >= 1, repr(d["conditional_formats"]))
    total_rules = sum(c["rules"] for c in d["conditional_formats"])
    ok("T7  one conditional rule per tier (5)", total_rules == len(TIER_FILLS),
       "got %d" % total_rules)

    # --- the no-green rule is an accuracy rule, so assert it on the file ---------------
    wb = load_workbook(matrix)
    ws = wb["Requirements"]
    hexes = []
    for _rng, rules in ws.conditional_formatting._cf_rules.items():
        for r in rules:
            if r.dxf is not None and r.dxf.fill is not None:
                # For a solid PatternFill the colour lives in fgColor. bgColor is present
                # but set to '00000000', which is a TRUTHY string, so an `or` here silently
                # reads every rule as black. That bug made the no-green check below pass
                # against an empty list, which is why T8a now asserts the list is non-empty.
                col = r.dxf.fill.fgColor.rgb
                if col:
                    hexes.append(str(col)[-6:].upper())
    expected = {bg.upper() for bg, _fg in TIER_FILLS.values()}
    ok("T8  every tier hex from the schema is present", expected.issubset(set(hexes)),
       "missing %s" % (expected - set(hexes)))
    ok("T8a the hex list is non-empty, so T9 cannot pass vacuously", len(hexes) >= 5,
       "got %d" % len(hexes))

    def greenish(h):
        try:
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        except ValueError:
            return False
        return g > r + 24 and g > b + 24
    ok("T9  NO green fill anywhere in the conditional formats",
       not any(greenish(h) for h in hexes), repr([h for h in hexes if greenish(h)]))

    # --- structure -------------------------------------------------------------------
    ok("T10 sheet protection is on", d["protected"] is True)
    ok("T11 autofilter is set", bool(d["autofilter"]), repr(d["autofilter"]))
    ok("T12 all 11 base columns present, no optional ones", len(d["headers"]) == 11,
       repr(d["headers"]))
    ok("T13 20 requirement rows", d["rows"] == 20, "got %d" % d["rows"])

    hdr = {h: i + 1 for i, h in enumerate(d["headers"])}
    resp = ws.cell(row=2, column=hdr["Supplier_Response"])
    req = ws.cell(row=2, column=hdr["Requirement"])
    comm = ws.cell(row=2, column=hdr["Supplier_Comments"])
    ok("T14 supplier response cells are UNLOCKED", resp.protection.locked is False)
    ok("T15 structure cells are LOCKED", req.protection.locked is True)
    ok("T16 comments column wraps text", comm.alignment.wrap_text is True)

    # --- optional columns appear only when populated ----------------------------------
    rows2 = sample_rows()
    rows2[0]["Dependencies"] = "Depends on: REQ-002"
    m2 = os.path.join(tmp, "with_optional.xlsx")
    build_requirements_matrix(rows2, m2)
    h2 = describe(m2)["headers"]
    ok("T17 a populated optional column IS added", "Dependencies" in h2)
    ok("T18 an unpopulated optional column is NOT added", "Amendment_Ref" not in h2, repr(h2))

    # --- refusals ---------------------------------------------------------------------
    def bad_weights():
        r = sample_rows(); r[0]["Evaluation_Weight"] = 25.0
        validate_requirements(r)
    raises("T19 refuses category weights that do not sum to 100 (kernel, G11)",
           WeightSumError, bad_weights)

    def dupe():
        r = sample_rows(); r[1]["Req_ID"] = r[0]["Req_ID"]
        validate_requirements(r)
    raises("T20 refuses a duplicate Req_ID", SchemaError, dupe)

    def bad_priority():
        r = sample_rows(); r[0]["Priority"] = "Critical"
        validate_requirements(r)
    raises("T21 refuses a Priority outside the vocabulary", SchemaError, bad_priority)

    def bad_format():
        r = sample_rows(); r[0]["Response_Format"] = "Freeform"
        validate_requirements(r)
    raises("T22 refuses a Response_Format outside the vocabulary", SchemaError, bad_format)

    def missing_weight():
        r = sample_rows(); del r[0]["Evaluation_Weight"]
        validate_requirements(r)
    raises("T23 refuses a missing weight (never treats it as 0)", SchemaError, missing_weight)

    def too_few():
        validate_requirements(sample_rows()[:8])
    raises("T24 refuses a Full package below the 20-row minimum", SchemaError, too_few)

    def too_few_cats():
        r = [x for x in sample_rows() if x["Category"] in ("Security", "Integration")]
        validate_requirements(r)
    raises("T25 refuses below the category minimum", SchemaError, too_few_cats)

    raises("T26 refuses an empty requirement set", SchemaError,
           lambda: validate_requirements([]))
    raises("T27 refuses an unknown package name", SchemaError,
           lambda: validate_requirements(sample_rows(), package="deluxe"))

    # A Brief package has different minimums and must still be buildable. Note it needs its
    # OWN well-formed set: slicing rows off a Full set cuts categories mid-way, so the
    # weights stop summing to 100 and the generator correctly refuses. That is the check
    # working, not a bug, and it caught this test's first version.
    brief = [r for r in sample_rows(n_per_cat=4)
             if r["Category"] in ("Security", "Integration", "Reporting")]
    ok("T28 a Brief package passes its own lower minimum",
       len(validate_requirements(brief, package="brief")) == 0)
    raises("T28a that same Brief set is REFUSED as a Full package", SchemaError,
           lambda: validate_requirements(brief, package="full"))

    # --- nothing is written when validation fails --------------------------------------
    never = os.path.join(tmp, "never_written.xlsx")
    try:
        bad = sample_rows(); bad[0]["Req_ID"] = bad[1]["Req_ID"]
        build_requirements_matrix(bad, never)
    except SchemaError:
        pass
    ok("T29 a refused build leaves NO partial file behind", not os.path.exists(never))

    # --- pricing template --------------------------------------------------------------
    build_pricing_template(pricing)
    pwb = load_workbook(pricing)
    ok("T30 pricing template has all seven required tabs",
       all(t in pwb.sheetnames for t in PRICING_TABS),
       repr([t for t in PRICING_TABS if t not in pwb.sheetnames]))
    cs = pwb["Commercial_Summary"]
    formulas = [c.value for row in cs.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    ok("T31 Commercial_Summary is formula-driven, not hardcoded", len(formulas) >= 4,
       "got %d" % len(formulas))
    vs = pwb["Volume_Scenarios"]
    scenarios = [vs.cell(row=r, column=1).value for r in (2, 3, 4)]
    ok("T32 Volume_Scenarios carries Base / +25% / +50%",
       scenarios == ["Base", "+25%", "+50%"], repr(scenarios))

    pwb2 = load_workbook(build_pricing_template(
        os.path.join(tmp, "p2.xlsx"), domain_tabs=["Managed_Services"]))
    ok("T33 a domain-specific tab can be added", "Managed_Services" in pwb2.sheetnames)

    print("=" * 84)
    print("SUMMARY: %d/%d passed, %d failed" % (len(PASS), len(PASS) + len(FAIL), len(FAIL)))
    print("artifacts: %s" % tmp)
    print("=" * 84)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(run())
