#!/usr/bin/env python3
"""
scope_artifacts_generator.py — the four structured artifacts (F9 build 1).

F9 ranked this first: four structured deliverables, zero generators, and the arithmetic
already sitting in the vendored kernel. Everything here was hand-assembled before, which
is the drift case E1/E2 exist to prevent.

  rate_card_and_payment_schedule.xlsx   rate card + milestone schedule, footing verified
  raci_matrix.csv                       RACI grid, orphans flagged not dropped
  change_control_log_template.xlsx      fixed-column register
  scope_findings.json                   findings ledger + the score calculation table

`Rewritten_SOW.docx` is deliberately NOT here: F9 classed it PROSE. A rewritten scope is
argument and specification, not assembly, and generating it would be the wrong kind of
automation. `scope_diagnostic_dashboard.jsx` is deferred to D1.

Division of labour: code owns validation, arithmetic, assembly and invariants; the model
owns narrative.

WHAT IT REFUSES TO DO
---------------------
  * a rate-card row whose rate x quantity does not equal its stated total   -> LineMathError
  * milestone payments that do not sum to the stated contract value        -> the kernel's
                                                                              ReconciliationError
  * a dimension scored above the ceiling its own findings impose           -> SeverityCapError
  * an orphaned RACI deliverable that is not flagged as an open finding    -> OrphanError
  * a weight set that does not sum to 1.0                                  -> the kernel's
                                                                              WeightSumError
The reconciliation rule is the skill's own, from `references/pass-artifacts.md`: "if it
still does not foot, the rewritten SOW carries the same defect it was meant to fix; do not
ship an unreconciled rewrite."

Stdlib plus openpyxl for the two workbooks. CSV and JSON need no third-party code, so a
missing openpyxl still lets those two artifacts build.
"""
from __future__ import annotations

import csv
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from numeric_kernel import (                                   # noqa: E402  (G11)
    ReconciliationError,
    assert_reconciles,
    verify_line_math,
    weighted_score,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:                                            # pragma: no cover
    _OPENPYXL = False


# --- the 10 dimensions, from references/scope-quality-scoring.md ------------------------
DIMENSION_WEIGHTS = {
    "deliverables_definition":   0.15,
    "scope_boundary":            0.10,
    "acceptance_criteria":       0.15,
    "assumptions_dependencies":  0.10,
    "raci_completeness":         0.10,
    "milestones_schedule":       0.10,
    "slas_kpis":                 0.10,
    "staffing_rate_card":        0.10,
    "payment_alignment":         0.05,
    "change_control":            0.05,
}

DIMENSION_LABELS = {
    "deliverables_definition":  "Deliverables Definition & Testability",
    "scope_boundary":           "In-Scope / Out-of-Scope Boundary Clarity",
    "acceptance_criteria":      "Acceptance Criteria Objectivity",
    "assumptions_dependencies": "Assumptions & Dependencies Completeness",
    "raci_completeness":        "Roles & Responsibilities (RACI) Completeness",
    "milestones_schedule":      "Milestones & Schedule Definition",
    "slas_kpis":                "SLAs / KPIs Measurability",
    "staffing_rate_card":       "Staffing & Rate-Card Structure Soundness",
    "payment_alignment":        "Payment-to-Deliverable Alignment",
    "change_control":           "Change-Control Trigger Definition",
}

# Findings drive dimension scores, not the other way round. A severity imposes a CEILING.
SEVERITY_CAPS = {"BLOCKING": 0.9, "HIGH": 3.4, "MEDIUM": 4.4, "LOW": None}

# A rate-card footing failure caps its own dimension at 2.4 regardless of how well the rest
# of the dimension reads (scope-quality-scoring.md).
FOOTING_FAILURE_CAP = 2.4

SCORE_BANDS = [
    (75, 100, "Low gap / Ready to Issue"),
    (50, 74, "Moderate gap / Needs Targeted Fixes"),
    (25, 49, "High gap / Major Rework"),
    (0, 24, "Critical gap / Not Priceable"),
]

RACI_ROLES = ("Responsible", "Accountable", "Consulted", "Informed")

HEADER_FILL = "D9D9D9"


class ScopeArtifactError(Exception):
    """Base for every refusal."""


class LineMathError(ScopeArtifactError):
    pass


class SeverityCapError(ScopeArtifactError):
    pass


class OrphanError(ScopeArtifactError):
    pass


class XlsxUnavailableError(ScopeArtifactError):
    pass


def score_band(score_0_100):
    for lo, hi, label in SCORE_BANDS:
        if lo <= score_0_100 <= hi:
            return label
    raise ScopeArtifactError("score %r is outside 0-100" % score_0_100)


def compute_scope_score(dimension_scores, findings, auto_clamp_findings=None):
    """Composite via the kernel, with the severity ceilings enforced first.

    Returns the calculation table too, because a score without its visible per-dimension
    derivation is invalid per the suite validation checklist. The table is the artifact,
    not a debugging aid.

    Two kinds of finding, treated differently on purpose:

      `findings`             the CALLER's ledger. A score above one of these ceilings is an
                             inconsistency in the caller's own input, so it REFUSES: the
                             score and the ledger disagree and only a human can say which
                             is right.
      `auto_clamp_findings`  defects THIS generator discovered (a rate card that does not
                             foot). The caller could not have reconciled a score against a
                             finding that did not exist yet, so refusing would punish them
                             for the generator's own discovery. These CLAMP the score down
                             to the ceiling and record that they did, visibly, in the
                             calculation table.
    """
    missing = [k for k in DIMENSION_WEIGHTS if k not in dimension_scores]
    if missing:
        raise ScopeArtifactError(
            "dimension score(s) missing: %s. All ten dimensions are scored every run; a "
            "dimension left out is not the same as a dimension scored 0, and guessing "
            "which was meant is exactly what this refuses to do." % missing
        )

    # Worst (lowest) ceiling per dimension, from that dimension's open findings.
    ceilings = {}
    for f in findings:
        if f.get("status", "open").lower() != "open":
            continue
        dim = f.get("dimension")
        sev = (f.get("severity") or "").upper()
        if dim not in DIMENSION_WEIGHTS:
            raise ScopeArtifactError(
                "finding %r is tagged to unknown dimension %r"
                % (f.get("id", "?"), dim)
            )
        if sev not in SEVERITY_CAPS:
            raise ScopeArtifactError(
                "finding %r has severity %r; expected one of %s"
                % (f.get("id", "?"), sev, sorted(SEVERITY_CAPS))
            )
        cap = SEVERITY_CAPS[sev]
        if cap is not None and (dim not in ceilings or cap < ceilings[dim]):
            ceilings[dim] = cap

    for dim, cap in sorted(ceilings.items()):
        if dimension_scores[dim] > cap:
            raise SeverityCapError(
                "dimension %r scored %.1f but an open finding caps it at %.1f. Findings "
                "drive dimension scores; a score above its own findings' ceiling means the "
                "score and the ledger disagree, and the ledger wins."
                % (DIMENSION_LABELS[dim], dimension_scores[dim], cap)
            )

    # Generator-discovered defects clamp rather than refuse. Each carries its own explicit
    # cap where the scoring doc names one (a footing failure caps its dimension at 2.4),
    # falling back to the severity ceiling otherwise.
    effective = dict(dimension_scores)
    clamped = {}
    for f in (auto_clamp_findings or []):
        dim = f.get("dimension")
        if dim not in DIMENSION_WEIGHTS:
            continue
        cap = f.get("cap")
        if cap is None:
            cap = SEVERITY_CAPS.get((f.get("severity") or "").upper())
        if cap is None:
            continue
        if effective[dim] > cap:
            clamped[dim] = {"from": effective[dim], "to": cap, "by": f.get("id")}
            effective[dim] = cap

    composite = weighted_score(effective, DIMENSION_WEIGHTS)   # refuses on bad weights
    score_100 = round(composite * 20)
    table = [
        {
            "dimension": DIMENSION_LABELS[k],
            "weight": DIMENSION_WEIGHTS[k],
            "score": effective[k],
            "score_as_submitted": dimension_scores[k],
            "contribution": round(effective[k] * DIMENSION_WEIGHTS[k], 4),
            "ceiling": ceilings.get(k),
            "clamped": clamped.get(k),
        }
        for k in DIMENSION_WEIGHTS
    ]
    return {
        "composite_0to5": round(composite, 3),
        "score_0to100": score_100,
        "band": score_band(score_100),
        "calculation_table": table,
    }


def verify_rate_card(rows):
    """Every row must foot. Returns the verified line totals."""
    totals = []
    for i, r in enumerate(rows):
        rate = r.get("rate")
        qty = r.get("quantity")
        stated = r.get("line_total")
        for name, v in (("rate", rate), ("quantity", qty), ("line_total", stated)):
            if v is None:
                raise LineMathError(
                    "rate-card row %d (%r) has no %s. Refusing to treat a missing number as "
                    "zero: it would understate the total and the error would be inherited "
                    "by the payment schedule." % (i, r.get("role", "?"), name)
                )
        if not verify_line_math(float(rate), float(qty), float(stated)):
            raise LineMathError(
                "rate-card row %d (%r) does not foot: %s x %s = %.2f, but the stated line "
                "total is %.2f." % (i, r.get("role", "?"), rate, qty,
                                    float(rate) * float(qty), float(stated))
            )
        totals.append(float(stated))
    return totals


def verify_payment_schedule(milestones, contract_value):
    """Milestone payments must sum to the contract value. Uses the kernel, not a hand sum."""
    amounts = []
    for i, m in enumerate(milestones):
        amt = m.get("amount")
        if amt is None:
            raise LineMathError(
                "milestone %d (%r) has no amount." % (i, m.get("name", "?"))
            )
        amounts.append(float(amt))
    # Raises ReconciliationError, which the caller does NOT catch: pass-artifacts.md says
    # an unreconciled rebuild must not ship.
    return assert_reconciles(amounts, float(contract_value), label="payment schedule")


def build_raci(deliverables, findings):
    """RACI rows plus the orphan check.

    pass-artifacts.md: "RACI matrix has zero orphaned deliverables, or each orphan is
    explicitly flagged as an open finding (not silently dropped)". So an orphan is allowed
    to EXIST; what is refused is an orphan nobody recorded.
    """
    flagged = {
        (f.get("deliverable") or "").strip()
        for f in findings
        if f.get("dimension") == "raci_completeness" and f.get("status", "open").lower() == "open"
    }
    rows, orphans = [], []
    for d in deliverables:
        name = (d.get("name") or "").strip()
        if not name:
            raise OrphanError("a deliverable has no name; it cannot be RACI-mapped")
        responsible = (d.get("responsible") or "").strip()
        if not responsible:
            orphans.append(name)
            if name not in flagged:
                raise OrphanError(
                    "deliverable %r has no Responsible party and no open RACI finding "
                    "flagging it. An orphan may exist, but it may not be silently dropped: "
                    "add a finding tagged raci_completeness naming this deliverable." % name
                )
        rows.append({
            "Deliverable": name,
            "Responsible": responsible or "[ORPHAN, no Responsible party]",
            "Accountable": (d.get("accountable") or "").strip(),
            "Consulted": (d.get("consulted") or "").strip(),
            "Informed": (d.get("informed") or "").strip(),
            "Orphan_Flag": "YES" if not responsible else "",
        })
    return rows, orphans


def write_raci_csv(rows, path):
    cols = ["Deliverable"] + list(RACI_ROLES) + ["Orphan_Flag"]
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return path


def write_scope_findings_json(findings, score, path):
    payload = {
        "scope_definition_score": {
            "composite_0to5": score["composite_0to5"],
            "score_0to100": score["score_0to100"],
            "band": score["band"],
            "calculation_table": score["calculation_table"],
        },
        "findings": findings,
        "finding_count": len(findings),
        "clamped_dimensions": [r for r in score["calculation_table"] if r.get("clamped")],
        "note": ("Sidecar mirrors the findings ledger and the score calculation. Every "
                 "finding in the Diagnostic Report appears here; the two are generated "
                 "from the same object so they cannot disagree."),
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)
    return path


def _header(ws, cols):
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        ws.column_dimensions[get_column_letter(c)].width = max(16, min(46, len(name) + 8))


def write_rate_card_xlsx(rows, milestones, contract_value, path):
    """Rate card + payment schedule, with the reconciliation SHOWN, not just asserted."""
    if not _OPENPYXL:
        raise XlsxUnavailableError(
            "openpyxl is unavailable, so the workbook cannot be built. Refusing to emit a "
            "CSV under an .xlsx name."
        )
    line_totals = verify_rate_card(rows)
    verify_payment_schedule(milestones, contract_value)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rate Card"
    _header(ws, ["Role", "Level", "Rate", "Unit", "Quantity", "Line Total",
                 "Footing Check (at build)", "Footing Check (live)"])
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("role"))
        ws.cell(row=i, column=2, value=r.get("level", ""))
        ws.cell(row=i, column=3, value=float(r["rate"]))
        ws.cell(row=i, column=4, value=r.get("unit", "hour"))
        ws.cell(row=i, column=5, value=float(r["quantity"]))
        ws.cell(row=i, column=6, value=float(r["line_total"]))
        # TWO checks, on purpose, because each covers the other's blind spot.
        #
        # Column G is the BUILD-TIME verdict, a static value. openpyxl writes formulas
        # without a cached result, so a live formula reads as None to every programmatic
        # reader (load_workbook(data_only=True), pandas, any non-Excel consumer). A check
        # that is blank unless a human opens Excel is not a check.
        #
        # Column H is the LIVE formula, which column G cannot be: it recomputes if someone
        # edits a rate or a quantity after the build, so a later edit that breaks the
        # footing is visible rather than silently contradicting the frozen verdict.
        ws.cell(row=i, column=7,
                value="OK" if verify_line_math(float(r["rate"]), float(r["quantity"]),
                                               float(r["line_total"])) else "DOES NOT FOOT")
        ws.cell(row=i, column=8, value="=IF(ABS(C%d*E%d-F%d)<0.01,\"OK\",\"DOES NOT FOOT\")"
                % (i, i, i))
    last = len(rows) + 1
    ws.cell(row=last + 1, column=5, value="Total").font = Font(bold=True)
    ws.cell(row=last + 1, column=6, value="=SUM(F2:F%d)" % last).font = Font(bold=True)

    ps = wb.create_sheet("Payment Schedule")
    _header(ps, ["Milestone", "Deliverable", "Trigger / Acceptance Gate", "Amount"])
    for i, m in enumerate(milestones, start=2):
        ps.cell(row=i, column=1, value=m.get("name"))
        ps.cell(row=i, column=2, value=m.get("deliverable", ""))
        ps.cell(row=i, column=3, value=m.get("trigger", ""))
        ps.cell(row=i, column=4, value=float(m["amount"]))
    mlast = len(milestones) + 1
    ps.cell(row=mlast + 1, column=3, value="Total").font = Font(bold=True)
    ps.cell(row=mlast + 1, column=4, value="=SUM(D2:D%d)" % mlast).font = Font(bold=True)
    ps.cell(row=mlast + 2, column=3, value="Stated contract value").font = Font(bold=True)
    ps.cell(row=mlast + 2, column=4, value=float(contract_value))
    ps.cell(row=mlast + 3, column=3, value="Reconciliation (at build)").font = Font(bold=True)
    # Static, for the same reason as the rate card: a formula-only verdict is None to every
    # programmatic reader. build_all() has already refused if this does not reconcile, so
    # this states the verdict rather than deciding it.
    ps.cell(row=mlast + 3, column=4, value="RECONCILES")
    ps.cell(row=mlast + 4, column=3, value="Reconciliation (live)").font = Font(bold=True)
    ps.cell(row=mlast + 4, column=4,
            value="=IF(ABS(D%d-D%d)<0.01,\"RECONCILES\",\"DOES NOT RECONCILE\")"
            % (mlast + 1, mlast + 2))
    ps.cell(row=mlast + 6, column=1,
            value=("Two verdicts on purpose. The build-time row is a static value, so a "
                   "programmatic reader sees it; the live row recomputes in Excel, so a "
                   "later edit that breaks the footing is visible rather than silently "
                   "contradicting the frozen verdict."))
    ps.cell(row=mlast + 6, column=1).alignment = Alignment(wrap_text=True)

    wb.save(path)
    return path, sum(line_totals)


def write_change_control_xlsx(path, triggers=None, drafted=False):
    """Change-control register. Pre-populated with the DRAFT default when none existed."""
    if not _OPENPYXL:
        raise XlsxUnavailableError("openpyxl is unavailable; cannot build the register.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Change Control"
    _header(ws, ["Change_ID", "Raised_Date", "Description", "Trigger Threshold",
                 "Scope Impact", "Schedule Impact", "Cost Impact", "Pricing Mechanism",
                 "Approval Authority", "Status", "Source"])
    for i, t in enumerate(triggers or [], start=2):
        ws.cell(row=i, column=4, value=t.get("threshold", ""))
        ws.cell(row=i, column=8, value=t.get("pricing_mechanism", ""))
        ws.cell(row=i, column=9, value=t.get("approval_authority", ""))
        ws.cell(row=i, column=10, value="OPEN")
        # Provenance: extracted from the SOW, or drafted from the clause library.
        ws.cell(row=i, column=11,
                value="DRAFT - confirm with the requesting stakeholder" if drafted
                else t.get("source", "Extracted from the SOW"))
    wb.save(path)
    return path


def _arithmetic_findings(spec):
    """Run the arithmetic and return any failures AS FINDINGS rather than as exceptions.

    This is the diagnosis path. A supplier's rate card that does not foot is precisely what
    this skill exists to CATCH; refusing to report it would be backwards.

    Not fabrication: each of these is a deterministic, recomputable defect with an exact
    delta, and the finding states the numbers so a reader can check it.
    """
    found = []
    for i, r in enumerate(spec.get("rate_card") or []):
        try:
            verify_rate_card([r])
        except LineMathError as e:
            found.append({
                "id": "GEN-RC-%03d" % i,
                "dimension": "staffing_rate_card",
                "severity": "HIGH",
                "status": "open",
                # scope-quality-scoring.md: a footing failure caps this dimension at 2.4
                # regardless of how well the rest of the dimension reads.
                "cap": FOOTING_FAILURE_CAP,
                "text": str(e),
                "detected_by": "scope_artifacts_generator.verify_rate_card",
            })
    try:
        verify_payment_schedule(spec.get("milestones") or [], spec["contract_value"])
    except (LineMathError, ReconciliationError) as e:
        found.append({
            "id": "GEN-PS-001",
            "dimension": "payment_alignment",
            "severity": "HIGH",
            "status": "open",
            "cap": FOOTING_FAILURE_CAP,
            "text": str(e),
            "detected_by": "scope_artifacts_generator.verify_payment_schedule",
        })
    return found


def build_all(spec, outdir):
    """Build the artifacts, separating DIAGNOSIS from REBUILD.

    The refusal in `references/pass-artifacts.md` is scoped precisely: "Payment/rate-card
    reconciliation in the REBUILT tables actually foots ... do not ship an unreconciled
    REWRITE." It forbids shipping a rebuilt commercial artifact that still does not foot.
    It does NOT forbid reporting the defect.

    So:
      * the DIAGNOSIS (`scope_findings.json`, `raci_matrix.csv`) is always written, and a
        failing rate card or payment schedule is recorded IN it as a finding;
      * the REBUILT commercial workbook is withheld when the arithmetic fails, because that
        is the artifact the rule is about.

    An earlier version of this function refused everything on any arithmetic failure, which
    suppressed the very sidecar that documents the defect. That was backwards for a
    diagnostic skill and is the reason this split is spelled out here.

    Score, severity-ceiling and orphan errors still refuse outright: those mean the ledger
    itself is inconsistent, so there is no trustworthy diagnosis to write.
    """
    findings = list(spec.get("findings") or [])
    gen_findings = _arithmetic_findings(spec)
    findings.extend(gen_findings)

    score = compute_scope_score(spec["dimension_scores"],
                                spec.get("findings") or [],
                                auto_clamp_findings=gen_findings)
    raci_rows, orphans = build_raci(spec.get("deliverables") or [], findings)

    os.makedirs(outdir, exist_ok=True)
    written = {}
    written["raci_matrix.csv"] = write_raci_csv(
        raci_rows, os.path.join(outdir, "raci_matrix.csv"))
    written["scope_findings.json"] = write_scope_findings_json(
        findings, score, os.path.join(outdir, "scope_findings.json"))

    withheld = []
    if _OPENPYXL:
        written["change_control_log_template.xlsx"] = write_change_control_xlsx(
            os.path.join(outdir, "change_control_log_template.xlsx"),
            spec.get("change_control_triggers"), spec.get("change_control_drafted", False))
        if gen_findings:
            withheld.append({
                "artifact": "rate_card_and_payment_schedule.xlsx",
                "reason": ("the arithmetic does not foot, so a rebuilt commercial table "
                           "would carry the same defect it was meant to fix. The failure "
                           "is recorded in scope_findings.json as %s."
                           % ", ".join(f["id"] for f in gen_findings)),
            })
        else:
            p, _ = write_rate_card_xlsx(
                spec["rate_card"], spec["milestones"], spec["contract_value"],
                os.path.join(outdir, "rate_card_and_payment_schedule.xlsx"))
            written["rate_card_and_payment_schedule.xlsx"] = p

    return {"score": score, "orphans": orphans, "written": written,
            "withheld": withheld, "generator_findings": gen_findings}


def main(argv):
    if not argv:
        print(__doc__.strip())
        print("\nusage: scope_artifacts_generator.py <spec.json> <outdir>")
        return 0
    with open(argv[0], encoding="utf-8") as fh:
        spec = json.load(fh)
    outdir = argv[1] if len(argv) > 1 else "."
    try:
        result = build_all(spec, outdir)
    except (ScopeArtifactError, ReconciliationError) as e:
        print("REFUSED: %s: %s" % (type(e).__name__, e), file=sys.stderr)
        return 2
    print(json.dumps({"score": result["score"]["score_0to100"],
                      "band": result["score"]["band"],
                      "orphans": result["orphans"],
                      "written": sorted(result["written"])}, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
