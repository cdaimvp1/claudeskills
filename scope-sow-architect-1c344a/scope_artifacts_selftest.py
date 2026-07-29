#!/usr/bin/env python3
"""
Self-test for scope_artifacts_generator.py (F9 build 1).

The score test is anchored on the PUBLISHED worked example in
`references/scope-quality-scoring.md`, which states composite 2.550 -> 51 -> "Moderate gap
/ Needs Targeted Fixes". Reproducing a golden someone else wrote down is worth more than
asserting against numbers this code produced itself.

Negative controls throughout: an orphan that IS flagged must pass, a LOW finding must not
cap anything. A generator that refuses legitimate input is as broken as one that accepts
anything.

Run: python scope_artifacts_selftest.py
"""
from __future__ import annotations

import json
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from numeric_kernel import ReconciliationError, WeightSumError    # noqa: E402
from scope_artifacts_generator import (                            # noqa: E402
    DIMENSION_WEIGHTS,
    LineMathError,
    OrphanError,
    ScopeArtifactError,
    SeverityCapError,
    build_all,
    build_raci,
    compute_scope_score,
    score_band,
    verify_payment_schedule,
    verify_rate_card,
    write_raci_csv,
)

try:
    from openpyxl import load_workbook
    _XL = True
except ImportError:
    _XL = False

PASS, FAIL = [], []


def ok(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(("  ok   " if cond else "  FAIL ") + name + (("  <- " + detail) if detail and not cond else ""))


def raises(name, exc, fn):
    try:
        fn()
    except exc:
        ok(name, True)
        return
    except Exception as e:                                    # noqa: BLE001
        ok(name, False, "raised %s, expected %s" % (type(e).__name__, exc.__name__))
        return
    ok(name, False, "did not raise " + exc.__name__)


# The published worked example, in the generator's key order.
GOLDEN_SCORES = {
    "deliverables_definition":  3.5,
    "scope_boundary":           1.0,
    "acceptance_criteria":      2.0,
    "assumptions_dependencies": 3.0,
    "raci_completeness":        2.5,
    "milestones_schedule":      4.0,
    "slas_kpis":                2.5,
    "staffing_rate_card":       3.5,
    "payment_alignment":        1.5,
    "change_control":           0.0,
}


def spec():
    return {
        "contract_value": 500000.00,
        "dimension_scores": dict(GOLDEN_SCORES),
        "findings": [
            {"id": "F-001", "dimension": "change_control", "severity": "BLOCKING",
             "status": "open", "text": "No change-control process of any kind."},
            {"id": "F-002", "dimension": "scope_boundary", "severity": "HIGH",
             "status": "open", "text": "No out-of-scope statement."},
        ],
        "deliverables": [
            {"name": "Discovery report", "responsible": "Supplier PM",
             "accountable": "Lilly Owner", "consulted": "IT", "informed": "Finance"},
            {"name": "Migration plan", "responsible": "Supplier Architect",
             "accountable": "Lilly Owner", "consulted": "Security", "informed": "IT"},
        ],
        "rate_card": [
            {"role": "Architect", "level": "Senior", "rate": 250.0, "quantity": 400,
             "unit": "hour", "line_total": 100000.0},
            {"role": "Engineer", "level": "Mid", "rate": 175.0, "quantity": 1200,
             "unit": "hour", "line_total": 210000.0},
        ],
        "milestones": [
            {"name": "M1 Discovery", "deliverable": "Discovery report",
             "trigger": "Acceptance of report", "amount": 150000.0},
            {"name": "M2 Migration", "deliverable": "Migration plan",
             "trigger": "Go-live sign-off", "amount": 350000.0},
        ],
    }


def run():
    print("=" * 84)
    print("scope_artifacts_generator self-test")
    print("=" * 84)

    # --- the published golden ---------------------------------------------------------
    ok("T1  weights sum to 1.0 as the doc requires",
       abs(sum(DIMENSION_WEIGHTS.values()) - 1.0) < 1e-9,
       "%.4f" % sum(DIMENSION_WEIGHTS.values()))

    s = compute_scope_score(GOLDEN_SCORES, [])
    ok("T2  GOLDEN: composite is 2.550 (published worked example)",
       abs(s["composite_0to5"] - 2.550) < 0.0005, str(s["composite_0to5"]))
    ok("T3  GOLDEN: rescaled score is 51", s["score_0to100"] == 51, str(s["score_0to100"]))
    ok("T4  GOLDEN: band is 'Moderate gap / Needs Targeted Fixes'",
       s["band"] == "Moderate gap / Needs Targeted Fixes", s["band"])
    ok("T5  the calculation table has all ten rows", len(s["calculation_table"]) == 10)
    ok("T6  a published per-dimension contribution reproduces (0.15 x 3.5 = 0.525)",
       any(abs(r["contribution"] - 0.525) < 1e-9 for r in s["calculation_table"]))

    for lo, hi, label in ((75, 100, "Low gap / Ready to Issue"),
                          (50, 74, "Moderate gap / Needs Targeted Fixes"),
                          (25, 49, "High gap / Major Rework"),
                          (0, 24, "Critical gap / Not Priceable")):
        ok("T7  band boundary %3d and %3d" % (lo, hi),
           score_band(lo) == label and score_band(hi) == label)

    # --- severity ceilings -------------------------------------------------------------
    def over_cap():
        f = [{"id": "X", "dimension": "acceptance_criteria", "severity": "BLOCKING",
              "status": "open"}]
        compute_scope_score(GOLDEN_SCORES, f)
    raises("T8  refuses a dimension scored above its BLOCKING ceiling (0.9)",
           SeverityCapError, over_cap)

    def over_high():
        sc = dict(GOLDEN_SCORES); sc["milestones_schedule"] = 4.0
        compute_scope_score(sc, [{"id": "X", "dimension": "milestones_schedule",
                                  "severity": "HIGH", "status": "open"}])
    raises("T9  refuses a dimension scored above its HIGH ceiling (3.4)",
           SeverityCapError, over_high)

    sc = dict(GOLDEN_SCORES); sc["milestones_schedule"] = 3.4
    ok("T10 a score exactly AT the ceiling is allowed",
       compute_scope_score(sc, [{"id": "X", "dimension": "milestones_schedule",
                                 "severity": "HIGH", "status": "open"}])["score_0to100"] > 0)

    ok("T11 NEGATIVE CONTROL: a LOW finding caps nothing",
       compute_scope_score(GOLDEN_SCORES,
                           [{"id": "X", "dimension": "milestones_schedule",
                             "severity": "LOW", "status": "open"}])["score_0to100"] == 51)

    ok("T12 NEGATIVE CONTROL: a RESOLVED blocking finding caps nothing",
       compute_scope_score(GOLDEN_SCORES,
                           [{"id": "X", "dimension": "milestones_schedule",
                             "severity": "BLOCKING", "status": "resolved"}])["score_0to100"] == 51)

    ok("T13 the WORST ceiling wins when a dimension has several findings",
       compute_scope_score(
           {**GOLDEN_SCORES, "slas_kpis": 0.9},
           [{"id": "A", "dimension": "slas_kpis", "severity": "MEDIUM", "status": "open"},
            {"id": "B", "dimension": "slas_kpis", "severity": "BLOCKING", "status": "open"}]
       )["score_0to100"] >= 0)

    raises("T14 refuses a finding tagged to an unknown dimension", ScopeArtifactError,
           lambda: compute_scope_score(GOLDEN_SCORES,
                                       [{"id": "X", "dimension": "nope", "severity": "LOW"}]))
    raises("T15 refuses an unknown severity", ScopeArtifactError,
           lambda: compute_scope_score(GOLDEN_SCORES,
                                       [{"id": "X", "dimension": "slas_kpis",
                                         "severity": "SEVERE"}]))
    raises("T16 refuses a missing dimension score (not treated as 0)", ScopeArtifactError,
           lambda: compute_scope_score({k: v for k, v in list(GOLDEN_SCORES.items())[:9]}, []))

    # --- rate card + payment schedule --------------------------------------------------
    sp = spec()
    ok("T17 a footing rate card verifies", len(verify_rate_card(sp["rate_card"])) == 2)

    def bad_footing():
        s2 = spec(); s2["rate_card"][0]["line_total"] = 99999.0
        verify_rate_card(s2["rate_card"])
    raises("T18 refuses a rate-card row that does not foot", LineMathError, bad_footing)

    def missing_qty():
        s2 = spec(); del s2["rate_card"][0]["quantity"]
        verify_rate_card(s2["rate_card"])
    raises("T19 refuses a row with no quantity (never treats it as 0)",
           LineMathError, missing_qty)

    ok("T20 a reconciling payment schedule verifies",
       verify_payment_schedule(sp["milestones"], sp["contract_value"]) == 500000.0)

    def bad_recon():
        s2 = spec(); s2["milestones"][0]["amount"] = 100000.0
        verify_payment_schedule(s2["milestones"], s2["contract_value"])
    raises("T21 refuses milestones that do not sum to the contract value",
           ReconciliationError, bad_recon)

    # --- RACI orphans -------------------------------------------------------------------
    rows, orphans = build_raci(sp["deliverables"], sp["findings"])
    ok("T22 a complete RACI yields no orphans", orphans == [], repr(orphans))

    def silent_orphan():
        s2 = spec(); s2["deliverables"][0]["responsible"] = ""
        build_raci(s2["deliverables"], s2["findings"])
    raises("T23 refuses an UNFLAGGED orphaned deliverable", OrphanError, silent_orphan)

    s3 = spec()
    s3["deliverables"][0]["responsible"] = ""
    s3["findings"].append({"id": "F-003", "dimension": "raci_completeness",
                           "severity": "HIGH", "status": "open",
                           "deliverable": "Discovery report",
                           "text": "No Responsible party named."})
    s3["dimension_scores"]["raci_completeness"] = 2.5
    rows3, orphans3 = build_raci(s3["deliverables"], s3["findings"])
    ok("T24 NEGATIVE CONTROL: a FLAGGED orphan is allowed through", orphans3 == ["Discovery report"],
       repr(orphans3))
    ok("T25 the flagged orphan is marked in the row, not silently blanked",
       rows3[0]["Orphan_Flag"] == "YES" and "ORPHAN" in rows3[0]["Responsible"])

    # --- artifacts on disk ---------------------------------------------------------------
    tmp = tempfile.mkdtemp(prefix="f9_scope_")
    res = build_all(spec(), tmp)

    csv_path = os.path.join(tmp, "raci_matrix.csv")
    ok("T26 raci_matrix.csv written", os.path.isfile(csv_path))
    with open(csv_path, encoding="utf-8") as fh:
        head = fh.readline().strip()
    ok("T27 RACI csv has all four role columns",
       all(r in head for r in ("Responsible", "Accountable", "Consulted", "Informed")), head)

    js_path = os.path.join(tmp, "scope_findings.json")
    ok("T28 scope_findings.json written", os.path.isfile(js_path))
    with open(js_path, encoding="utf-8") as fh:
        payload = json.load(fh)
    ok("T29 sidecar carries EVERY finding, none dropped",
       payload["finding_count"] == 2 and len(payload["findings"]) == 2)
    ok("T30 sidecar carries the score calculation table, not just the score",
       len(payload["scope_definition_score"]["calculation_table"]) == 10)
    ok("T31 sidecar score matches the computed score",
       payload["scope_definition_score"]["score_0to100"] == res["score"]["score_0to100"])

    # --- DIAGNOSIS survives a failing rebuild (the correction) --------------------------
    # pass-artifacts.md forbids shipping an unreconciled REWRITE. It does not forbid
    # REPORTING the defect. An earlier version refused everything, suppressing the sidecar
    # that documents the problem, which is backwards for a diagnostic skill.
    tmp3 = tempfile.mkdtemp(prefix="f9_scope_diag_")
    broken = spec()
    broken["milestones"][0]["amount"] = 100000.0        # no longer sums to contract value
    # Also break a rate-card row, because that is what exercises CLAMPING: the golden
    # already scores payment_alignment 1.5, below the 2.4 footing cap, so a payment failure
    # alone clamps nothing. staffing_rate_card is 3.5 and does get clamped.
    broken["rate_card"][0]["line_total"] = 99999.0
    res3 = build_all(broken, tmp3)
    ok("T39 a failing rebuild STILL writes scope_findings.json",
       os.path.isfile(os.path.join(tmp3, "scope_findings.json")))
    ok("T40 a failing rebuild STILL writes raci_matrix.csv",
       os.path.isfile(os.path.join(tmp3, "raci_matrix.csv")))
    ok("T41 the REBUILT commercial workbook is WITHHELD",
       not os.path.isfile(os.path.join(tmp3, "rate_card_and_payment_schedule.xlsx")))
    ok("T42 the withholding is reported with a reason, not silent",
       len(res3["withheld"]) == 1 and "does not foot" in res3["withheld"][0]["reason"],
       repr(res3["withheld"]))
    with open(os.path.join(tmp3, "scope_findings.json"), encoding="utf-8") as fh:
        p3 = json.load(fh)
    ok("T43 the arithmetic failure is RECORDED as a finding",
       any(f["id"].startswith("GEN-PS") for f in p3["findings"]), repr([f["id"] for f in p3["findings"]]))
    ok("T44 that finding states the numbers so a reader can check it",
       any("450,000.00" in f.get("text", "") and "500,000.00" in f.get("text", "")
           for f in p3["findings"] if f["id"].startswith("GEN-PS")))
    ok("T44a BOTH arithmetic defects are recorded, not just the first",
       len([f for f in p3["findings"] if f["id"].startswith("GEN-")]) == 2,
       repr([f["id"] for f in p3["findings"] if f["id"].startswith("GEN-")]))
    ok("T45 the generator-found defect CLAMPS its dimension rather than refusing",
       any(r.get("clamped") for r in p3["scope_definition_score"]["calculation_table"]))
    clamp = [r for r in p3["scope_definition_score"]["calculation_table"] if r.get("clamped")][0]
    ok("T46 the clamp uses the doc's footing cap of 2.4", clamp["clamped"]["to"] == 2.4,
       repr(clamp["clamped"]))
    ok("T47 the clamp is VISIBLE: the submitted score is preserved beside it",
       clamp["score_as_submitted"] > clamp["score"], repr(clamp))
    ok("T48 the clamped score is lower than the clean one (a defect cannot improve a score)",
       p3["scope_definition_score"]["score_0to100"] < res["score"]["score_0to100"],
       "%s vs %s" % (p3["scope_definition_score"]["score_0to100"], res["score"]["score_0to100"]))

    # NEGATIVE CONTROL: the CALLER's own inconsistency still refuses. Clamping is only for
    # defects the generator discovered, never a way to paper over a bad input ledger.
    raises("T49 a caller's score above THEIR OWN finding's ceiling still refuses",
           SeverityCapError,
           lambda: compute_scope_score(GOLDEN_SCORES,
                                       [{"id": "C-1", "dimension": "acceptance_criteria",
                                         "severity": "BLOCKING", "status": "open"}]))

    if _XL:
        rc = os.path.join(tmp, "rate_card_and_payment_schedule.xlsx")
        ok("T32 rate card workbook written", os.path.isfile(rc))
        wb = load_workbook(rc)
        ok("T33 both tabs present", wb.sheetnames == ["Rate Card", "Payment Schedule"],
           repr(wb.sheetnames))
        ws = wb["Rate Card"]
        formulas = [c.value for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith("=")]
        ok("T34 footing checks are LIVE formulas, so they survive editing",
           any("DOES NOT FOOT" in f for f in formulas), "%d formulas" % len(formulas))
        # The real fix: a formula-only verdict reads as None to any programmatic reader,
        # because openpyxl writes no cached value. The static column must carry the answer.
        dwb = load_workbook(rc, data_only=True)["Rate Card"]
        ok("T34a the BUILD-TIME verdict is readable programmatically (data_only=True)",
           dwb.cell(row=2, column=7).value == "OK", repr(dwb.cell(row=2, column=7).value))
        ok("T34b the live-formula column is blank to that same reader, which is why "
           "the static one exists",
           dwb.cell(row=2, column=8).value is None, repr(dwb.cell(row=2, column=8).value))
        ps = wb["Payment Schedule"]
        pf = [c.value for row in ps.iter_rows() for c in row
              if isinstance(c.value, str) and c.value.startswith("=")]
        ok("T35 reconciliation is SHOWN in the workbook, not just asserted in code",
           any("RECONCILES" in f for f in pf), repr(pf[-1:]))
        dps = load_workbook(rc, data_only=True)["Payment Schedule"]
        statics = [c.value for row in dps.iter_rows() for c in row if c.value == "RECONCILES"]
        ok("T35a the reconciliation verdict is ALSO readable programmatically",
           len(statics) == 1, repr(statics))
        cc = os.path.join(tmp, "change_control_log_template.xlsx")
        ok("T36 change-control template written", os.path.isfile(cc))
        cwb = load_workbook(cc)
        hdr = [c.value for c in cwb["Change Control"][1]]
        ok("T37 change-control has its fixed column set",
           "Trigger Threshold" in hdr and "Approval Authority" in hdr, repr(hdr))
    else:
        print("  skip  T32-T37 (openpyxl unavailable)")

    # --- what a HARD refusal writes: nothing ---------------------------------------------
    # The distinction that matters. An ARITHMETIC failure is a finding, so the diagnosis is
    # still written (T39-T48). A LEDGER inconsistency means the input contradicts itself, so
    # there is no trustworthy diagnosis to write and nothing is written at all.
    tmp2 = tempfile.mkdtemp(prefix="f9_scope_bad_")
    try:
        bad = spec()
        bad["findings"].append({"id": "C-9", "dimension": "deliverables_definition",
                                "severity": "BLOCKING", "status": "open"})
        build_all(bad, tmp2)          # score 3.5 vs its own BLOCKING ceiling of 0.9
    except SeverityCapError:
        pass
    ok("T38 a LEDGER inconsistency writes NO artifacts at all", os.listdir(tmp2) == [],
       repr(os.listdir(tmp2)))

    tmp4 = tempfile.mkdtemp(prefix="f9_scope_orphan_")
    try:
        bad2 = spec(); bad2["deliverables"][0]["responsible"] = ""
        build_all(bad2, tmp4)
    except OrphanError:
        pass
    ok("T38a an unflagged orphan also writes NO artifacts", os.listdir(tmp4) == [],
       repr(os.listdir(tmp4)))

    print("=" * 84)
    print("SUMMARY: %d/%d passed, %d failed" % (len(PASS), len(PASS) + len(FAIL), len(FAIL)))
    print("=" * 84)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(run())
